"""
项目全生命周期全景管控系统（CC-010）— 数据模型、业务计算与报表导出

依据《项目全生命周期全景管控系统-产品架构规格文档（V1.0 可迭代版）》实现。
顶层纲领：四算为纲、财经为尺、PMO 为缰、人力为本。

表清单（统一 plm_ 前缀，与既有 contract_compare.db 同库）：
  plm_opportunity      商机档案（模块一）
  plm_presale_doc      售前资料归档（模块一）
  plm_contract         合同主数据（模块二）
  plm_project          项目立项（模块二）
  plm_baseline         四算基线：概算/预算/核算[预留]/决算[预留]（模块二/三）
  plm_baseline_item    基线分项明细
  plm_milestone        里程碑（粗/细两级，parent_id 自关联）（模块三）
  plm_task             任务拆解（模块三）
  plm_staff            人员池（模块四）
  plm_assignment       人员精准分配（模块四）
  plm_timesheet        工时填报（模块四）
  plm_ledger           收支台账（收入/成本，含工时自动归集）（模块五）
  plm_alert_rule       预警规则配置（模块七）
  plm_alert            预警实例与闭环（模块七）
  plm_op_log           操作日志（模块九）
  plm_dict             全局字典（模块九）
  plm_config           全局参数（含四算基线管控开关）（模块九）

约束：
  - 本模块数据全部页面手工录入，不从 datasource 上传表取数。
  - 核算 / 决算仅占位存储，本期不做计算与校验（规格 FR-3）。
  - 引用本模块必须使用 `import plm_models as plm`，禁止符号导入污染 main.py 全局命名空间。
"""

import os
import io
import json
import sqlite3
from datetime import datetime, date

# 四算「读本体」：优先从 ontos 取四算类型枚举（唯一真源），失败回退本地常量（保持一致）
try:
    import sys as _sys
    import os as _os
    _ontos_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', 'ontos')
    if _ontos_dir not in _sys.path:
        _sys.path.insert(0, _ontos_dir)
    from ontos.domain_business import (  # noqa: F401
        COST_BASELINE_CALC_TYPES, COST_BASELINE_CALC_TYPE_CN,
        COST_BASELINE_CALCULATED as ONTOS_CALC_BASELINE)
except Exception:  # pragma: no cover
    COST_BASELINE_CALC_TYPES = ("概算", "基准预算", "生产预算", "核算", "决算")
    COST_BASELINE_CALC_TYPE_CN = {k: k for k in COST_BASELINE_CALC_TYPES}

# calc_type(本体) → stage(历史兼容) 映射：新写入以 calc_type 为准，stage 同步填充以兼容旧 compare/lock 逻辑
CALC_TYPE_TO_STAGE = {
    "概算": "estimate_locked",
    "基准预算": "budget",
    "生产预算": "budget",
    "核算": "accounting",
    "决算": "final",
}
STAGE_TO_CALC_TYPE = {v: k for k, v in CALC_TYPE_TO_STAGE.items()}

DB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
DB_PATH = os.path.join(DB_DIR, 'contract_compare.db')


# ===================== 基础工具 =====================

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _today():
    return date.today().strftime('%Y-%m-%d')


def _r(v, n=2):
    """金额安全四舍五入；非数值返回 None。"""
    if v is None:
        return None
    try:
        return round(float(v), n)
    except (TypeError, ValueError):
        return None


def _f(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _rate(numerator, denominator):
    """比率（小数）；分母 <= 0 或任一为空返回 None。"""
    if numerator is None or denominator is None:
        return None
    try:
        d = float(denominator)
    except (TypeError, ValueError):
        return None
    if d <= 0:
        return None
    return round(float(numerator) / d, 6)


def _days_between(d1, d2):
    """d2 - d1 的天数；任一为空或格式异常返回 None。"""
    if not d1 or not d2:
        return None
    try:
        a = datetime.strptime(str(d1)[:10], '%Y-%m-%d')
        b = datetime.strptime(str(d2)[:10], '%Y-%m-%d')
    except (ValueError, TypeError):
        return None
    return (b - a).days


def _rows(cur):
    return [dict(r) for r in cur.fetchall()]


def _first(cur):
    rs = cur.fetchall()
    return dict(rs[0]) if rs else None


# ===================== 枚举与常量 =====================

OPP_STATUS = ('跟进中', '投标中', '中标', '流标')
PROJECT_STATUS = ('待启动', '执行中', '暂停', '结项')
CONTRACT_STATUS = ('草签', '已签署', '执行中', '已结项', '已终止')
BASELINE_STAGE = ('estimate_bid', 'estimate_locked', 'budget', 'accounting', 'final')
BASELINE_STAGE_CN = {
    'estimate_bid': '概算（投标）', 'estimate_locked': '概算（立项锁定）',
    'budget': '预算', 'accounting': '核算（预留）', 'final': '决算（预留）',
}
# 本期仅前三项参与计算；核算/决算为预留（FR-3）
BASELINE_CALCULATED = ('estimate_bid', 'estimate_locked', 'budget')
BASELINE_RESERVED = ('accounting', 'final')
BASELINE_STATUS = ('草稿', '已确认', '已锁定')
COST_CATEGORY = ('人力成本', '分包成本', '硬件成本', '软件成本', '服务成本', '其他费用')
BUDGET_CATEGORY = ('人力成本', '其他费用')
INCOME_CATEGORY = ('签单收入', '变更收入', '其他收入')
MILESTONE_STATUS = ('未开始', '进行中', '已完成', '延期')
MILESTONE_LEVEL = ('粗', '细')
TASK_STATUS = ('未开始', '进行中', '已完成', '延期', '已取消')
STAFF_STATUS = ('可用', '占用', '休假', '离职')
ASSIGN_STATUS = ('生效中', '已解除')
ALERT_DIM = ('cost', 'gross', 'schedule', 'staff')
ALERT_LEVEL = ('提醒', '警告', '严重')
ALERT_STATUS = ('待处理', '处理中', '已闭环')
BASELINE_ORDER = ('estimate_locked', 'budget', 'accounting', 'final')
RESERVED_NOTE = '预留：本期不做计算与校验'


# ===================== 建表 =====================

def init_plm_db():
    """初始化 PLM 全部表（幂等）。"""
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_opportunity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opp_no TEXT UNIQUE NOT NULL,
            opp_name TEXT NOT NULL,
            customer TEXT DEFAULT '',
            industry TEXT DEFAULT '',
            region TEXT DEFAULT '',
            dept TEXT DEFAULT '',
            owner TEXT DEFAULT '',
            status TEXT DEFAULT '跟进中',
            bid_date TEXT DEFAULT '',
            expect_income REAL DEFAULT 0,
            est_cost REAL DEFAULT 0,
            est_gross REAL DEFAULT 0,
            est_gross_rate REAL,
            follow_log TEXT DEFAULT '[]',
            won_at TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_presale_doc (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opportunity_id INTEGER NOT NULL,
            doc_name TEXT NOT NULL,
            doc_type TEXT DEFAULT '其他',
            file_name TEXT DEFAULT '',
            file_path TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            uploader TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_contract (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_no TEXT UNIQUE NOT NULL,
            contract_name TEXT DEFAULT '',
            customer TEXT DEFAULT '',
            industry TEXT DEFAULT '',
            region TEXT DEFAULT '',
            dept TEXT DEFAULT '',
            sign_amount REAL DEFAULT 0,
            sign_date TEXT DEFAULT '',
            project_cycle TEXT DEFAULT '',
            status TEXT DEFAULT '已签署',
            owner TEXT DEFAULT '',
            opportunity_id INTEGER,
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_project (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_no TEXT UNIQUE NOT NULL,
            project_name TEXT NOT NULL,
            customer TEXT DEFAULT '',
            dept TEXT DEFAULT '',
            region TEXT DEFAULT '',
            manager TEXT DEFAULT '',
            status TEXT DEFAULT '待启动',
            contract_id INTEGER,
            opportunity_id INTEGER,
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            kickoff_date TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_baseline (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scope_type TEXT NOT NULL,
            scope_id INTEGER NOT NULL,
            stage TEXT NOT NULL,
            total_income REAL DEFAULT 0,
            total_cost REAL DEFAULT 0,
            gross REAL DEFAULT 0,
            gross_rate REAL,
            status TEXT DEFAULT '草稿',
            source_baseline_id INTEGER,
            version INTEGER DEFAULT 1,
            locked_at TEXT DEFAULT '',
            locked_by TEXT DEFAULT '',
            created_by TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # 四算重构（读本体）：补齐 calc_type(对齐 ontos CostBaseline.calc_type) 与 scope_key(合同号字符串归集锚)
    # 注：SQLite 不支持 ADD COLUMN IF NOT EXISTS，先判列存在
    _bl_cols = {r[1] for r in c.execute("PRAGMA table_info(plm_baseline)")}
    if 'calc_type' not in _bl_cols:
        c.execute("ALTER TABLE plm_baseline ADD COLUMN calc_type TEXT DEFAULT ''")
    if 'scope_key' not in _bl_cols:
        c.execute("ALTER TABLE plm_baseline ADD COLUMN scope_key TEXT DEFAULT ''")

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_baseline_item (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            baseline_id INTEGER NOT NULL,
            category TEXT DEFAULT '其他费用',
            item_name TEXT DEFAULT '',
            plan_amount REAL DEFAULT 0,
            actual_amount REAL,
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_milestone (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            parent_id INTEGER,
            level TEXT DEFAULT '细',
            name TEXT NOT NULL,
            owner TEXT DEFAULT '',
            plan_start TEXT DEFAULT '',
            plan_end TEXT DEFAULT '',
            actual_start TEXT DEFAULT '',
            actual_end TEXT DEFAULT '',
            progress REAL DEFAULT 0,
            status TEXT DEFAULT '未开始',
            is_key INTEGER DEFAULT 0,
            plan_output REAL DEFAULT 0,
            task_no TEXT DEFAULT '',
            plan_payback_date TEXT DEFAULT '',
            payback_date TEXT DEFAULT '',
            payback_amount REAL DEFAULT 0,
            deliverable TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_task (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            milestone_id INTEGER,
            name TEXT NOT NULL,
            owner TEXT DEFAULT '',
            plan_hours REAL DEFAULT 0,
            actual_hours REAL DEFAULT 0,
            progress REAL DEFAULT 0,
            status TEXT DEFAULT '未开始',
            plan_start TEXT DEFAULT '',
            plan_end TEXT DEFAULT '',
            actual_end TEXT DEFAULT '',
            deliverable TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_staff (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            role TEXT DEFAULT '',
            dept TEXT DEFAULT '',
            cost_rate REAL DEFAULT 0,
            available_hours REAL DEFAULT 160,
            status TEXT DEFAULT '可用',
            skills TEXT DEFAULT '',
            efficiency_bonus REAL,
            revenue_per_day REAL,
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_assignment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            milestone_id INTEGER,
            task_id INTEGER,
            role_in_proj TEXT DEFAULT '',
            planned_hours REAL DEFAULT 0,
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            status TEXT DEFAULT '生效中',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_timesheet (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            task_id INTEGER,
            work_date TEXT DEFAULT '',
            hours REAL DEFAULT 0,
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            contract_id INTEGER,
            kind TEXT NOT NULL,
            category TEXT DEFAULT '',
            plan_or_actual TEXT DEFAULT '实际',
            amount REAL DEFAULT 0,
            occur_date TEXT DEFAULT '',
            source TEXT DEFAULT '手工录入',
            milestone_id INTEGER,
            ref_type TEXT DEFAULT '',
            ref_id TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_alert_rule (
            rule_key TEXT PRIMARY KEY,
            rule_name TEXT NOT NULL,
            dim TEXT NOT NULL,
            metric TEXT NOT NULL,
            op TEXT DEFAULT '>',
            threshold REAL NOT NULL,
            level TEXT DEFAULT '提醒',
            enabled INTEGER DEFAULT 1,
            description TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_alert (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            staff_id INTEGER,
            rule_key TEXT NOT NULL,
            dim TEXT DEFAULT '',
            level TEXT DEFAULT '提醒',
            title TEXT DEFAULT '',
            detail TEXT DEFAULT '',
            metric_value REAL,
            threshold REAL,
            status TEXT DEFAULT '待处理',
            handler TEXT DEFAULT '',
            handle_note TEXT DEFAULT '',
            handle_time TEXT DEFAULT '',
            last_scan_at TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_op_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            target_type TEXT NOT NULL,
            target_id TEXT DEFAULT '',
            target_name TEXT DEFAULT '',
            action TEXT NOT NULL,
            change_json TEXT DEFAULT '{}',
            operator TEXT DEFAULT '',
            role TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_dict (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            key TEXT NOT NULL,
            label TEXT NOT NULL,
            sort INTEGER DEFAULT 0,
            enabled INTEGER DEFAULT 1,
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS plm_config (
            key TEXT PRIMARY KEY,
            value TEXT DEFAULT '',
            description TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # 字典同分类下 key 唯一（create_dict 依赖 IntegrityError 做冲突提示）
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS ux_plm_dict_cat_key "
              "ON plm_dict (category, key)")
    # 常用查询索引
    c.execute("CREATE INDEX IF NOT EXISTS ix_plm_baseline_scope "
              "ON plm_baseline (scope_type, scope_id, stage)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_plm_ledger_project "
              "ON plm_ledger (project_id, kind, plan_or_actual)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_plm_alert_open "
              "ON plm_alert (project_id, rule_key, status)")

    # —— plm_milestone 增量补列（幂等：列已在则跳过）——
    # 承载导入的「项目里程碑表」回款信息，供 payment_cycle / milestone_payback_point 使用
    _milestone_add_cols = [
        ('task_no', 'TEXT DEFAULT ""'),
        ('plan_payback_date', 'TEXT DEFAULT ""'),
        ('payback_date', 'TEXT DEFAULT ""'),
        ('payback_amount', 'REAL DEFAULT 0'),
    ]
    _ms_existing = {r[1] for r in c.execute("PRAGMA table_info(plm_milestone)").fetchall()}
    for _name, _typedef in _milestone_add_cols:
        if _name in _ms_existing:
            continue
        c.execute('ALTER TABLE plm_milestone ADD COLUMN %s %s' % (_name, _typedef))

    conn.commit()
    conn.close()


# ===================== 预置字典 / 规则 / 参数 =====================

DEFAULT_DICT = [
    ('opp_status', '跟进中', '跟进中', 1), ('opp_status', '投标中', '投标中', 2),
    ('opp_status', '中标', '中标', 3), ('opp_status', '流标', '流标', 4),
    ('project_status', '待启动', '待启动', 1), ('project_status', '执行中', '执行中', 2),
    ('project_status', '暂停', '暂停', 3), ('project_status', '结项', '结项', 4),
    ('contract_status', '草签', '草签', 1), ('contract_status', '已签署', '已签署', 2),
    ('contract_status', '执行中', '执行中', 3), ('contract_status', '已结项', '已结项', 4),
    ('contract_status', '已终止', '已终止', 5),
    ('milestone_type', '粗', '粗里程碑（立项级）', 1),
    ('milestone_type', '细', '细里程碑（执行级）', 2),
    ('milestone_status', '未开始', '未开始', 1), ('milestone_status', '进行中', '进行中', 2),
    ('milestone_status', '已完成', '已完成', 3), ('milestone_status', '延期', '延期', 4),
    ('task_status', '未开始', '未开始', 1), ('task_status', '进行中', '进行中', 2),
    ('task_status', '已完成', '已完成', 3), ('task_status', '延期', '延期', 4),
    ('task_status', '已取消', '已取消', 5),
    ('cost_category', '人力成本', '人力成本', 1), ('cost_category', '分包成本', '分包成本', 2),
    ('cost_category', '硬件成本', '硬件成本', 3), ('cost_category', '软件成本', '软件成本', 4),
    ('cost_category', '服务成本', '服务成本', 5), ('cost_category', '其他费用', '其他费用', 6),
    ('income_category', '签单收入', '签单收入', 1), ('income_category', '变更收入', '变更收入', 2),
    ('income_category', '其他收入', '其他收入', 3),
    ('budget_category', '人力成本', '人力成本', 1), ('budget_category', '其他费用', '其他费用', 2),
    ('staff_status', '可用', '可用', 1), ('staff_status', '占用', '占用', 2),
    ('staff_status', '休假', '休假', 3), ('staff_status', '离职', '离职', 4),
    ('presale_doc_type', '投标方案', '投标方案', 1), ('presale_doc_type', '报价单', '报价单', 2),
    ('presale_doc_type', '沟通记录', '沟通记录', 3), ('presale_doc_type', '其他', '其他', 4),
    ('role', '项目经理', '项目经理', 1), ('role', 'PMO', 'PMO', 2), ('role', '财务', '财务', 3),
    ('role', '管理员', '管理员', 4), ('role', '普通查看者', '普通查看者', 5),
    ('role', '开发工程师', '开发工程师', 6), ('role', '实施工程师', '实施工程师', 7),
    ('role', '测试工程师', '测试工程师', 8), ('role', '架构师', '架构师', 9),
]

DEFAULT_CONFIG = [
    ('baseline_constraint', 'off',
     '四算基线逐级约束开关：on 时预算超概算将被拒绝保存（本期默认 off）'),
    ('labor_day_hours', '8', '工时折算人天的口径（小时/人天）'),
    ('alert_staff_overload', '1.2', '人员负荷率过载判定系数兜底值'),
]

DEFAULT_ALERT_RULES = [
    ('cost_overrun_warn', '预算超耗预警', 'cost', 'budget_usage_rate', '>=', 0.80, '提醒',
     '预算消耗占比达到 80%，需关注成本节奏'),
    ('cost_overrun_crit', '预算超支预警', 'cost', 'budget_usage_rate', '>=', 1.00, '严重',
     '累计实际成本已超出执行预算总额'),
    ('gross_low_warn', '毛利偏低预警', 'gross', 'actual_gross_rate', '<', 0.15, '提醒',
     '实际毛利率低于 15% 底线'),
    ('gross_negative_crit', '项目亏损预警', 'gross', 'actual_gross_rate', '<', 0.0, '严重',
     '实际毛利为负，项目处于亏损状态'),
    ('schedule_overdue_warn', '进度延期预警', 'schedule', 'max_overdue_days', '>', 7, '提醒',
     '存在延期超过 7 天的里程碑/任务节点'),
    ('schedule_overdue_severe', '关键节点严重延期', 'schedule', 'max_overdue_days', '>', 30, '警告',
     '存在延期超过 30 天的节点'),
    ('staff_overload_warn', '人员工时过载预警', 'staff', 'max_load_rate', '>', 1.2, '提醒',
     '人员已分配工时超过其月可用工时的 120%'),
]


def seed_plm_master():
    """幂等预置字典、预警规则与全局参数（已存在则不覆盖用户改动）。"""
    conn = get_db()
    c = conn.cursor()
    for category, key, label, sort in DEFAULT_DICT:
        c.execute("INSERT OR IGNORE INTO plm_dict (category,key,label,sort) VALUES (?,?,?,?)",
                  (category, key, label, sort))
    for rule_key, name, dim, metric, op, th, level, desc in DEFAULT_ALERT_RULES:
        c.execute("""INSERT OR IGNORE INTO plm_alert_rule
                     (rule_key,rule_name,dim,metric,op,threshold,level,enabled,description)
                     VALUES (?,?,?,?,?,?,?,?,?)""",
                  (rule_key, name, dim, metric, op, th, level, 1, desc))
    for key, value, desc in DEFAULT_CONFIG:
        c.execute("INSERT OR IGNORE INTO plm_config (key,value,description) VALUES (?,?,?)",
                  (key, value, desc))
    conn.commit()
    conn.close()


# ===================== 配置 / 字典 / 日志 =====================

def get_config(key, default=''):
    conn = get_db()
    r = conn.execute("SELECT value FROM plm_config WHERE key=?", (key,)).fetchone()
    conn.close()
    return r['value'] if r else default


def set_config(key, value, description='', operator='admin'):
    conn = get_db()
    if description:
        conn.execute("""INSERT INTO plm_config (key,value,description,updated_at) VALUES (?,?,?,?)
                        ON CONFLICT(key) DO UPDATE SET value=excluded.value,
                        description=excluded.description, updated_at=excluded.updated_at""",
                     (key, str(value), description, _now()))
    else:
        conn.execute("""INSERT INTO plm_config (key,value,updated_at) VALUES (?,?,?)
                        ON CONFLICT(key) DO UPDATE SET value=excluded.value,
                        updated_at=excluded.updated_at""", (key, str(value), _now()))
    conn.commit()
    conn.close()
    log_op('config', key, key, '修改全局参数', {'value': value}, operator=operator)
    return get_config(key)


def list_config():
    conn = get_db()
    rows = _rows(conn.execute("SELECT * FROM plm_config ORDER BY key"))
    conn.close()
    return rows


def list_dict(category=None):
    conn = get_db()
    if category:
        rows = _rows(conn.execute(
            "SELECT * FROM plm_dict WHERE category=? AND enabled=1 ORDER BY sort,id", (category,)))
    else:
        rows = _rows(conn.execute(
            "SELECT * FROM plm_dict WHERE enabled=1 ORDER BY category,sort,id"))
    conn.close()
    return rows


def dict_options(category):
    """返回某分类的 label 列表，供表单下拉使用。"""
    return [d['label'] for d in list_dict(category)]


def create_dict(category, key, label, sort=0, remark='', operator='admin'):
    if not category or not key or not label:
        return {'success': False, 'error': 'category / key / label 均为必填'}
    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO plm_dict (category,key,label,sort,remark) VALUES (?,?,?,?,?)",
            (category, key, label, int(sort or 0), remark))
        conn.commit()
        did = cur.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        return {'success': False, 'error': '同分类下 key 已存在'}
    conn.close()
    log_op('dict', did, '%s:%s' % (category, key), '新增字典', {'label': label}, operator=operator)
    return {'success': True, 'id': did}


def delete_dict(dict_id, operator='admin'):
    conn = get_db()
    r = conn.execute("SELECT * FROM plm_dict WHERE id=?", (dict_id,)).fetchone()
    if not r:
        conn.close()
        return {'success': False, 'error': '字典项不存在'}
    conn.execute("UPDATE plm_dict SET enabled=0 WHERE id=?", (dict_id,))
    conn.commit()
    conn.close()
    log_op('dict', dict_id, r['label'], '停用字典', operator=operator)
    return {'success': True}


def log_op(target_type, target_id, target_name, action, change=None,
           operator='', role='', remark=''):
    conn = get_db()
    conn.execute("""INSERT INTO plm_op_log
                    (target_type,target_id,target_name,action,change_json,operator,role,remark)
                    VALUES (?,?,?,?,?,?,?,?)""",
                 (target_type, str(target_id), target_name, action,
                  json.dumps(change or {}, ensure_ascii=False), operator, role, remark))
    conn.commit()
    conn.close()


def list_logs(target_type=None, target_id=None, limit=200):
    conn = get_db()
    sql, args, conds = "SELECT * FROM plm_op_log", [], []
    if target_type:
        conds.append("target_type=?"); args.append(target_type)
    if target_id:
        conds.append("target_id=?"); args.append(str(target_id))
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    for r in rows:
        try:
            r['change'] = json.loads(r.get('change_json') or '{}')
        except (ValueError, TypeError):
            r['change'] = {}
    return rows


# ===================== 通用工具：编号与字段更新 =====================

def _next_no(conn, table, column, prefix):
    """生成流水编号：<PREFIX>-<年份>-<3 位序号>。"""
    year = datetime.now().strftime('%Y')
    like = '%s-%s-%%' % (prefix, year)
    rows = _rows(conn.execute(
        "SELECT %s n FROM %s WHERE %s LIKE ?" % (column, table, column), (like,)))
    seqs = []
    for r in rows:
        try:
            seqs.append(int(str(r['n']).rsplit('-', 1)[1]))
        except (IndexError, ValueError):
            continue
    nxt = (max(seqs) + 1) if seqs else 1
    return '%s-%s-%03d' % (prefix, year, nxt)


def _update_by_fields(table, log_type, obj_id, fields, payload, numeric=(),
                      label_field='id', action='修改', operator=''):
    """按允许字段做 UPDATE，并把前后差异写入操作日志。"""
    sets, args, touched = [], [], []
    for k in fields:
        if k in payload:
            sets.append("%s=?" % k)
            args.append(_f(payload[k]) if k in numeric else payload[k])
            touched.append(k)
    if not sets:
        return {'success': False, 'error': '无可更新字段'}
    conn = get_db()
    old = _first(conn.execute("SELECT * FROM %s WHERE id=?" % table, (obj_id,)))
    if not old:
        conn.close()
        return {'success': False, 'error': '记录不存在'}
    try:
        conn.execute("UPDATE %s SET %s, updated_at=? WHERE id=?" % (table, ",".join(sets)),
                     args + [_now(), obj_id])
        conn.commit()
    except sqlite3.IntegrityError as e:
        conn.close()
        return {'success': False, 'error': '唯一键冲突：%s' % e}
    new = _first(conn.execute("SELECT * FROM %s WHERE id=?" % table, (obj_id,)))
    conn.close()
    changed = {}
    for k in touched:
        if old.get(k) != (new or {}).get(k):
            changed[k] = {'before': old.get(k), 'after': (new or {}).get(k)}
    log_op(log_type, obj_id, str(old.get(label_field, obj_id)), action, changed,
           operator=operator)
    return {'success': True, 'changed': changed}


# ===================== 模块一：商机与售前投标概算 =====================

def _follow_list(raw):
    try:
        v = json.loads(raw or '[]')
        return v if isinstance(v, list) else []
    except (ValueError, TypeError):
        return []


def list_opportunities(keyword=None, status=None, limit=500):
    conn = get_db()
    sql, args, conds = "SELECT * FROM plm_opportunity", [], []
    if status and status != '全部':
        conds.append("status=?"); args.append(status)
    if keyword:
        conds.append("(opp_no LIKE ? OR opp_name LIKE ? OR customer LIKE ? OR owner LIKE ?)")
        args += ['%' + keyword + '%'] * 4
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    for r in rows:
        b = _first(conn.execute(
            "SELECT id FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=? "
            "AND stage='estimate_bid' ORDER BY id DESC LIMIT 1", (r['id'],)))
        r['has_estimate'] = bool(b)
        r['item_count'] = conn.execute(
            "SELECT COUNT(*) n FROM plm_baseline_item WHERE baseline_id=?",
            (b['id'] if b else -1,)).fetchone()['n']
    conn.close()
    return rows


def get_opportunity(opp_id):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_opportunity WHERE id=?", (opp_id,)))
    if r:
        r['follow_records'] = _follow_list(r.get('follow_log'))
        r['docs'] = _rows(conn.execute(
            "SELECT * FROM plm_presale_doc WHERE opportunity_id=? ORDER BY id DESC", (opp_id,)))
        b = _first(conn.execute(
            "SELECT * FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=? "
            "AND stage='estimate_bid' ORDER BY id DESC LIMIT 1", (opp_id,)))
        if b:
            b['stage_cn'] = BASELINE_STAGE_CN.get(b['stage'], b['stage'])
            b['items'] = _rows(conn.execute(
                "SELECT * FROM plm_baseline_item WHERE baseline_id=? ORDER BY id", (b['id'],)))
        r['estimate'] = b
        r['contracts'] = _rows(conn.execute(
            "SELECT id,contract_no,contract_name,sign_amount FROM plm_contract "
            "WHERE opportunity_id=?", (opp_id,)))
        r['projects'] = _rows(conn.execute(
            "SELECT id,project_no,project_name,status FROM plm_project WHERE opportunity_id=?",
            (opp_id,)))
    conn.close()
    return r


def create_opportunity(payload, operator=''):
    opp_no = (payload.get('opp_no') or '').strip()
    opp_name = (payload.get('opp_name') or '').strip()
    if not opp_name:
        return {'success': False, 'error': '商机名称必填'}
    conn = get_db()
    if not opp_no:
        opp_no = _next_no(conn, 'plm_opportunity', 'opp_no', 'SJ')
    try:
        cur = conn.execute("""INSERT INTO plm_opportunity
            (opp_no,opp_name,customer,industry,region,dept,owner,status,bid_date,
             expect_income,won_at,remark) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (opp_no, opp_name, payload.get('customer', ''), payload.get('industry', ''),
             payload.get('region', ''), payload.get('dept', ''), payload.get('owner', ''),
             payload.get('status', '跟进中'), payload.get('bid_date', ''),
             _f(payload.get('expect_income')), payload.get('won_at', ''),
             payload.get('remark', '')))
        oid = cur.lastrowid
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {'success': False, 'error': '商机编号已存在：%s' % opp_no}
    conn.close()
    log_op('opportunity', oid, opp_name, '新增商机', {'opp_no': opp_no}, operator=operator)
    return {'success': True, 'id': oid, 'opp_no': opp_no}


def update_opportunity(opp_id, payload, operator=''):
    fields = ('opp_name', 'customer', 'industry', 'region', 'dept', 'owner', 'status',
              'bid_date', 'expect_income', 'won_at', 'remark')
    res = _update_by_fields('plm_opportunity', 'opportunity', opp_id, fields, payload,
                            numeric=('expect_income',), label_field='opp_name',
                            action='修改商机', operator=operator)
    if res.get('success'):
        recompute_opportunity_rollup(opp_id)
    return res


def delete_opportunity(opp_id, operator=''):
    conn = get_db()
    o = _first(conn.execute("SELECT * FROM plm_opportunity WHERE id=?", (opp_id,)))
    if not o:
        conn.close()
        return {'success': False, 'error': '商机不存在'}
    refs = {
        '合同': conn.execute("SELECT COUNT(*) n FROM plm_contract WHERE opportunity_id=?",
                            (opp_id,)).fetchone()['n'],
        '项目': conn.execute("SELECT COUNT(*) n FROM plm_project WHERE opportunity_id=?",
                            (opp_id,)).fetchone()['n'],
    }
    if any(refs.values()):
        conn.close()
        return {'success': False, 'error': '该商机已被下游引用，无法删除', 'refs': refs}
    bids = [r['id'] for r in conn.execute(
        "SELECT id FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=?",
        (opp_id,)).fetchall()]
    for bid in bids:
        conn.execute("DELETE FROM plm_baseline_item WHERE baseline_id=?", (bid,))
    conn.execute("DELETE FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=?",
                 (opp_id,))
    conn.execute("DELETE FROM plm_presale_doc WHERE opportunity_id=?", (opp_id,))
    conn.execute("DELETE FROM plm_opportunity WHERE id=?", (opp_id,))
    conn.commit()
    conn.close()
    log_op('opportunity', opp_id, o['opp_name'], '删除商机', operator=operator)
    return {'success': True}


def add_follow_record(opp_id, content, operator='', follow_time=None):
    if not content:
        return {'success': False, 'error': '跟进内容不能为空'}
    conn = get_db()
    o = _first(conn.execute("SELECT * FROM plm_opportunity WHERE id=?", (opp_id,)))
    if not o:
        conn.close()
        return {'success': False, 'error': '商机不存在'}
    rec = _follow_list(o['follow_log'])
    rec.append({'time': follow_time or _now(), 'owner': operator or o['owner'] or '-',
                'content': content})
    conn.execute("UPDATE plm_opportunity SET follow_log=?, updated_at=? WHERE id=?",
                 (json.dumps(rec, ensure_ascii=False), _now(), opp_id))
    conn.commit()
    conn.close()
    log_op('opportunity', opp_id, o['opp_name'], '新增跟进记录', {'content': content},
           operator=operator)
    return {'success': True, 'records': rec}


def get_opportunity_estimate(opp_id):
    conn = get_db()
    b = _first(conn.execute(
        "SELECT * FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=? "
        "AND stage='estimate_bid' ORDER BY id DESC LIMIT 1", (opp_id,)))
    if b:
        b['stage_cn'] = BASELINE_STAGE_CN.get(b['stage'], b['stage'])
        b['items'] = _rows(conn.execute(
            "SELECT * FROM plm_baseline_item WHERE baseline_id=? ORDER BY id", (b['id'],)))
    conn.close()
    return b


def save_opportunity_estimate(opp_id, payload, operator=''):
    """录入/更新投标概算：总额 + 分项明细，自动汇总概算成本、预估毛利与毛利率（FR-1）。"""
    conn = get_db()
    o = _first(conn.execute("SELECT * FROM plm_opportunity WHERE id=?", (opp_id,)))
    if not o:
        conn.close()
        return {'success': False, 'error': '商机不存在'}
    income = _f(payload.get('total_income', o['expect_income']))
    items = payload.get('items') or []
    b = _first(conn.execute(
        "SELECT * FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=? "
        "AND stage='estimate_bid' ORDER BY id DESC LIMIT 1", (opp_id,)))
    bid = b['id'] if b else None
    if bid:
        conn.execute("DELETE FROM plm_baseline_item WHERE baseline_id=?", (bid,))
    else:
        cur = conn.execute("""INSERT INTO plm_baseline
            (scope_type,scope_id,stage,total_income,total_cost,gross,gross_rate,status,
             created_by,remark,updated_at)
            VALUES ('opportunity',?,'estimate_bid',?,0,0,NULL,'草稿',?,?,?)""",
            (opp_id, income, operator or o['owner'], payload.get('remark', ''), _now()))
        bid = cur.lastrowid
    total_cost = 0.0
    for it in items:
        amt = _f(it.get('plan_amount'))
        total_cost += amt
        conn.execute("""INSERT INTO plm_baseline_item
            (baseline_id,category,item_name,plan_amount,actual_amount,remark) VALUES (?,?,?,?,?,?)""",
            (bid, it.get('category', '其他费用'), it.get('item_name', ''), amt,
             it.get('actual_amount'), it.get('remark', '')))
    if not items and payload.get('total_cost') is not None:
        total_cost = _f(payload.get('total_cost'))
    total_cost = _r(total_cost)
    gross = _r(income - total_cost)
    rate = _rate(gross, income)
    conn.execute("""UPDATE plm_baseline SET total_income=?,total_cost=?,gross=?,gross_rate=?,
                    remark=?,updated_at=? WHERE id=?""",
                 (income, total_cost, gross, rate,
                  payload.get('remark', (b or {}).get('remark', '')), _now(), bid))
    conn.execute("""UPDATE plm_opportunity SET expect_income=?,est_cost=?,est_gross=?,
                    est_gross_rate=?,updated_at=? WHERE id=?""",
                 (income, total_cost, gross, rate, _now(), opp_id))
    conn.commit()
    conn.close()
    log_op('estimate', bid, o['opp_name'], '录入投标概算',
           {'total_income': income, 'total_cost': total_cost}, operator=operator)
    return {'success': True, 'baseline_id': bid, 'total_income': income,
            'total_cost': total_cost, 'gross': gross, 'gross_rate': rate}


def recompute_opportunity_rollup(opp_id):
    """预估收入变化后同步概算基线收入与毛利指标；无概算时毛利置空。"""
    conn = get_db()
    o = _first(conn.execute("SELECT * FROM plm_opportunity WHERE id=?", (opp_id,)))
    if not o:
        conn.close()
        return None
    income = _f(o['expect_income'])
    b = _first(conn.execute(
        "SELECT * FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=? "
        "AND stage='estimate_bid' ORDER BY id DESC LIMIT 1", (opp_id,)))
    if not b:
        conn.execute("UPDATE plm_opportunity SET est_gross=NULL, est_gross_rate=NULL WHERE id=?",
                     (opp_id,))
        conn.commit()
        conn.close()
        return None
    items = _rows(conn.execute("SELECT plan_amount FROM plm_baseline_item WHERE baseline_id=?",
                              (b['id'],)))
    cost = sum(_f(i['plan_amount']) for i in items) if items else _f(b['total_cost'])
    gross = _r(income - cost)
    rate = _rate(gross, income)
    conn.execute("UPDATE plm_baseline SET total_income=?,gross=?,gross_rate=?,updated_at=? WHERE id=?",
                 (income, gross, rate, _now(), b['id']))
    conn.execute("""UPDATE plm_opportunity SET est_cost=?,est_gross=?,est_gross_rate=?,
                    updated_at=? WHERE id=?""", (_r(cost), gross, rate, _now(), opp_id))
    conn.commit()
    conn.close()
    return {'total_cost': _r(cost), 'gross': gross, 'gross_rate': rate}


# ---------- 售前资料归档 ----------

def list_presale_docs(opp_id=None):
    conn = get_db()
    if opp_id:
        rows = _rows(conn.execute(
            "SELECT * FROM plm_presale_doc WHERE opportunity_id=? ORDER BY id DESC", (opp_id,)))
    else:
        rows = _rows(conn.execute("SELECT * FROM plm_presale_doc ORDER BY id DESC LIMIT 500"))
    conn.close()
    return rows


def create_presale_doc(payload, operator=''):
    opp_id = payload.get('opportunity_id')
    name = (payload.get('doc_name') or '').strip()
    if not opp_id or not name:
        return {'success': False, 'error': 'opportunity_id 与 doc_name 必填'}
    conn = get_db()
    if not _first(conn.execute("SELECT id FROM plm_opportunity WHERE id=?", (opp_id,))):
        conn.close()
        return {'success': False, 'error': '商机不存在'}
    cur = conn.execute("""INSERT INTO plm_presale_doc
        (opportunity_id,doc_name,doc_type,file_name,file_path,remark,uploader)
        VALUES (?,?,?,?,?,?,?)""",
        (opp_id, name, payload.get('doc_type', '其他'), payload.get('file_name', ''),
         payload.get('file_path', ''), payload.get('remark', ''), operator))
    did = cur.lastrowid
    conn.commit()
    conn.close()
    log_op('presale_doc', did, name, '归档售前资料', {'doc_type': payload.get('doc_type')},
           operator=operator)
    return {'success': True, 'id': did}


def delete_presale_doc(doc_id, operator=''):
    conn = get_db()
    d = _first(conn.execute("SELECT * FROM plm_presale_doc WHERE id=?", (doc_id,)))
    if not d:
        conn.close()
        return {'success': False, 'error': '资料不存在'}
    conn.execute("DELETE FROM plm_presale_doc WHERE id=?", (doc_id,))
    conn.commit()
    conn.close()
    log_op('presale_doc', doc_id, d['doc_name'], '删除售前资料', operator=operator)
    return {'success': True}


# ===================== 模块二：合同与项目立项 =====================

def list_contracts(keyword=None, limit=500):
    conn = get_db()
    sql, args = "SELECT * FROM plm_contract", []
    if keyword:
        sql += " WHERE (contract_no LIKE ? OR contract_name LIKE ? OR customer LIKE ?)"
        args += ['%' + keyword + '%'] * 3
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    for r in rows:
        r['project_count'] = conn.execute(
            "SELECT COUNT(*) n FROM plm_project WHERE contract_id=?", (r['id'],)).fetchone()['n']
        opp = r.get('opportunity_id')
        o = _first(conn.execute("SELECT opp_no,opp_name FROM plm_opportunity WHERE id=?", (opp,))) \
            if opp else None
        r['opp_no'] = o['opp_no'] if o else ''
        r['opp_name'] = o['opp_name'] if o else ''
    conn.close()
    return rows


def get_contract(contract_id):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_contract WHERE id=?", (contract_id,)))
    if r:
        r['projects'] = _rows(conn.execute(
            "SELECT id,project_no,project_name,status,manager FROM plm_project WHERE contract_id=?",
            (contract_id,)))
        if r.get('opportunity_id'):
            r['opportunity'] = _first(conn.execute(
                "SELECT id,opp_no,opp_name,status,customer FROM plm_opportunity WHERE id=?",
                (r['opportunity_id'],)))
    conn.close()
    return r


def create_contract(payload, operator=''):
    no = (payload.get('contract_no') or '').strip()
    conn = get_db()
    if not no:
        no = _next_no(conn, 'plm_contract', 'contract_no', 'HT')
    try:
        cur = conn.execute("""INSERT INTO plm_contract
            (contract_no,contract_name,customer,industry,region,dept,sign_amount,sign_date,
             project_cycle,status,owner,opportunity_id,remark) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (no, payload.get('contract_name', ''), payload.get('customer', ''),
             payload.get('industry', ''), payload.get('region', ''), payload.get('dept', ''),
             _f(payload.get('sign_amount')), payload.get('sign_date', ''),
             payload.get('project_cycle', ''), payload.get('status', '已签署'),
             payload.get('owner', ''), payload.get('opportunity_id'),
             payload.get('remark', '')))
        cid = cur.lastrowid
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {'success': False, 'error': '合同编号已存在：%s' % no}
    conn.close()
    log_op('contract', cid, no, '新增合同', {'sign_amount': payload.get('sign_amount')},
           operator=operator)
    return {'success': True, 'id': cid, 'contract_no': no}


def update_contract(contract_id, payload, operator=''):
    fields = ('contract_no', 'contract_name', 'customer', 'industry', 'region', 'dept',
              'sign_amount', 'sign_date', 'project_cycle', 'status', 'owner',
              'opportunity_id', 'remark')
    return _update_by_fields('plm_contract', 'contract', contract_id, fields, payload,
                             numeric=('sign_amount',), label_field='contract_no',
                             action='修改合同', operator=operator)


def delete_contract(contract_id, operator=''):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_contract WHERE id=?", (contract_id,)))
    if not r:
        conn.close()
        return {'success': False, 'error': '合同不存在'}
    refs = {'项目': conn.execute("SELECT COUNT(*) n FROM plm_project WHERE contract_id=?",
                                (contract_id,)).fetchone()['n']}
    if any(refs.values()):
        conn.close()
        return {'success': False, 'error': '合同下仍有项目，无法删除', 'refs': refs}
    conn.execute("DELETE FROM plm_contract WHERE id=?", (contract_id,))
    conn.commit()
    conn.close()
    log_op('contract', contract_id, r['contract_no'], '删除合同', operator=operator)
    return {'success': True}


def list_projects(keyword=None, status=None, limit=500):
    conn = get_db()
    sql, args, conds = "SELECT * FROM plm_project", [], []
    if status and status != '全部':
        conds.append("status=?"); args.append(status)
    if keyword:
        conds.append("(project_no LIKE ? OR project_name LIKE ? OR customer LIKE ? OR manager LIKE ?)")
        args += ['%' + keyword + '%'] * 4
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    for r in rows:
        r.update(_project_brief_metrics(r['id']))
    return rows


def _project_brief_metrics(pid):
    conn = get_db()
    est = _first(conn.execute(
        "SELECT total_cost FROM plm_baseline WHERE scope_type='project' AND scope_id=? "
        "AND stage IN ('estimate_locked','estimate_bid') ORDER BY id DESC LIMIT 1", (pid,)))
    bud = _first(conn.execute(
        "SELECT total_cost FROM plm_baseline WHERE scope_type='project' AND scope_id=? "
        "AND stage='budget' ORDER BY id DESC LIMIT 1", (pid,)))
    actual = conn.execute("SELECT COALESCE(SUM(amount),0) a FROM plm_ledger WHERE project_id=? "
                          "AND kind='cost' AND plan_or_actual='实际'", (pid,)).fetchone()['a']
    income = conn.execute("SELECT COALESCE(SUM(amount),0) a FROM plm_ledger WHERE project_id=? "
                          "AND kind='income'", (pid,)).fetchone()['a']
    ms = _rows(conn.execute("SELECT status,plan_end FROM plm_milestone WHERE project_id=?", (pid,)))
    alerts = conn.execute("SELECT COUNT(*) n FROM plm_alert WHERE project_id=? "
                          "AND status<>'已闭环'", (pid,)).fetchone()['n']
    conn.close()
    today = _today()
    total_ms = len(ms)
    done_ms = len([m for m in ms if m['status'] == '已完成'])
    overdue_ms = len([m for m in ms if m['status'] != '已完成' and m['plan_end']
                      and str(m['plan_end'])[:10] < today])
    income, actual = _f(income), _f(actual)
    return {
        'estimate_cost': _r(_f(est['total_cost'])) if est else None,
        'budget_cost': _r(_f(bud['total_cost'])) if bud else None,
        'actual_cost': _r(actual),
        'income': _r(income),
        'actual_gross': _r(income - actual) if (income or actual) else None,
        'actual_gross_rate': _rate(income - actual, income),
        'milestone_total': total_ms,
        'milestone_done': done_ms,
        'milestone_overdue': overdue_ms,
        'open_alerts': alerts,
    }


def get_project(project_id):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_project WHERE id=?", (project_id,)))
    if not r:
        conn.close()
        return None
    if r.get('contract_id'):
        r['contract'] = _first(conn.execute(
            "SELECT id,contract_no,contract_name,customer,sign_amount,sign_date,status "
            "FROM plm_contract WHERE id=?", (r['contract_id'],)))
    if r.get('opportunity_id'):
        r['opportunity'] = _first(conn.execute(
            "SELECT id,opp_no,opp_name,status,customer FROM plm_opportunity WHERE id=?",
            (r['opportunity_id'],)))
    conn.close()
    return r


def create_project(payload, operator=''):
    name = (payload.get('project_name') or '').strip()
    if not name:
        return {'success': False, 'error': '项目名称必填'}
    conn = get_db()
    no = (payload.get('project_no') or '').strip()
    if not no:
        no = _next_no(conn, 'plm_project', 'project_no', 'XM')
    if payload.get('contract_id') and not _first(
            conn.execute("SELECT id FROM plm_contract WHERE id=?", (payload['contract_id'],))):
        conn.close()
        return {'success': False, 'error': '所关联合同不存在'}
    try:
        cur = conn.execute("""INSERT INTO plm_project
            (project_no,project_name,customer,dept,region,manager,status,contract_id,
             opportunity_id,start_date,end_date,kickoff_date,remark)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (no, name, payload.get('customer', ''), payload.get('dept', ''),
             payload.get('region', ''), payload.get('manager', ''),
             payload.get('status', '待启动'), payload.get('contract_id'),
             payload.get('opportunity_id'), payload.get('start_date', ''),
             payload.get('end_date', ''), payload.get('kickoff_date', ''),
             payload.get('remark', '')))
        pid = cur.lastrowid
        if payload.get('opportunity_id'):
            conn.execute("UPDATE plm_opportunity SET status='中标', updated_at=? "
                         "WHERE id=? AND status IN ('跟进中','投标中')", (_now(),
                                                                          payload['opportunity_id']))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {'success': False, 'error': '项目编号已存在：%s' % no}
    conn.close()
    log_op('project', pid, name, '项目立项', {'project_no': no}, operator=operator)
    return {'success': True, 'id': pid, 'project_no': no}


def update_project(project_id, payload, operator=''):
    fields = ('project_no', 'project_name', 'customer', 'dept', 'region', 'manager',
              'status', 'contract_id', 'opportunity_id', 'start_date', 'end_date',
              'kickoff_date', 'remark')
    return _update_by_fields('plm_project', 'project', project_id, fields, payload,
                             label_field='project_no', action='修改项目', operator=operator)


def delete_project(project_id, operator=''):
    conn = get_db()
    p = _first(conn.execute("SELECT * FROM plm_project WHERE id=?", (project_id,)))
    if not p:
        conn.close()
        return {'success': False, 'error': '项目不存在'}
    refs = {
        '里程碑': conn.execute("SELECT COUNT(*) n FROM plm_milestone WHERE project_id=?",
                              (project_id,)).fetchone()['n'],
        '任务': conn.execute("SELECT COUNT(*) n FROM plm_task WHERE project_id=?",
                            (project_id,)).fetchone()['n'],
        '收支台账': conn.execute("SELECT COUNT(*) n FROM plm_ledger WHERE project_id=?",
                               (project_id,)).fetchone()['n'],
        '工时': conn.execute("SELECT COUNT(*) n FROM plm_timesheet WHERE project_id=?",
                            (project_id,)).fetchone()['n'],
        '未闭环预警': conn.execute("SELECT COUNT(*) n FROM plm_alert WHERE project_id=? "
                                 "AND status<>'已闭环'", (project_id,)).fetchone()['n'],
    }
    if any(refs.values()):
        conn.close()
        return {'success': False, 'error': '项目下仍有执行数据，无法删除；请先清理或将项目结项',
                'refs': refs}
    bids = [r['id'] for r in conn.execute(
        "SELECT id FROM plm_baseline WHERE scope_type='project' AND scope_id=?",
        (project_id,)).fetchall()]
    for bid in bids:
        conn.execute("DELETE FROM plm_baseline_item WHERE baseline_id=?", (bid,))
    conn.execute("DELETE FROM plm_baseline WHERE scope_type='project' AND scope_id=?", (project_id,))
    conn.execute("DELETE FROM plm_alert WHERE project_id=?", (project_id,))
    conn.execute("DELETE FROM plm_assignment WHERE project_id=?", (project_id,))
    conn.execute("DELETE FROM plm_milestone WHERE project_id=?", (project_id,))
    conn.execute("DELETE FROM plm_task WHERE project_id=?", (project_id,))
    conn.execute("DELETE FROM plm_project WHERE id=?", (project_id,))
    conn.commit()
    conn.close()
    log_op('project', project_id, p['project_name'], '删除项目', operator=operator)
    return {'success': True}


# ===================== 模块二/三：四算基线 =====================

def list_baselines(project_id=None, scope_type='project', scope_id=None, scope_key=None):
    conn = get_db()
    if project_id is not None:
        scope_type, scope_id = 'project', project_id
    if scope_key is not None:
        rows = _rows(conn.execute(
            "SELECT * FROM plm_baseline WHERE scope_type=? AND scope_key=? ORDER BY calc_type, version",
            (scope_type, scope_key)))
    else:
        rows = _rows(conn.execute(
            "SELECT * FROM plm_baseline WHERE scope_type=? AND scope_id=? ORDER BY id",
            (scope_type, scope_id)))
    for r in rows:
        r['stage_cn'] = BASELINE_STAGE_CN.get(r['stage'], r['stage'])
        r['calc_type_cn'] = COST_BASELINE_CALC_TYPE_CN.get(r.get('calc_type') or '', r.get('calc_type'))
        r['items'] = _rows(conn.execute(
            "SELECT * FROM plm_baseline_item WHERE baseline_id=? ORDER BY id", (r['id'],)))
        r['calculated'] = r['stage'] in BASELINE_CALCULATED
        r['reserved'] = r['stage'] in BASELINE_RESERVED
    conn.close()
    return rows


def get_baseline(baseline_id):
    conn = get_db()
    b = _first(conn.execute("SELECT * FROM plm_baseline WHERE id=?", (baseline_id,)))
    if b:
        b['items'] = _rows(conn.execute(
            "SELECT * FROM plm_baseline_item WHERE baseline_id=? ORDER BY id", (baseline_id,)))
        b['stage_cn'] = BASELINE_STAGE_CN.get(b['stage'], b['stage'])
        b['reserved'] = b['stage'] in BASELINE_RESERVED
    conn.close()
    return b


def _estimate_total(conn, project_id):
    r = _first(conn.execute(
        "SELECT total_cost FROM plm_baseline WHERE scope_type='project' AND scope_id=? "
        "AND stage IN ('estimate_locked','estimate_bid') "
        "ORDER BY CASE stage WHEN 'estimate_locked' THEN 0 ELSE 1 END, id DESC LIMIT 1",
        (project_id,)))
    return _f(r['total_cost']) if r else None


def _check_baseline_constraint(conn, stage, scope_type, scope_id, total_cost):
    """FR-3：管控开关开启时预算超概算拒绝保存。返回 (ok, message)。"""
    if stage != 'budget':
        return True, ''
    if str(get_config('baseline_constraint', 'off')).strip().lower() != 'on':
        return True, ''
    if scope_type != 'project':
        return True, ''
    est = _estimate_total(conn, scope_id)
    if est is None:
        return False, '尚未录入概算基线，基线管控开启时不允许直接保存预算'
    if total_cost > est + 1e-6:
        return False, '预算超出概算 %s 元，基线管控开关已开启，禁止保存' % _r(total_cost - est)
    return True, ''


def _write_items(conn, baseline_id, items):
    conn.execute("DELETE FROM plm_baseline_item WHERE baseline_id=?", (baseline_id,))
    for it in items:
        conn.execute("""INSERT INTO plm_baseline_item
            (baseline_id,category,item_name,plan_amount,actual_amount,remark) VALUES (?,?,?,?,?,?)""",
            (baseline_id, it.get('category', '其他费用'), it.get('item_name', ''),
             _f(it.get('plan_amount')), it.get('actual_amount'), it.get('remark', '')))


def save_baseline(payload, operator=''):
    """新建或更新基线（分项明细 + 总额）。四算重构：优先以 calc_type(本体) 为准，
    scope_type='contract' 时以 scope_key(合同号) 归集，对齐 ontos CostBaseline。"""
    scope_type = payload.get('scope_type', 'project')
    scope_id = payload.get('scope_id') or payload.get('project_id')
    scope_key = payload.get('scope_key') or (str(scope_id) if scope_type == 'contract' else None)
    # calc_type(本体) 优先；未传则按 stage 反推（兼容旧项目级调用）
    calc_type = payload.get('calc_type')
    stage = payload.get('stage')
    if calc_type and calc_type not in COST_BASELINE_CALC_TYPES:
        return {'success': False, 'error': 'calc_type 需属于 %s' % list(COST_BASELINE_CALC_TYPES)}
    if calc_type:
        stage = CALC_TYPE_TO_STAGE.get(calc_type, stage)
    elif stage:
        calc_type = STAGE_TO_CALC_TYPE.get(stage, '')
    if not scope_id and not scope_key:
        return {'success': False, 'error': 'scope_id/scope_key 必填'}
    if stage not in BASELINE_STAGE:
        return {'success': False, 'error': 'stage 需属于 %s' % list(BASELINE_STAGE)}
    items = payload.get('items') or []
    total_cost = _f(payload.get('total_cost'))
    if items:
        total_cost = _r(sum(_f(i.get('plan_amount')) for i in items))
    reserved = stage in BASELINE_RESERVED
    conn = get_db()
    if scope_type == 'project':
        if not _first(conn.execute("SELECT id FROM plm_project WHERE id=?", (scope_id,))):
            conn.close()
            return {'success': False, 'error': '项目不存在'}
    elif scope_type == 'contract':
        if not _first(conn.execute('SELECT "合同编号" FROM md_contract WHERE "合同编号"=?', (scope_key,))):
            conn.close()
            return {'success': False, 'error': '合同 %s 不存在' % scope_key}
    ok, msg = _check_baseline_constraint(conn, stage, scope_type, scope_id, total_cost)
    if not ok:
        conn.close()
        return {'success': False, 'error': msg, 'blocked': True}
    income = payload.get('total_income')
    if income is None:
        income = _default_income(conn, scope_type, scope_id)
    income = _f(income)
    bid = payload.get('id')
    if bid:
        old = _first(conn.execute("SELECT * FROM plm_baseline WHERE id=?", (bid,)))
        if not old:
            conn.close()
            return {'success': False, 'error': '基线不存在'}
        _write_items(conn, bid, items)
        gross = None if reserved else _r(income - total_cost)
        rate = None if reserved else _rate(gross, income)
        conn.execute("""UPDATE plm_baseline SET total_income=?,total_cost=?,gross=?,gross_rate=?,
                        remark=?,updated_at=? WHERE id=?""",
                     (income, total_cost, gross, rate,
                      payload.get('remark', old['remark']), _now(), bid))
        conn.commit()
        conn.close()
        warned = old['status'] == '已锁定'
        log_op('baseline', bid, BASELINE_STAGE_CN.get(stage, stage),
               '调整已锁定概算基线' if warned else '修改基线',
               {'total_cost': {'before': old['total_cost'], 'after': total_cost}},
               operator=operator)
        return {'success': True, 'id': bid, 'total_cost': total_cost, 'gross': gross, 'calc_type': calc_type,
                'warning': '已锁定基线被调整，变更已留痕' if warned else ''}
    gross = None if reserved else _r(income - total_cost)
    rate = None if reserved else _rate(gross, income)
    cur = conn.execute("""INSERT INTO plm_baseline
        (scope_type,scope_id,stage,total_income,total_cost,gross,gross_rate,status,
         source_baseline_id,created_by,remark,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (scope_type, scope_id, stage, income, total_cost, gross, rate, '草稿',
         payload.get('source_baseline_id'), operator, payload.get('remark', ''), _now()))
    bid = cur.lastrowid
    _write_items(conn, bid, items)
    conn.commit()
    est = _estimate_total(conn, scope_id) if scope_type == 'project' else None
    conn.close()
    over = _r(total_cost - est) if (stage == 'budget' and est is not None
                                    and total_cost > est) else None
    log_op('baseline', bid, BASELINE_STAGE_CN.get(stage, stage), '新增基线',
           {'total_cost': total_cost}, operator=operator)
    return {'success': True, 'id': bid, 'total_cost': total_cost, 'gross': gross,
            'over_estimate': over,
            'warning': ('预算超出概算 %s 元（当前仅提示，未拦截）' % over) if over else '',
            'reserved': reserved}


def _default_income(conn, scope_type, scope_id):
    if scope_type != 'project':
        return 0.0
    p = _first(conn.execute("SELECT contract_id FROM plm_project WHERE id=?", (scope_id,)))
    if p and p.get('contract_id'):
        ct = _first(conn.execute("SELECT sign_amount FROM plm_contract WHERE id=?",
                                 (p['contract_id'],)))
        if ct:
            return _f(ct['sign_amount'])
    led = conn.execute("SELECT COALESCE(SUM(amount),0) a FROM plm_ledger WHERE project_id=? "
                       "AND kind='income'", (scope_id,)).fetchone()['a']
    return _f(led)


def confirm_baseline(baseline_id, operator=''):
    conn = get_db()
    b = _first(conn.execute("SELECT * FROM plm_baseline WHERE id=?", (baseline_id,)))
    if not b:
        conn.close()
        return {'success': False, 'error': '基线不存在'}
    conn.execute("UPDATE plm_baseline SET status='已确认', updated_at=? WHERE id=?",
                 (_now(), baseline_id))
    conn.commit()
    conn.close()
    log_op('baseline', baseline_id, BASELINE_STAGE_CN.get(b['stage'], b['stage']),
           '确认基线', operator=operator)
    return {'success': True}


def lock_baseline(baseline_id, operator=''):
    """锁定概算基线，形成项目顶层管控基线（FR-2）。"""
    conn = get_db()
    b = _first(conn.execute("SELECT * FROM plm_baseline WHERE id=?", (baseline_id,)))
    if not b:
        conn.close()
        return {'success': False, 'error': '基线不存在'}
    if b['stage'] not in ('estimate_bid', 'estimate_locked'):
        conn.close()
        return {'success': False, 'error': '仅概算基线可锁定'}
    conn.execute("""UPDATE plm_baseline SET status='已锁定', locked_at=?, locked_by=?,
                    updated_at=? WHERE id=?""", (_now(), operator, _now(), baseline_id))
    conn.commit()
    conn.close()
    log_op('baseline', baseline_id, BASELINE_STAGE_CN.get(b['stage'], b['stage']),
           '锁定概算基线', {'total_cost': b['total_cost']}, operator=operator)
    return {'success': True}


def delete_baseline(baseline_id, operator=''):
    conn = get_db()
    b = _first(conn.execute("SELECT * FROM plm_baseline WHERE id=?", (baseline_id,)))
    if not b:
        conn.close()
        return {'success': False, 'error': '基线不存在'}
    if b['status'] == '已锁定':
        conn.close()
        return {'success': False, 'error': '已锁定基线不可删除'}
    conn.execute("DELETE FROM plm_baseline_item WHERE baseline_id=?", (baseline_id,))
    conn.execute("DELETE FROM plm_baseline WHERE id=?", (baseline_id,))
    conn.commit()
    conn.close()
    log_op('baseline', baseline_id, BASELINE_STAGE_CN.get(b['stage'], b['stage']),
           '删除基线', operator=operator)
    return {'success': True}


def compare_baselines(project_id):
    """四算基线对比：概算 / 预算 /【预留】核算 /【预留】决算。"""
    rows = list_baselines(project_id=project_id)
    by_stage = {}
    for r in rows:
        by_stage[r['stage']] = r  # 同 stage 取最新（rows 已按 id 升序）
    est = by_stage.get('estimate_locked') or by_stage.get('estimate_bid')
    bud = by_stage.get('budget')

    def pack(b):
        if not b:
            return None
        return {'baseline_id': b['id'], 'stage': b['stage'],
                'stage_cn': b['stage_cn'], 'total_income': _r(b['total_income']),
                'total_cost': _r(b['total_cost']), 'gross': _r(b['gross']),
                'gross_rate': b['gross_rate'], 'status': b['status'],
                'item_count': len(b['items']),
                'items': b['items'], 'source_baseline_id': b.get('source_baseline_id')}
    out = {'estimate': pack(est), 'budget': pack(bud),
           'accounting': _pack_reserved(by_stage.get('accounting')),
           'final': _pack_reserved(by_stage.get('final')),
           'estimate_vs_budget': None, 'budget_usage_note': '',
           'constraint_on': str(get_config('baseline_constraint', 'off')).lower() == 'on'}
    if est and bud:
        diff = _r(_f(bud['total_cost']) - _f(est['total_cost']))
        out['estimate_vs_budget'] = diff
        if diff and diff > 0:
            out['budget_usage_note'] = '预算超出概算 %s 元' % diff
    return out


def _pack_reserved(b):
    return {'baseline_id': b['id'] if b else None,
            'stage': b['stage'] if b else None,
            'stage_cn': (BASELINE_STAGE_CN.get(b['stage'], b['stage']) if b else
                         ('核算（预留）')),
            'total_cost': _r(_f(b['total_cost'])) if b else None,
            'gross': None, 'gross_rate': None,
            'status': b['status'] if b else None,
            'reserved': True, 'note': RESERVED_NOTE}


# ---------- 四算重构：合同级对比（读 ontos CostBaseline 驱动） ----------

def compare_contract_baselines(contract_no):
    """四算基线对比（合同级）：按 ontos calc_type 分组取最新 version；同时兼容旧键 estimate/budget/..。"""
    rows = list_baselines(scope_type='contract', scope_key=contract_no)
    by_calc = {}
    for r in rows:
        ct = r.get('calc_type') or STAGE_TO_CALC_TYPE.get(r['stage'], '概算')
        if ct not in by_calc or (r.get('version') or 0) >= (by_calc[ct].get('version') or 0):
            by_calc[ct] = r

    def pack(b):
        if not b:
            return None
        return {'baseline_id': b['id'], 'calc_type': b.get('calc_type'),
                'calc_type_cn': COST_BASELINE_CALC_TYPE_CN.get(b.get('calc_type') or '', b.get('calc_type')),
                'stage': b['stage'], 'stage_cn': b['stage_cn'], 'version': b.get('version'),
                'total_income': _r(b['total_income']), 'total_cost': _r(b['total_cost']),
                'gross': _r(b['gross']), 'gross_rate': b['gross_rate'], 'status': b['status'],
                'item_count': len(b['items']), 'items': b['items'],
                'source_baseline_id': b.get('source_baseline_id')}

    out = {ct: pack(by_calc.get(ct)) for ct in COST_BASELINE_CALC_TYPES}
    out['estimate'] = out.get('概算')          # 兼容旧键
    out['budget'] = out.get('基准预算')
    out['accounting'] = out.get('核算')
    out['final'] = out.get('决算')
    # 一号可度量目标：决算毛利率 − 签单毛利率（≥0 达标）
    est_b, fin = by_calc.get('概算'), by_calc.get('决算')
    out['margin_goal'] = None
    if est_b and fin and est_b.get('gross_rate') is not None and fin.get('gross_rate') is not None:
        out['margin_goal'] = _r(fin['gross_rate'] - est_b['gross_rate'])
    out['margin_goal_note'] = '签单毛利率 %s / 决算毛利率 %s' % (
        (round(est_b['gross_rate'], 2) if est_b and est_b['gross_rate'] is not None else '-'),
        (round(fin['gross_rate'], 2) if fin and fin['gross_rate'] is not None else '-'))
    # 概算 vs 基准预算 差异
    eb, bud = by_calc.get('概算'), by_calc.get('基准预算')
    out['estimate_vs_budget'] = None
    out['budget_usage_note'] = ''
    if eb and bud:
        diff = _r(_f(bud['total_cost']) - _f(eb['total_cost']))
        out['estimate_vs_budget'] = diff
        if diff and diff > 0:
            out['budget_usage_note'] = '基准预算超出概算 %s 元' % diff
    out['constraint_on'] = str(get_config('baseline_constraint', 'off')).lower() == 'on'
    out['calc_types'] = list(COST_BASELINE_CALC_TYPES)
    return out


def _upsert_contract_baseline(conn, contract_no, calc_type, income, cost, operator):
    stage = CALC_TYPE_TO_STAGE.get(calc_type, 'budget')
    gross = _r(income - cost)
    rate = _rate(gross, income)
    exist = _first(conn.execute(
        "SELECT id FROM plm_baseline WHERE scope_type='contract' AND scope_key=? AND calc_type=?",
        (contract_no, calc_type)))
    if exist:
        conn.execute("""UPDATE plm_baseline SET total_income=?,total_cost=?,gross=?,gross_rate=?,
                        updated_at=? WHERE id=?""",
                     (income, cost, gross, rate, _now(), exist['id']))
    else:
        conn.execute("""INSERT INTO plm_baseline
            (scope_type,scope_id,scope_key,calc_type,stage,total_income,total_cost,gross,gross_rate,status,created_by,updated_at)
            VALUES ('contract',0,?,?,?,?,?,?,?,'草稿',?,?)""",
            (contract_no, calc_type, stage, income, cost, gross, rate, operator, _now()))


def seed_baselines_from_contracts(operator='seed'):
    """从 md_contract 灌入四算种子（读本体对齐）：累计实施成本预估→基准预算；累计实施成本实际→核算。"""
    conn = get_db()
    cols = [r[1] for r in conn.execute("PRAGMA table_info(md_contract)")]
    col_est = '累计实施成本预估' if '累计实施成本预估' in cols else None
    col_act = '累计实施成本实际' if '累计实施成本实际' in cols else None
    # 收入口径（签单收入）= 合同总金额（主合同额列）。md_contract 无纯 "合同额" 列，
    # 含"合同额"的均为衍生列（原币合同额/软件合同额/合同额差异…），故按候选链回退探测。
    _AMT_CANDIDATES = ['合同总金额', '原币合同额', '合同额', '"合同额"']
    col_amt = next((c for c in _AMT_CANDIDATES if c in cols), None)
    if not col_est or not col_act or not col_amt:
        conn.close()
        return {'success': False,
                'error': 'md_contract 缺少成本/合同额列（已探测候选：%s）' % _AMT_CANDIDATES}
    total = conn.execute('SELECT COUNT(*) FROM md_contract').fetchone()[0]
    rows = _rows(conn.execute('SELECT "合同编号","%s","%s","%s" FROM md_contract' % (col_amt, col_est, col_act)))
    n_budget = n_account = 0
    n_skip = 0
    for r in rows:
        cno = r['合同编号']
        # 表头污染行过滤：合同编号为空/等于表头/非编码格式的行跳过
        if not cno or str(cno).strip() == '合同编号' or str(cno).strip() == '':
            n_skip += 1
            continue
        amt = _f(r.get(col_amt))
        est = _f(r.get(col_est))
        act = _f(r.get(col_act))
        if est and est > 0:
            _upsert_contract_baseline(conn, cno, '基准预算', amt, est, operator)
            n_budget += 1
        if act and act > 0:
            _upsert_contract_baseline(conn, cno, '核算', amt, act, operator)
            n_account += 1
    conn.commit()
    conn.close()
    return {'success': True, 'total': total, 'n_budget': n_budget, 'n_account': n_account,
            'n_skip': n_skip,
            'income_col': col_amt,
            'coverage_note': '基准预算覆盖 %d/%d 合同；核算覆盖 %d/%d 合同（其余主数据无实际成本，非超支）；跳过 %d 行'
                            % (n_budget, total, n_account, total, n_skip)}


def list_master_contract_summary(keyword='', limit=2000):
    """轻量合同清单（四算归集锚数据源，读 md_contract 而非空表 plm_contract）。

    财经域四算基线的合同下拉用：每行只取 {contract_no, customer, amount, status}，
    避免下拉全量 204 列。合同号与 plm_baseline.scope_key 对齐（来源同为 md_contract）。
    过滤表头污染行（合同编号为空/等于表头）。
    """
    conn = get_db()
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(md_contract)")]
        cno_col = '合同编号' if '合同编号' in cols else None
        if not cno_col:
            return {'success': False, 'error': 'md_contract 缺合同编号列'}
        def _col(cands):
            return next((c for c in cands if c in cols), None)
        cust_col = _col(['甲方名称', '客户简称', '客户名称'])
        amt_col = _col(['合同总金额', '原币合同额'])
        st_col = _col(['合同状态', '状态'])
        sel = ['"%s"' % c for c in [cno_col, cust_col, amt_col, st_col] if c]
        sql = 'SELECT %s FROM md_contract' % ','.join(sel)
        args = []
        if keyword:
            sql += ' WHERE "%s" LIKE ?' % cno_col
            args.append('%' + keyword + '%')
        sql += ' LIMIT ?'; args.append(int(limit))
        rows = []
        for r in _rows(conn.execute(sql, args)):
            cno = r[cno_col]
            if not cno or str(cno).strip() in ('', '合同编号'):
                continue
            rows.append({
                'contract_no': str(cno).strip(),
                'customer': (r[cust_col] if cust_col else None) or '',
                'amount': _r(r[amt_col] if amt_col else 0),
                'status': (r[st_col] if st_col else None) or '',
            })
        # 保持合同号唯一（md_contract 有 2 行重复合同号）
        seen, uniq = set(), []
        for x in rows:
            if x['contract_no'] in seen:
                continue
            seen.add(x['contract_no']); uniq.append(x)
        return {'success': True, 'total': len(uniq), 'data': uniq}
    finally:
        conn.close()


# ---------- 中标商机一键联动立项（FR-2） ----------

def convert_opportunity(payload, operator=''):
    """中标商机 → 合同 + 项目 + 概算基线（草稿，待确认锁定）+ 顶层粗里程碑 + 签单收入。"""
    opp_id = payload.get('opportunity_id')
    conn = get_db()
    o = _first(conn.execute("SELECT * FROM plm_opportunity WHERE id=?", (opp_id,)))
    if not o:
        conn.close()
        return {'success': False, 'error': '商机不存在'}
    if o['status'] != '中标':
        conn.close()
        return {'success': False, 'error': '仅「中标」状态商机可联动立项，当前状态：%s' % o['status']}
    est = _first(conn.execute(
        "SELECT * FROM plm_baseline WHERE scope_type='opportunity' AND scope_id=? "
        "AND stage='estimate_bid' ORDER BY id DESC LIMIT 1", (opp_id,)))
    est_items = _rows(conn.execute("SELECT * FROM plm_baseline_item WHERE baseline_id=?",
                                   (est['id'],))) if est else []
    cp = payload.get('project') or {}
    contract_id = None
    contract_no = ''
    try:
        if payload.get('create_contract', True):
            ct = payload.get('contract') or {}
            cno = (ct.get('contract_no') or '').strip() or _next_no(
                conn, 'plm_contract', 'contract_no', 'HT')
            contract_no = cno
            cur = conn.execute("""INSERT INTO plm_contract
                (contract_no,contract_name,customer,industry,region,dept,sign_amount,sign_date,
                 project_cycle,status,owner,opportunity_id,remark) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cno, ct.get('contract_name') or o['opp_name'],
                 ct.get('customer') or o['customer'], ct.get('industry') or o['industry'],
                 ct.get('region') or o['region'], ct.get('dept') or o['dept'],
                 _f(ct.get('sign_amount', o['expect_income'])), ct.get('sign_date', ''),
                 ct.get('project_cycle', ''), ct.get('status', '已签署'),
                 ct.get('owner') or o['owner'], opp_id, ct.get('remark', '')))
            contract_id = cur.lastrowid
        project_no = (cp.get('project_no') or '').strip() or _next_no(
            conn, 'plm_project', 'project_no', 'XM')
        cur = conn.execute("""INSERT INTO plm_project
            (project_no,project_name,customer,dept,region,manager,status,contract_id,
             opportunity_id,start_date,end_date,kickoff_date,remark)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (project_no, cp.get('project_name') or o['opp_name'],
             cp.get('customer') or o['customer'], cp.get('dept') or o['dept'],
             cp.get('region') or o['region'], cp.get('manager') or o['owner'],
             cp.get('status', '待启动'), contract_id, opp_id, cp.get('start_date', ''),
             cp.get('end_date', ''), cp.get('kickoff_date') or _today(), cp.get('remark', '')))
        project_id = cur.lastrowid
        bid = None
        if est:
            cur = conn.execute("""INSERT INTO plm_baseline
                (scope_type,scope_id,stage,total_income,total_cost,gross,gross_rate,status,
                 source_baseline_id,created_by,remark,updated_at)
                VALUES ('project',?,'estimate_locked',?,?,?,?, '草稿',?,?,?,?)""",
                (project_id, est['total_income'], est['total_cost'], est['gross'],
                 est['gross_rate'], est['id'], operator,
                 '由商机投标概算自动带入：%s' % o['opp_no'], _now()))
            bid = cur.lastrowid
            for it in est_items:
                conn.execute("""INSERT INTO plm_baseline_item
                    (baseline_id,category,item_name,plan_amount,actual_amount,remark)
                    VALUES (?,?,?,?,?,?)""",
                    (bid, it['category'], it['item_name'], it['plan_amount'],
                     it['actual_amount'], it['remark']))
        ms_ids = []
        for m in (cp.get('milestones') or []):
            nm = (m.get('name') or '').strip()
            if not nm:
                continue
            cur = conn.execute("""INSERT INTO plm_milestone
                (project_id,parent_id,level,name,owner,plan_start,plan_end,is_key,deliverable,
                 status,remark) VALUES (?,NULL,'粗',?,?,?,?,?,?,'未开始','')""",
                (project_id, nm, m.get('owner', ''), m.get('plan_start', ''),
                 m.get('plan_end', ''), 1 if m.get('is_key') else 0, m.get('deliverable', '')))
            ms_ids.append((cur.lastrowid, nm))
        if payload.get('book_sign_income', True) and contract_id:
            amt = conn.execute("SELECT sign_amount a FROM plm_contract WHERE id=?",
                               (contract_id,)).fetchone()['a']
            if _f(amt) > 0:
                conn.execute("""INSERT INTO plm_ledger
                    (project_id,contract_id,kind,category,plan_or_actual,amount,occur_date,
                     source,remark) VALUES (?,?,'income','签单收入','实际',?,?, '手工录入',
                     '联动立项自动归集')""",
                    (project_id, contract_id, _f(amt), _today()))
        conn.commit()
    except sqlite3.IntegrityError as e:
        conn.close()
        return {'success': False, 'error': '编号冲突：%s' % e}
    conn.close()
    log_op('project', project_id, project_no, '商机联动立项',
           {'opportunity_no': o['opp_no'], 'contract_id': contract_id, 'baseline_id': bid},
           operator=operator)
    if bid:
        log_op('baseline', bid, BASELINE_STAGE_CN['estimate_locked'], '联动立项带入概算基线',
               {'source_opp_no': o['opp_no'], 'source_baseline_id': est['id'] if est else None},
               operator=operator)
    if contract_id:
        log_op('contract', contract_id, contract_no, '联动立项生成合同',
               {'opportunity_no': o['opp_no']}, operator=operator)
    for mid, nm in ms_ids:
        log_op('milestone', mid, nm, '立项创建粗里程碑', {'level': '粗'}, operator=operator)
    return {'success': True, 'project_id': project_id, 'project_no': project_no,
            'contract_id': contract_id, 'estimate_baseline_id': bid,
            'milestone_ids': [m[0] for m in ms_ids],
            'next': '请在「四算基线」确认概算并锁定，随后录入执行预算'}


# ===================== 模块三：里程碑与任务 =====================

def list_milestones(project_id):
    conn = get_db()
    rows = _rows(conn.execute("SELECT * FROM plm_milestone WHERE project_id=? "
                              "ORDER BY parent_id IS NOT NULL, plan_start, id", (project_id,)))
    today = _today()
    for r in rows:
        r['children'] = [x for x in rows if x.get('parent_id') == r['id']]
        r['task_count'] = conn.execute(
            "SELECT COUNT(*) n FROM plm_task WHERE milestone_id=?", (r['id'],)).fetchone()['n']
        r['is_overdue'] = bool(r['plan_end'] and r['status'] != '已完成'
                               and str(r['plan_end'])[:10] < today)
        r['overdue_days'] = (_days_between(str(r['plan_end'])[:10], today) or 0
                             if r['is_overdue'] else 0)
    conn.close()
    top = [r for r in rows if not r.get('parent_id')]
    for r in top:
        kids = r['children']
        if kids:
            r['progress_rollup'] = _r(sum(_f(k['progress']) for k in kids) / len(kids), 2)
    return rows


def list_all_milestones(keyword=None):
    """跨项目里程碑全量列表（含项目 / 合同关联），供「里程碑」视图使用。"""
    conn = get_db()
    sql = ("SELECT m.*, pp.project_no, pp.project_name, pp.contract_id, pc.contract_no "
           "FROM plm_milestone m "
           "LEFT JOIN plm_project pp ON pp.id = m.project_id "
           "LEFT JOIN plm_contract pc ON pc.id = pp.contract_id")
    conds, args = [], []
    if keyword:
        conds.append("(m.name LIKE ? OR pp.project_no LIKE ? OR pp.project_name LIKE ? "
                     "OR pc.contract_no LIKE ?)")
        args += ['%' + keyword + '%'] * 4
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY COALESCE(pp.project_no,''), m.project_id, m.plan_start, m.id"
    rows = _rows(conn.execute(sql, args))
    conn.close()
    return rows


# 「项目里程碑表」列 → plm_milestone 字段映射（按优先级依次尝试候选列名）
MILESTONE_IMPORT_FIELDS = {
    'name': ('任务名称', '合同里程碑名称', '任务名称'),
    'task_no': ('任务编号',),
    'owner': ('项目经理',),
    'plan_start': ('任务计划开始日期', '计划开始日期', '计划开始时间'),
    'plan_end': ('任务计划结束日期', '计划结束日期'),
    'actual_start': ('任务实际开始日期',),
    'actual_end': ('任务实际结束日期',),
    'progress': ('实际完成百分比',),
    'status': ('任务状态', '项目状态'),
    'plan_output': ('计划产值(元)', '里程碑产值(元)', '计划产值', '里程碑产值'),
    'plan_payback_date': ('计划回款时间',),
    'payback_date': ('回款时间',),
    'payback_amount': ('回款金额',),
}


def _pick_import(row, candidates):
    for c in candidates:
        v = row.get(c)
        if v is not None and str(v).strip() != '':
            return str(v).strip()
    return ''


def import_milestones(rows, operator=''):
    """按「项目里程碑表」行写入 plm_milestone。

    项目关联：用 core 主数据（core_project：project_no/contract_no）把
    行内「项目编号 / 合同编号」解析成项目号，再落到 plm_project 以取 project_id；
    无法匹配的项目行跳过并计数。
    返回 {'success','inserted','skipped','total','matched_columns'}。
    """
    from core import project as _core_proj
    # ① core 主数据索引：project_no 与 contract_no → 项目号（即 plm_project.project_no）
    core_idx = {}
    for cp in (_core_proj.list_projects() or []):
        pno = (cp.get('project_no') or '').strip()
        cno = (cp.get('contract_no') or '').strip()
        if pno:
            core_idx.setdefault(pno, pno)
        if cno:
            core_idx.setdefault(cno, pno)

    conn = get_db()
    try:
        # ② plm_project 索引：项目号 → id
        plm_proj = { (p['project_no'] or '').strip(): p['id']
                     for p in _rows(conn.execute("SELECT id, project_no FROM plm_project")) }
        inserted, skipped, matched_cols, op_logs = 0, 0, set(), []
        for rw in rows:
            if not rw:
                continue
            pno = _pick_import(rw, ('项目编号', 'project_no'))
            cno = _pick_import(rw, ('合同编号', 'contract_no'))
            matched_no = core_idx.get(pno) or core_idx.get(cno) or core_idx.get(pno or cno)
            if not matched_no:
                skipped += 1
                continue
            plm_id = plm_proj.get(matched_no)
            if not plm_id:
                # 主数据已存在但 plm_project 无该立项 → 自动登记，里程碑才有 project_id 可挂
                pname = _pick_import(rw, ('项目名称',)) or matched_no
                cur = conn.execute(
                    "INSERT INTO plm_project (project_no, project_name) VALUES (?,?)",
                    (matched_no, pname))
                plm_id = cur.lastrowid
                plm_proj[matched_no] = plm_id

            f = {field: _pick_import(rw, cands)
                 for field, cands in MILESTONE_IMPORT_FIELDS.items()}
            name = f['name']
            if not name:
                skipped += 1
                continue
            progress = None
            if f['progress']:
                try:
                    progress = float(f['progress'])
                    if progress > 1:
                        progress = progress / 100.0
                except (ValueError, TypeError):
                    progress = None
            amount = None
            if f['payback_amount']:
                try:
                    amount = round(float(str(f['payback_amount']).replace(',', '')), 2)
                except (ValueError, TypeError):
                    amount = None
            plan_output = None
            if f['plan_output']:
                plan_output = _f(str(f['plan_output']).replace(',', '')) or None

            cur = conn.execute(
                """INSERT INTO plm_milestone
                  (project_id, level, name, owner, task_no, plan_start, plan_end,
                   actual_start, actual_end, progress, status, plan_output,
                   plan_payback_date, payback_date, payback_amount, remark)
                  VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (plm_id, '细', name, f['owner'], f['task_no'], f['plan_start'],
                 f['plan_end'], f['actual_start'], f['actual_end'], progress,
                 f['status'] or '未开始', plan_output,
                 f['plan_payback_date'], f['payback_date'], amount, ''))
            inserted += 1
            matched_cols.update(field for field, v in f.items() if v)
            op_logs.append((cur.lastrowid, name, f['task_no']))
        conn.commit()
        # 操作日志在提交后写入（避免占用写锁）
        for mid, mname, tno in op_logs:
            log_op('milestone', mid, mname, '导入项目里程碑表',
                   {'source': 'xlsx', 'task_no': tno}, operator=operator)
        return {'success': True, 'inserted': inserted, 'skipped': skipped,
                'total': len(rows),
                'matched_columns': sorted(matched_cols)}
    finally:
        conn.close()


def create_milestone(payload, operator=''):
    name = (payload.get('name') or '').strip()
    pid = payload.get('project_id')
    if not pid or not name:
        return {'success': False, 'error': 'project_id 与 name 必填'}
    conn = get_db()
    if not _first(conn.execute("SELECT id FROM plm_project WHERE id=?", (pid,))):
        conn.close()
        return {'success': False, 'error': '项目不存在'}
    parent = payload.get('parent_id')
    if parent:
        pm = _first(conn.execute("SELECT project_id,level FROM plm_milestone WHERE id=?", (parent,)))
        if not pm:
            conn.close()
            return {'success': False, 'error': '父里程碑不存在'}
        if pm['project_id'] != pid:
            conn.close()
            return {'success': False, 'error': '子里程碑与父里程碑不属于同一项目'}
    level = payload.get('level') or ('细' if parent else '粗')
    if level not in MILESTONE_LEVEL:
        level = '细'
    cur = conn.execute("""INSERT INTO plm_milestone
        (project_id,parent_id,level,name,owner,plan_start,plan_end,actual_start,actual_end,
         progress,status,is_key,plan_output,deliverable,remark)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (pid, parent, level, name, payload.get('owner', ''), payload.get('plan_start', ''),
         payload.get('plan_end', ''), payload.get('actual_start', ''),
         payload.get('actual_end', ''), _f(payload.get('progress')),
         payload.get('status', '未开始'), 1 if payload.get('is_key') else 0,
         _f(payload.get('plan_output')), payload.get('deliverable', ''),
         payload.get('remark', '')))
    mid = cur.lastrowid
    conn.commit()
    conn.close()
    _rollup_parent(parent)
    log_op('milestone', mid, name, '新增里程碑', {'level': level}, operator=operator)
    return {'success': True, 'id': mid}


def update_milestone(mid, payload, operator=''):
    conn = get_db()
    old = _first(conn.execute("SELECT * FROM plm_milestone WHERE id=?", (mid,)))
    conn.close()
    if not old:
        return {'success': False, 'error': '里程碑不存在'}
    if 'is_key' in payload:
        payload = dict(payload)
        payload['is_key'] = 1 if payload['is_key'] in (1, True, '1', 'true') else 0
    fields = ('name', 'owner', 'plan_start', 'plan_end', 'actual_start', 'actual_end',
              'progress', 'status', 'is_key', 'plan_output', 'deliverable', 'remark', 'level')
    res = _update_by_fields('plm_milestone', 'milestone', mid, fields, payload,
                            numeric=('progress', 'plan_output'), label_field='name',
                            action='修改里程碑', operator=operator)
    if res.get('success'):
        _rollup_parent(old.get('parent_id'))
    return res


def delete_milestone(mid, operator=''):
    conn = get_db()
    m = _first(conn.execute("SELECT * FROM plm_milestone WHERE id=?", (mid,)))
    if not m:
        conn.close()
        return {'success': False, 'error': '里程碑不存在'}
    refs = {
        '子里程碑': conn.execute("SELECT COUNT(*) n FROM plm_milestone WHERE parent_id=?",
                               (mid,)).fetchone()['n'],
        '任务': conn.execute("SELECT COUNT(*) n FROM plm_task WHERE milestone_id=?",
                            (mid,)).fetchone()['n'],
    }
    if any(refs.values()):
        conn.close()
        return {'success': False, 'error': '里程碑下仍有子节点或任务，无法删除', 'refs': refs}
    conn.execute("DELETE FROM plm_milestone WHERE id=?", (mid,))
    conn.execute("DELETE FROM plm_assignment WHERE milestone_id=?", (mid,))
    conn.commit()
    parent = m.get('parent_id')
    conn.close()
    _rollup_parent(parent)
    log_op('milestone', mid, m['name'], '删除里程碑', operator=operator)
    return {'success': True}


def _rollup_parent(parent_id):
    """子级完成百分比与状态向上汇总。"""
    if not parent_id:
        return
    conn = get_db()
    kids = _rows(conn.execute("SELECT progress,status FROM plm_milestone WHERE parent_id=?",
                              (parent_id,)))
    if kids:
        avg = sum(_f(k['progress']) for k in kids) / len(kids)
        if all(k['status'] == '已完成' for k in kids):
            st = '已完成'
        elif any(_f(k['progress']) > 0 or k['status'] != '未开始' for k in kids):
            st = '进行中'
        else:
            st = '未开始'
        conn.execute("UPDATE plm_milestone SET progress=?,status=?,updated_at=? WHERE id=?",
                     (_r(avg, 4), st, _now(), parent_id))
        conn.commit()
    conn.close()


def list_tasks(project_id=None, milestone_id=None, owner=None, limit=1000):
    conn = get_db()
    sql, args, conds = "SELECT * FROM plm_task", [], []
    if project_id:
        conds.append("project_id=?"); args.append(project_id)
    if milestone_id:
        conds.append("milestone_id=?"); args.append(milestone_id)
    if owner:
        conds.append("owner=?"); args.append(owner)
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    mids = sorted({r['milestone_id'] for r in rows if r.get('milestone_id')})
    names = {}
    if mids:
        for r in conn.execute("SELECT id,name FROM plm_milestone WHERE id IN (%s)"
                              % ",".join("?" * len(mids)), mids):
            names[r['id']] = r['name']
    conn.close()
    today = _today()
    for r in rows:
        r['milestone_name'] = names.get(r.get('milestone_id'), '')
        r['is_overdue'] = bool(r['plan_end'] and r['status'] not in ('已完成', '已取消')
                               and str(r['plan_end'])[:10] < today)
        r['overdue_days'] = (_days_between(str(r['plan_end'])[:10], today) or 0
                             if r['is_overdue'] else 0)
    return rows


def get_task(task_id):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_task WHERE id=?", (task_id,)))
    conn.close()
    return r


def create_task(payload, operator=''):
    name = (payload.get('name') or '').strip()
    pid = payload.get('project_id')
    if not pid or not name:
        return {'success': False, 'error': 'project_id 与 name 必填'}
    conn = get_db()
    if payload.get('milestone_id'):
        m = _first(conn.execute("SELECT project_id FROM plm_milestone WHERE id=?",
                                (payload['milestone_id'],)))
        if not m:
            conn.close()
            return {'success': False, 'error': '所绑定的里程碑不存在'}
        if m['project_id'] != pid:
            conn.close()
            return {'success': False, 'error': '任务与里程碑不属于同一项目，无法绑定'}
    cur = conn.execute("""INSERT INTO plm_task
        (project_id,milestone_id,name,owner,plan_hours,actual_hours,progress,status,
         plan_start,plan_end,actual_end,deliverable,remark) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (pid, payload.get('milestone_id'), name, payload.get('owner', ''),
         _f(payload.get('plan_hours')), _f(payload.get('actual_hours')),
         _f(payload.get('progress')), payload.get('status', '未开始'),
         payload.get('plan_start', ''), payload.get('plan_end', ''),
         payload.get('actual_end', ''), payload.get('deliverable', ''),
         payload.get('remark', '')))
    tid = cur.lastrowid
    conn.commit()
    conn.close()
    log_op('task', tid, name, '新增任务', {'plan_hours': payload.get('plan_hours')},
           operator=operator)
    return {'success': True, 'id': tid}


def update_task(tid, payload, operator=''):
    conn = get_db()
    old = _first(conn.execute("SELECT * FROM plm_task WHERE id=?", (tid,)))
    conn.close()
    if not old:
        return {'success': False, 'error': '任务不存在'}
    fields = ('milestone_id', 'name', 'owner', 'plan_hours', 'actual_hours', 'progress',
              'status', 'plan_start', 'plan_end', 'actual_end', 'deliverable', 'remark')
    return _update_by_fields('plm_task', 'task', tid, fields, payload,
                             numeric=('plan_hours', 'actual_hours', 'progress'),
                             label_field='name', action='修改任务', operator=operator)


def delete_task(tid, operator=''):
    conn = get_db()
    t = _first(conn.execute("SELECT * FROM plm_task WHERE id=?", (tid,)))
    if not t:
        conn.close()
        return {'success': False, 'error': '任务不存在'}
    conn.execute("DELETE FROM plm_task WHERE id=?", (tid,))
    conn.execute("DELETE FROM plm_assignment WHERE task_id=?", (tid,))
    conn.commit()
    conn.close()
    log_op('task', tid, t['name'], '删除任务', operator=operator)
    return {'success': True}


# ===================== 模块四：人力池 / 分配 / 工时 =====================

def list_staff(keyword=None, status=None, limit=500):
    conn = get_db()
    sql, args, conds = "SELECT * FROM plm_staff", [], []
    if status and status != '全部':
        conds.append("status=?"); args.append(status)
    if keyword:
        conds.append("(name LIKE ? OR role LIKE ? OR dept LIKE ?)")
        args += ['%' + keyword + '%'] * 3
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    load = {x['staff_id']: x for x in staff_load()}
    for r in rows:
        r.update({k: v for k, v in load.get(r['id'], {}).items() if k != 'assignments'})
        for k in ('planned_hours', 'actual_hours', 'load_rate', 'load_state', 'parallel_projects'):
            r.setdefault(k, 0 if k.endswith('hours') or k == 'parallel_projects' else None)
        r.setdefault('load_state', '闲置')
    return rows


def get_staff(staff_id):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_staff WHERE id=?", (staff_id,)))
    if r:
        r['assignments'] = _rows(conn.execute(
            "SELECT * FROM plm_assignment WHERE staff_id=? ORDER BY id DESC", (staff_id,)))
        r['timesheets'] = _rows(conn.execute(
            "SELECT * FROM plm_timesheet WHERE staff_id=? ORDER BY work_date DESC LIMIT 200",
            (staff_id,)))
    conn.close()
    return r


def create_staff(payload, operator=''):
    name = (payload.get('name') or '').strip()
    if not name:
        return {'success': False, 'error': '人员姓名必填'}
    cost_rate = _f(payload.get('cost_rate'))
    avail = _f(payload.get('available_hours'), 160) or 160
    if cost_rate < 0 or avail <= 0:
        return {'success': False, 'error': '人力成本单价需 >= 0，可用工时需 > 0'}
    conn = get_db()
    try:
        cur = conn.execute("""INSERT INTO plm_staff
            (name,role,dept,cost_rate,available_hours,status,skills,remark) VALUES (?,?,?,?,?,?,?,?)""",
            (name, payload.get('role', ''), payload.get('dept', ''), cost_rate, avail,
             payload.get('status', '可用'), payload.get('skills', ''), payload.get('remark', '')))
        sid = cur.lastrowid
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {'success': False, 'error': '人员已存在：%s' % name}
    conn.close()
    log_op('staff', sid, name, '新增人员', {'cost_rate': cost_rate}, operator=operator)
    return {'success': True, 'id': sid}


def update_staff(sid, payload, operator=''):
    conn = get_db()
    old = _first(conn.execute("SELECT project_id FROM plm_timesheet WHERE staff_id=? "
                              "GROUP BY project_id", (sid,)))
    pids = [r['project_id'] for r in conn.execute(
        "SELECT DISTINCT project_id FROM plm_timesheet WHERE staff_id=?", (sid,)).fetchall()]
    conn.close()
    fields = ('name', 'role', 'dept', 'cost_rate', 'available_hours', 'status', 'skills',
              'efficiency_bonus', 'revenue_per_day', 'remark')
    res = _update_by_fields('plm_staff', 'staff', sid, fields, payload,
                            numeric=('cost_rate', 'available_hours', 'efficiency_bonus',
                                     'revenue_per_day'),
                            label_field='name', action='修改人员', operator=operator)
    if res.get('success'):
        for pid in pids:
            sync_labor_cost(project_id=pid)
    return res


def delete_staff(sid, operator=''):
    conn = get_db()
    s = _first(conn.execute("SELECT * FROM plm_staff WHERE id=?", (sid,)))
    if not s:
        conn.close()
        return {'success': False, 'error': '人员不存在'}
    refs = {
        '分配': conn.execute("SELECT COUNT(*) n FROM plm_assignment WHERE staff_id=?",
                            (sid,)).fetchone()['n'],
        '工时': conn.execute("SELECT COUNT(*) n FROM plm_timesheet WHERE staff_id=?",
                            (sid,)).fetchone()['n'],
    }
    if any(refs.values()):
        conn.close()
        return {'success': False, 'error': '该人员仍有分配或工时记录，无法删除', 'refs': refs}
    conn.execute("DELETE FROM plm_staff WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    log_op('staff', sid, s['name'], '删除人员', operator=operator)
    return {'success': True}


def list_assignments(project_id=None, staff_id=None, status=None, limit=1000):
    conn = get_db()
    sql = """SELECT a.*, s.name staff_name, s.role staff_role, s.cost_rate, s.available_hours,
                    p.project_no, p.project_name, m.name milestone_name, t.name task_name
             FROM plm_assignment a
             LEFT JOIN plm_staff s ON s.id=a.staff_id
             LEFT JOIN plm_project p ON p.id=a.project_id
             LEFT JOIN plm_milestone m ON m.id=a.milestone_id
             LEFT JOIN plm_task t ON t.id=a.task_id"""
    args, conds = [], []
    if project_id:
        conds.append("a.project_id=?"); args.append(project_id)
    if staff_id:
        conds.append("a.staff_id=?"); args.append(staff_id)
    if status and status != '全部':
        conds.append("a.status=?"); args.append(status)
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY a.id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    return rows


def create_assignment(payload, operator=''):
    sid, pid = payload.get('staff_id'), payload.get('project_id')
    if not sid or not pid:
        return {'success': False, 'error': 'staff_id 与 project_id 必填'}
    conn = get_db()
    if not _first(conn.execute("SELECT id FROM plm_staff WHERE id=?", (sid,))):
        conn.close()
        return {'success': False, 'error': '人员不存在'}
    if not _first(conn.execute("SELECT id FROM plm_project WHERE id=?", (pid,))):
        conn.close()
        return {'success': False, 'error': '项目不存在'}
    if payload.get('task_id'):
        t = _first(conn.execute("SELECT project_id FROM plm_task WHERE id=?", (payload['task_id'],)))
        if not t:
            conn.close()
            return {'success': False, 'error': '任务不存在'}
        if t['project_id'] != pid:
            conn.close()
            return {'success': False, 'error': '任务与项目不一致'}
    if payload.get('milestone_id'):
        m = _first(conn.execute("SELECT project_id FROM plm_milestone WHERE id=?",
                                (payload['milestone_id'],)))
        if not m or m['project_id'] != pid:
            conn.close()
            return {'success': False, 'error': '里程碑不属于该项目'}
    cur = conn.execute("""INSERT INTO plm_assignment
        (staff_id,project_id,milestone_id,task_id,role_in_proj,planned_hours,start_date,
         end_date,status,remark) VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (sid, pid, payload.get('milestone_id'), payload.get('task_id'),
         payload.get('role_in_proj', ''), _f(payload.get('planned_hours')),
         payload.get('start_date', ''), payload.get('end_date', ''),
         payload.get('status', '生效中'), payload.get('remark', '')))
    aid = cur.lastrowid
    conn.commit()
    conn.close()
    log_op('assignment', aid, 'staff#%s → project#%s' % (sid, pid), '新增人员分配',
           {'planned_hours': payload.get('planned_hours')}, operator=operator)
    return {'success': True, 'id': aid}


def update_assignment(aid, payload, operator=''):
    fields = ('milestone_id', 'task_id', 'role_in_proj', 'planned_hours', 'start_date',
              'end_date', 'status', 'remark')
    return _update_by_fields('plm_assignment', 'assignment', aid, fields, payload,
                             numeric=('planned_hours',), label_field='id',
                             action='修改人员分配', operator=operator)


def delete_assignment(aid, operator=''):
    conn = get_db()
    a = _first(conn.execute("SELECT * FROM plm_assignment WHERE id=?", (aid,)))
    if not a:
        conn.close()
        return {'success': False, 'error': '分配记录不存在'}
    conn.execute("DELETE FROM plm_assignment WHERE id=?", (aid,))
    conn.commit()
    conn.close()
    log_op('assignment', aid, str(aid), '删除人员分配', operator=operator)
    return {'success': True}


def staff_load():
    """
    FR-6：人员负荷 = Σ生效中计划工时 ÷ 可用工时。
    三态：闲置（无任何分配与工时）/ 过载（> 系数，默认 1.2）/ 正常。
    """
    factor = _f(get_config('alert_staff_overload', '1.2'), 1.2) or 1.2
    conn = get_db()
    staff = _rows(conn.execute("SELECT * FROM plm_staff WHERE status<>'离职' ORDER BY id"))
    out = []
    for s in staff:
        asg = _rows(conn.execute(
            "SELECT a.*, p.project_no, p.project_name FROM plm_assignment a "
            "LEFT JOIN plm_project p ON p.id=a.project_id "
            "WHERE a.staff_id=? AND a.status='生效中' ORDER BY a.id", (s['id'],)))
        planned = round(sum(_f(a['planned_hours']) for a in asg), 4)
        actual = _f(conn.execute("SELECT COALESCE(SUM(hours),0) h FROM plm_timesheet "
                                 "WHERE staff_id=?", (s['id'],)).fetchone()['h'])
        avail = _f(s['available_hours'])
        rate = _rate(planned, avail)
        if planned <= 0 and actual <= 0:
            state = '闲置'
        elif rate is not None and rate > factor:
            state = '过载'
        else:
            state = '正常'
        seen, projs = set(), []
        for a in asg:
            key = a.get('project_id')
            if key and key not in seen:
                seen.add(key)
                projs.append({'project_id': key, 'project_no': a.get('project_no') or '',
                              'project_name': a.get('project_name') or '',
                              'hours': _f(a['planned_hours'])})
        for pr in projs:
            pr['hours'] = round(sum(_f(a['planned_hours']) for a in asg
                                    if a.get('project_id') == pr['project_id']), 2)
        out.append({'staff_id': s['id'], 'name': s['name'], 'role': s['role'],
                    'dept': s['dept'], 'cost_rate': _f(s['cost_rate']),
                    'available_hours': avail, 'planned_hours': _r(planned),
                    'actual_hours': _r(actual), 'load_rate': rate, 'load_state': state,
                    'parallel_projects': len(projs), 'projects': projs, 'assignments': asg})
    conn.close()
    return out


def list_timesheets(project_id=None, staff_id=None, limit=1000):
    conn = get_db()
    sql = """SELECT t.*, s.name staff_name, s.role staff_role, s.cost_rate,
                    p.project_no, p.project_name, tk.name task_name
             FROM plm_timesheet t
             LEFT JOIN plm_staff s ON s.id=t.staff_id
             LEFT JOIN plm_project p ON p.id=t.project_id
             LEFT JOIN plm_task tk ON tk.id=t.task_id"""
    args, conds = [], []
    if project_id:
        conds.append("t.project_id=?"); args.append(project_id)
    if staff_id:
        conds.append("t.staff_id=?"); args.append(staff_id)
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY t.work_date DESC, t.id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    return rows


def create_timesheet(payload, operator=''):
    sid, pid = payload.get('staff_id'), payload.get('project_id')
    hours = _f(payload.get('hours'))
    if not sid or not pid:
        return {'success': False, 'error': 'staff_id 与 project_id 必填'}
    if hours <= 0:
        return {'success': False, 'error': '工时必须大于 0'}
    conn = get_db()
    if not _first(conn.execute("SELECT id FROM plm_staff WHERE id=?", (sid,))):
        conn.close()
        return {'success': False, 'error': '人员不存在'}
    if not _first(conn.execute("SELECT id FROM plm_project WHERE id=?", (pid,))):
        conn.close()
        return {'success': False, 'error': '项目不存在'}
    if payload.get('task_id') and not _first(
            conn.execute("SELECT id FROM plm_task WHERE id=?", (payload['task_id'],))):
        conn.close()
        return {'success': False, 'error': '任务不存在'}
    cur = conn.execute("""INSERT INTO plm_timesheet
        (staff_id,project_id,task_id,work_date,hours,remark) VALUES (?,?,?,?,?,?)""",
        (sid, pid, payload.get('task_id'), payload.get('work_date') or _today(), hours,
         payload.get('remark', '')))
    ts_id = cur.lastrowid
    if payload.get('task_id'):
        conn.execute("UPDATE plm_task SET actual_hours=COALESCE(actual_hours,0)+? WHERE id=?",
                     (hours, payload['task_id']))
    conn.commit()
    conn.close()
    sync_labor_cost(project_id=pid)
    log_op('timesheet', ts_id, 'staff#%s project#%s' % (sid, pid), '填报工时',
           {'hours': hours}, operator=operator)
    return {'success': True, 'id': ts_id}


def update_timesheet(ts_id, payload, operator=''):
    conn = get_db()
    old = _first(conn.execute("SELECT * FROM plm_timesheet WHERE id=?", (ts_id,)))
    conn.close()
    if not old:
        return {'success': False, 'error': '工时记录不存在'}
    res = _update_by_fields('plm_timesheet', 'timesheet', ts_id,
                            ('task_id', 'work_date', 'hours', 'remark'), payload,
                            numeric=('hours',), label_field='id', action='修改工时',
                            operator=operator)
    if res.get('success'):
        sync_labor_cost(project_id=old['project_id'])
    return res


def delete_timesheet(ts_id, operator=''):
    conn = get_db()
    t = _first(conn.execute("SELECT * FROM plm_timesheet WHERE id=?", (ts_id,)))
    if not t:
        conn.close()
        return {'success': False, 'error': '工时记录不存在'}
    conn.execute("DELETE FROM plm_timesheet WHERE id=?", (ts_id,))
    if t.get('task_id'):
        conn.execute("UPDATE plm_task SET actual_hours=MAX(0,COALESCE(actual_hours,0)-?) WHERE id=?",
                     (_f(t['hours']), t['task_id']))
    conn.commit()
    pid = t['project_id']
    conn.close()
    sync_labor_cost(project_id=pid)
    log_op('timesheet', ts_id, str(ts_id), '删除工时', operator=operator)
    return {'success': True}


def sync_labor_cost(project_id=None, staff_id=None):
    """
    FR-6：把工时折算为实际人力成本并归集到台账（source='工时归集'）。
      金额 = Σhours ÷ labor_day_hours × staff.cost_rate
    以 ref_id='agg:<staff>:<project>' 幂等覆写；手工录入的成本记录不受影响。
    """
    day_hours = _f(get_config('labor_day_hours', '8'), 8) or 8
    conn = get_db()
    where, args = [], []
    if project_id:
        where.append("t.project_id=?"); args.append(project_id)
    if staff_id:
        where.append("t.staff_id=?"); args.append(staff_id)
    wsql = (" WHERE " + " AND ".join(where)) if where else ""
    agg = _rows(conn.execute(
        "SELECT t.project_id pid, t.staff_id sid, SUM(t.hours) h FROM plm_timesheet t%s "
        "GROUP BY t.project_id, t.staff_id" % wsql, args))
    touched = set()
    for row in agg:
        pid, sid, hours = int(row['pid']), int(row['sid']), _f(row['h'])
        s = _first(conn.execute("SELECT cost_rate,name FROM plm_staff WHERE id=?", (sid,)))
        if not s:
            continue
        amount = _r(hours / day_hours * _f(s['cost_rate']))
        ref = 'agg:%s:%s' % (sid, pid)
        note = '%s 工时归集（%.1f 小时）' % (s['name'], hours)
        exist = _first(conn.execute("SELECT id FROM plm_ledger WHERE ref_type='timesheet_agg' "
                                    "AND ref_id=?", (ref,)))
        if exist:
            conn.execute("UPDATE plm_ledger SET amount=?,remark=?,occur_date=? WHERE id=?",
                         (amount, note, _today(), exist['id']))
        else:
            conn.execute("""INSERT INTO plm_ledger
                (project_id,kind,category,plan_or_actual,amount,occur_date,source,ref_type,
                 ref_id,remark) VALUES (?,'cost','人力成本','实际',?,?, '工时归集',
                 'timesheet_agg',?,?)""", (pid, amount, _today(), ref, note))
        touched.add((sid, pid))
    stale_sql = "SELECT id,ref_id FROM plm_ledger WHERE ref_type='timesheet_agg'"
    stale_args = []
    if project_id:
        stale_sql += " AND project_id=?"; stale_args.append(project_id)
    if staff_id:
        stale_sql += " AND ref_id LIKE ?"; stale_args.append('agg:%s:%%' % staff_id)
    for row in _rows(conn.execute(stale_sql, stale_args)):
        parts = str(row['ref_id']).split(':')
        if len(parts) != 3:
            continue
        try:
            key = (int(parts[1]), int(parts[2]))
        except ValueError:
            continue
        if key not in touched:
            conn.execute("DELETE FROM plm_ledger WHERE id=?", (row['id'],))
    conn.commit()
    conn.close()
    return {'success': True, 'synced': len(touched)}


# ===================== 模块五：收支台账与毛利 =====================

def list_ledger(project_id=None, kind=None, category=None, source=None, limit=2000):
    conn = get_db()
    sql, args, conds = "SELECT * FROM plm_ledger", [], []
    if project_id:
        conds.append("project_id=?"); args.append(project_id)
    if kind and kind != '全部':
        conds.append("kind=?"); args.append(kind)
    if category and category != '全部':
        conds.append("category=?"); args.append(category)
    if source and source != '全部':
        conds.append("source=?"); args.append(source)
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += " ORDER BY id DESC LIMIT ?"; args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    return rows


def create_ledger(payload, operator=''):
    pid = payload.get('project_id')
    kind = payload.get('kind')
    if not pid or kind not in ('income', 'cost'):
        return {'success': False, 'error': 'project_id 必填，kind 需为 income / cost'}
    conn = get_db()
    if not _first(conn.execute("SELECT id FROM plm_project WHERE id=?", (pid,))):
        conn.close()
        return {'success': False, 'error': '项目不存在'}
    cur = conn.execute("""INSERT INTO plm_ledger
        (project_id,contract_id,kind,category,plan_or_actual,amount,occur_date,source,
         milestone_id,remark) VALUES (?,?,?,?,?,?,?, '手工录入',?,?)""",
        (pid, payload.get('contract_id'), kind, payload.get('category', ''),
         payload.get('plan_or_actual', '实际'), _f(payload.get('amount')),
         payload.get('occur_date') or _today(), payload.get('milestone_id'),
         payload.get('remark', '')))
    lid = cur.lastrowid
    conn.commit()
    conn.close()
    log_op('ledger', lid, payload.get('category') or kind, '新增台账记录',
           {'amount': payload.get('amount')}, operator=operator)
    return {'success': True, 'id': lid}


def update_ledger(lid, payload, operator=''):
    conn = get_db()
    old = _first(conn.execute("SELECT * FROM plm_ledger WHERE id=?", (lid,)))
    conn.close()
    if not old:
        return {'success': False, 'error': '台账记录不存在'}
    if old['source'] == '工时归集':
        return {'success': False, 'error': '工时归集记录由系统自动维护，请改工时而非台账'}
    return _update_by_fields('plm_ledger', 'ledger', lid,
                             ('contract_id', 'kind', 'category', 'plan_or_actual', 'amount',
                              'occur_date', 'milestone_id', 'remark'), payload,
                             numeric=('amount',), label_field='id', action='修改台账',
                             operator=operator)


def delete_ledger(lid, operator=''):
    conn = get_db()
    r = _first(conn.execute("SELECT * FROM plm_ledger WHERE id=?", (lid,)))
    if not r:
        conn.close()
        return {'success': False, 'error': '台账记录不存在'}
    if r['source'] == '工时归集':
        conn.close()
        return {'success': False, 'error': '工时归集记录由系统自动维护，不可手工删除'}
    conn.execute("DELETE FROM plm_ledger WHERE id=?", (lid,))
    conn.commit()
    conn.close()
    log_op('ledger', lid, str(r['category']), '删除台账记录', operator=operator)
    return {'success': True}


def project_finance(project_id):
    """FR-7：收入、成本、毛利四指标 + 概算-预算-实际三线差异。"""
    conn = get_db()

    def s(where, extra=None):
        return _f(conn.execute("SELECT COALESCE(SUM(amount),0) a FROM plm_ledger WHERE "
                               "project_id=? AND " + where,
                               [project_id] + (extra or [])).fetchone()['a'])
    income_signed = s("kind='income' AND category='签单收入'")
    income_change = s("kind='income' AND category='变更收入'")
    income_other = s("kind='income' AND category NOT IN ('签单收入','变更收入')")
    income_total = income_signed + income_change + income_other
    cost_est = s("kind='cost' AND plan_or_actual='预估'")
    cost_actual = s("kind='cost' AND plan_or_actual='实际'")
    labor_auto = s("kind='cost' AND source='工时归集'")
    hours_total = _f(conn.execute("SELECT COALESCE(SUM(hours),0) h FROM plm_timesheet "
                                  "WHERE project_id=?", (project_id,)).fetchone()['h'])
    est = _first(conn.execute(
        "SELECT * FROM plm_baseline WHERE scope_type='project' AND scope_id=? "
        "AND stage IN ('estimate_locked','estimate_bid') "
        "ORDER BY CASE stage WHEN 'estimate_locked' THEN 0 ELSE 1 END, id DESC LIMIT 1",
        (project_id,)))
    bud = _first(conn.execute(
        "SELECT * FROM plm_baseline WHERE scope_type='project' AND scope_id=? AND stage='budget' "
        "ORDER BY id DESC LIMIT 1", (project_id,)))
    ct = _f(conn.execute("SELECT COALESCE(SUM(sign_amount),0) a FROM plm_contract WHERE id IN "
                         "(SELECT contract_id FROM plm_project WHERE id=?)",
                         (project_id,)).fetchone()['a'])
    conn.close()
    estimate_cost = _f(est['total_cost']) if est else None
    budget_total = _f(bud['total_cost']) if bud else None
    baseline_cost = estimate_cost if estimate_cost is not None else cost_est
    signed_gross = _r(income_signed - baseline_cost)
    est_gross = _r(income_total - baseline_cost)
    actual_gross = _r(income_total - cost_actual)
    usage = _rate(cost_actual, budget_total)
    return {
        'income': {'signed': _r(income_signed), 'change': _r(income_change),
                   'other': _r(income_other), 'total': _r(income_total),
                   'contract_sign_amount': _r(ct)},
        'cost': {'estimate': _r(cost_est), 'actual_cum': _r(cost_actual),
                 'labor_auto': _r(labor_auto), 'hours_total': _r(hours_total)},
        'baseline': {'estimate_cost': _r(estimate_cost), 'budget_total': _r(budget_total)},
        'gross': {'signed': signed_gross, 'estimate': est_gross, 'actual': actual_gross,
                  'signed_rate': _rate(signed_gross, income_signed),
                  'estimate_rate': _rate(est_gross, income_total),
                  'actual_rate': _rate(actual_gross, income_total)},
        'variance': {
            'estimate_vs_budget': (_r(budget_total - estimate_cost)
                                   if (budget_total is not None and estimate_cost is not None)
                                   else None),
            'budget_vs_actual': (_r(cost_actual - budget_total)
                                 if budget_total is not None else None),
            'estimate_vs_actual': (_r(cost_actual - estimate_cost)
                                   if estimate_cost is not None else None),
            'budget_usage_rate': usage,
            'estimate_usage_rate': _rate(cost_actual, estimate_cost),
            'direction': ('超支' if (budget_total is not None and cost_actual > budget_total)
                          else '节支' if (budget_total is not None and cost_actual < budget_total)
                          else '持平'),
        },
        'reserved': {'accounting': None, 'final': None, 'note': RESERVED_NOTE},
    }


# ===================== 模块三：双维度进度 =====================

def project_progress(project_id):
    """FR-5：按期进度 + 按预算进度。"""
    conn = get_db()
    p = _first(conn.execute("SELECT * FROM plm_project WHERE id=?", (project_id,)))
    ms = _rows(conn.execute("SELECT * FROM plm_milestone WHERE project_id=?", (project_id,)))
    tk = _rows(conn.execute("SELECT * FROM plm_task WHERE project_id=?", (project_id,)))
    conn.close()
    today = _today()
    fine = [m for m in ms if m['level'] == '细'] or ms
    done = [m for m in fine if m['status'] == '已完成']
    on_time = [m for m in done if m['actual_end'] and m['plan_end']
               and str(m['actual_end'])[:10] <= str(m['plan_end'])[:10]]
    overdue_nodes, max_overdue = [], 0
    for m in fine:
        if m['status'] != '已完成' and m['plan_end'] and str(m['plan_end'])[:10] < today:
            d = _days_between(str(m['plan_end'])[:10], today) or 0
            max_overdue = max(max_overdue, d)
            overdue_nodes.append({'type': 'milestone', 'id': m['id'], 'name': m['name'],
                                  'plan_end': m['plan_end'], 'overdue_days': d,
                                  'is_key': bool(m['is_key']), 'owner': m['owner']})
    for t in tk:
        if t['status'] not in ('已完成', '已取消') and t['plan_end'] and str(t['plan_end'])[:10] < today:
            d = _days_between(str(t['plan_end'])[:10], today) or 0
            max_overdue = max(max_overdue, d)
            overdue_nodes.append({'type': 'task', 'id': t['id'], 'name': t['name'],
                                  'plan_end': t['plan_end'], 'overdue_days': d,
                                  'is_key': False, 'owner': t['owner']})
    overdue_nodes.sort(key=lambda x: -x['overdue_days'])
    if fine and all(_f(m['plan_output']) > 0 for m in fine):
        total_out = sum(_f(m['plan_output']) for m in fine)
        progress_rate = (round(sum(_f(m['plan_output']) * _f(m['progress']) / 100.0
                                   for m in fine) / total_out, 6)
                         if total_out > 0 else None)
        caliber = '计划产值加权'
    else:
        progress_rate = (round(sum(_f(m['progress']) for m in fine) / len(fine) / 100.0, 6)
                         if fine else None)
        caliber = '节点等权平均'
    tk_done = [t for t in tk if t['status'] == '已完成']
    tk_on_time = [t for t in tk_done if t['actual_end'] and t['plan_end']
                  and str(t['actual_end'])[:10] <= str(t['plan_end'])[:10]]
    fin = project_finance(project_id)
    usage = fin['variance']['budget_usage_rate']
    budget_total = fin['baseline']['budget_total']
    actual_cum = _f(fin['cost']['actual_cum'])
    over_budget_nodes = []
    if budget_total and usage and usage > 1.0:
        over_budget_nodes.append({'name': '项目整体', 'amount': _r(actual_cum - budget_total)})
    return {
        'project': {'id': project_id,
                    'project_no': p['project_no'] if p else '',
                    'project_name': p['project_name'] if p else '',
                    'status': p['status'] if p else ''},
        'schedule': {
            'milestone_total': len(fine), 'milestone_done': len(done),
            'milestone_overdue': len([n for n in overdue_nodes if n['type'] == 'milestone']),
            'on_time_rate': _rate(len(on_time), len(done)) if done else None,
            'progress_rate': progress_rate, 'progress_caliber': caliber,
            'task_total': len(tk), 'task_done': len(tk_done),
            'task_overdue': len([n for n in overdue_nodes if n['type'] == 'task']),
            'task_on_time_rate': _rate(len(tk_on_time), len(tk_done)) if tk_done else None,
            'max_overdue_days': max_overdue,
            'overdue_nodes': overdue_nodes,
        },
        'budget': {
            'estimate_total': fin['baseline']['estimate_cost'],
            'budget_total': _r(budget_total) if budget_total is not None else None,
            'actual_cum': _r(actual_cum),
            'budget_usage_rate': usage,
            'remaining': _r(budget_total - actual_cum) if budget_total is not None else None,
            'estimate_vs_budget_diff': fin['variance']['estimate_vs_budget'],
            'over_budget_nodes': over_budget_nodes,
            'time_vs_cost_gap': (_r((usage or 0) - (progress_rate or 0), 6)
                                 if usage is not None and progress_rate is not None else None),
            'note': '' if budget_total else '未录入执行预算',
        },
    }


# ===================== 模块六：项目全景视图 =====================

def project_panorama(project_id):
    """FR-8：固定 7 板块聚合；空数据返回空集合。"""
    p = get_project(project_id)
    if not p:
        return None
    bases = compare_baselines(project_id)
    prog = project_progress(project_id)
    fin = project_finance(project_id)
    asg = list_assignments(project_id=project_id)
    ts = list_timesheets(project_id=project_id, limit=200)
    ms = list_milestones(project_id)
    alerts = list_alerts(project_id=project_id, status='全部')
    assigned_staff = sorted({a['staff_id'] for a in asg if a.get('staff_id')})
    load_all = {x['staff_id']: x for x in staff_load()}
    participants = []
    for sid in assigned_staff:
        x = load_all.get(sid)
        if not x:
            continue
        mine = [a for a in x['assignments'] if a['project_id'] == project_id]
        participants.append({
            'staff_id': sid, 'name': x['name'], 'role': x['role'],
            'project_hours': round(sum(_f(a['planned_hours']) for a in mine), 2),
            'planned_hours': x['planned_hours'], 'actual_hours': x['actual_hours'],
            'available_hours': x['available_hours'],
            'load_state': x['load_state'], 'load_rate': x['load_rate'],
            'parallel_projects': x['parallel_projects']})
    hours_total = fin['cost']['hours_total']
    labor_cost = fin['cost']['labor_auto']
    opp = p.get('opportunity') or {}
    ct = p.get('contract') or {}
    return {
        'base_info': {
            'project_id': project_id, 'project_no': p['project_no'],
            'project_name': p['project_name'], 'customer': p['customer'],
            'contract_id': p.get('contract_id'), 'contract_no': ct.get('contract_no', ''),
            'contract_name': ct.get('contract_name', ''),
            'sign_amount': ct.get('sign_amount'),
            'opportunity_id': p.get('opportunity_id'), 'opportunity_no': opp.get('opp_no', ''),
            'opportunity_name': opp.get('opp_name', ''),
            'manager': p['manager'], 'dept': p['dept'], 'region': p['region'],
            'status': p['status'],
            'period': '%s ~ %s' % (p['start_date'] or '-', p['end_date'] or '-'),
            'kickoff_date': p['kickoff_date'], 'remark': p['remark'],
        },
        'baseline_area': {
            'estimate': bases['estimate'], 'budget': bases['budget'],
            'accounting': bases['accounting'], 'final': bases['final'],
            'estimate_vs_budget': bases['estimate_vs_budget'],
            'budget_usage_note': bases['budget_usage_note'],
            'estimate_items': (bases['estimate'] or {}).get('items', []),
            'budget_items': (bases['budget'] or {}).get('items', []),
            'note': '核算 / 决算为预留字段，本期不参与计算',
        },
        'pmo_area': {
            'schedule': prog['schedule'], 'budget': prog['budget'],
            'milestones': [{'id': m['id'], 'name': m['name'], 'level': m['level'],
                            'parent_id': m.get('parent_id'), 'plan_start': m['plan_start'],
                            'plan_end': m['plan_end'], 'actual_end': m['actual_end'],
                            'progress': m['progress'], 'status': m['status'],
                            'is_key': bool(m['is_key']), 'owner': m['owner'],
                            'is_overdue': m['is_overdue'], 'task_count': m['task_count']}
                           for m in ms],
        },
        'hr_area': {
            'participants': participants,
            'assignment_count': len(asg),
            'timesheets': ts[:100],
            'hours_total': hours_total,
            'labor_cost': labor_cost,
            'efficiency': {'note': '人效 / 元效为预留数据源，本期不计算',
                           'revenue_per_hour': _rate(_f(fin['income']['total']), hours_total),
                           'cost_per_hour': _rate(labor_cost, hours_total)},
        },
        'finance_area': fin,
        'alert_area': {
            'total': len([a for a in alerts if a['status'] != '已闭环']),
            'pending': len([a for a in alerts if a['status'] == '待处理']),
            'processing': len([a for a in alerts if a['status'] == '处理中']),
            'items': alerts,
        },
        'quick_links': [
            {'label': '商机档案', 'target': 'opportunity', 'id': p.get('opportunity_id')},
            {'label': '合同资料', 'target': 'contract', 'id': p.get('contract_id')},
            {'label': '任务清单', 'target': 'pmo', 'id': project_id},
            {'label': '工时明细', 'target': 'labor', 'id': project_id},
            {'label': '报表导出', 'target': 'export', 'id': project_id},
        ],
    }


# ===================== 模块七：风险预警 =====================

def list_alert_rules():
    conn = get_db()
    rows = _rows(conn.execute("SELECT * FROM plm_alert_rule ORDER BY dim,threshold"))
    conn.close()
    for r in rows:
        r['enabled'] = bool(r['enabled'])
    return rows


def update_alert_rule(rule_key, payload, operator=''):
    conn = get_db()
    old = _first(conn.execute("SELECT * FROM plm_alert_rule WHERE rule_key=?", (rule_key,)))
    if not old:
        conn.close()
        return {'success': False, 'error': '预警规则不存在'}
    sets, args, changed = [], [], {}
    for k in ('threshold', 'level', 'enabled', 'description', 'rule_name', 'op'):
        if k in payload:
            v = payload[k]
            if k == 'enabled':
                v = 1 if v in (1, True, '1', 'true', 'on') else 0
            if k == 'threshold':
                v = _f(v)
            sets.append("%s=?" % k); args.append(v)
            changed[k] = {'before': old.get(k), 'after': v}
    if not sets:
        conn.close()
        return {'success': False, 'error': '无可更新字段'}
    conn.execute("UPDATE plm_alert_rule SET %s, updated_at=? WHERE rule_key=?" % ",".join(sets),
                 args + [_now(), rule_key])
    conn.commit()
    conn.close()
    log_op('alert_rule', rule_key, old['rule_name'], '调整预警规则', changed, operator=operator)
    return {'success': True}


def _cmp(value, op, threshold):
    if value is None:
        return False
    try:
        v, t = float(value), float(threshold)
    except (TypeError, ValueError):
        return False
    return {'>': v > t, '>=': v >= t, '<': v < t, '<=': v <= t, '=': v == t}.get(op, False)


def _fmt_val(dim, v):
    if v is None:
        return '-'
    if dim in ('cost', 'gross', 'staff'):
        return '%.1f%%' % (float(v) * 100)
    return '%d 天' % int(float(v))


def _alert_text(p, ru, val):
    shown, thr = _fmt_val(ru['dim'], val), _fmt_val(ru['dim'], ru['threshold'])
    title = '%s：%s（当前 %s / 阈值 %s %s）' % (p['project_no'], ru['rule_name'], shown,
                                              ru['op'], thr)
    detail = '项目 %s 触发【%s】，当前值 %s，判定条件 %s %s。%s' % (
        p['project_name'], ru['rule_name'], shown, ru['op'], thr, ru['description'] or '')
    return title, detail


def project_alert_metrics(project_id, staff_rows=None):
    """预警判定所需指标集合。"""
    prog = project_progress(project_id)
    fin = project_finance(project_id)
    if staff_rows is None:
        staff_rows = staff_load()
    rates = [x['load_rate'] for x in staff_rows
             if any(a['project_id'] == project_id for a in x.get('assignments', []))
             and x.get('load_rate') is not None]
    return {
        'budget_usage_rate': prog['budget']['budget_usage_rate'],
        'actual_gross_rate': fin['gross']['actual_rate'],
        'max_overdue_days': prog['schedule']['max_overdue_days'],
        'max_load_rate': (max(rates) if rates else None),
    }


def scan_alerts(operator='system'):
    """FR-9：按启用规则扫描全部项目；同项目同规则未闭环只保留一条；风险消除自动闭环。"""
    conn = get_db()
    projects = _rows(conn.execute("SELECT * FROM plm_project"))
    rules = _rows(conn.execute("SELECT * FROM plm_alert_rule WHERE enabled=1"))
    conn.close()
    staff_rows = staff_load()
    fired = set()
    created, updated, auto_closed = [], 0, []
    conn = get_db()
    for p in projects:
        closed_proj = (p['status'] == '结项')
        metrics = project_alert_metrics(p['id'], staff_rows)
        for ru in rules:
            if closed_proj and ru['dim'] in ('cost', 'schedule'):
                continue
            if ru['dim'] == 'staff':
                continue  # 人员维度单独按人登记
            val = metrics.get(ru['metric'])
            if val is None or not _cmp(val, ru['op'], ru['threshold']):
                continue
            fired.add((p['id'], ru['rule_key'], None))
            title, detail = _alert_text(p, ru, val)
            exist = _first(conn.execute(
                "SELECT * FROM plm_alert WHERE project_id=? AND rule_key=? AND staff_id IS NULL "
                "AND status<>'已闭环' ORDER BY id DESC LIMIT 1", (p['id'], ru['rule_key'])))
            if exist:
                new_status = exist['status'] if exist['status'] == '处理中' else '待处理'
                conn.execute("""UPDATE plm_alert SET level=?,title=?,detail=?,metric_value=?,
                                threshold=?,status=?,last_scan_at=?,updated_at=? WHERE id=?""",
                             (ru['level'], title, detail, _r(val, 6), _f(ru['threshold']),
                              new_status, _now(), _now(), exist['id']))
                updated += 1
            else:
                conn.execute("""INSERT INTO plm_alert
                    (project_id,rule_key,dim,level,title,detail,metric_value,threshold,status,
                     last_scan_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (p['id'], ru['rule_key'], ru['dim'], ru['level'], title, detail,
                     _r(val, 6), _f(ru['threshold']), '待处理', _now(), _now()))
                created.append({'project_id': p['id'], 'rule_key': ru['rule_key'], 'title': title})
        # 人员过载：按项目 × 人员登记
        staff_rule = [r for r in rules if r['dim'] == 'staff']
        for s in staff_rows:
            if not any(a['project_id'] == p['id'] for a in s.get('assignments', [])):
                continue
            if s['load_state'] != '过载':
                continue
            for rule in staff_rule:
                fired.add((p['id'], rule['rule_key'], s['staff_id']))
                detail = '%s 已分配 %.1f 小时 / 月可用 %.1f 小时，负荷率 %s%%（并行 %d 个项目）' % (
                    s['name'], _f(s['planned_hours']), _f(s['available_hours']),
                    ('%.0f' % (_f(s['load_rate']) * 100)) if s['load_rate'] is not None else '-',
                    s['parallel_projects'])
                title = '人员工时过载：%s（%s）' % (s['name'], p['project_no'])
                exist = _first(conn.execute(
                    "SELECT * FROM plm_alert WHERE project_id=? AND rule_key=? AND staff_id=? "
                    "AND status<>'已闭环' ORDER BY id DESC LIMIT 1",
                    (p['id'], rule['rule_key'], s['staff_id'])))
                if exist:
                    conn.execute("UPDATE plm_alert SET title=?,detail=?,metric_value=?,"
                                 "level=?,last_scan_at=?,updated_at=? WHERE id=?",
                                 (title, detail, s['load_rate'], rule['level'], _now(),
                                  _now(), exist['id']))
                    updated += 1
                else:
                    conn.execute("""INSERT INTO plm_alert
                        (project_id,staff_id,rule_key,dim,level,title,detail,metric_value,
                         threshold,status,last_scan_at,updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (p['id'], s['staff_id'], rule['rule_key'], 'staff', rule['level'],
                         title, detail, s['load_rate'], _f(rule['threshold']), '待处理',
                         _now(), _now()))
                    created.append({'project_id': p['id'], 'rule_key': rule['rule_key'],
                                    'title': title})
    # 风险消除 → 自动闭环
    for a in _rows(conn.execute("SELECT * FROM plm_alert WHERE status<>'已闭环'")):
        if (a['project_id'], a['rule_key'], a.get('staff_id')) not in fired:
            conn.execute("""UPDATE plm_alert SET status='已闭环', handler='system',
                            handle_note='风险已消除（系统自动闭环）', handle_time=?, updated_at=?
                            WHERE id=?""", (_now(), _now(), a['id']))
            auto_closed.append(a['id'])
    conn.commit()
    conn.close()
    log_op('alert', '-', '预警扫描', '执行预警扫描',
           {'created': len(created), 'updated': updated, 'auto_closed': len(auto_closed)},
           operator=operator)
    return {'success': True, 'created': len(created), 'updated': updated,
            'auto_closed': len(auto_closed), 'created_items': created}


def list_alerts(project_id=None, dim=None, status=None, level=None, limit=1000):
    conn = get_db()
    sql = """SELECT a.*, p.project_no, p.project_name, p.status AS project_status,
                    s.name AS staff_name, r.rule_name
             FROM plm_alert a
             LEFT JOIN plm_project p ON p.id=a.project_id
             LEFT JOIN plm_staff s ON s.id=a.staff_id
             LEFT JOIN plm_alert_rule r ON r.rule_key=a.rule_key"""
    args, conds = [], []
    if project_id:
        conds.append("a.project_id=?"); args.append(project_id)
    if dim and dim != '全部':
        conds.append("a.dim=?"); args.append(dim)
    if level and level != '全部':
        conds.append("a.level=?"); args.append(level)
    if status and status != '全部':
        targets = {'待处理': ['待处理'], '处理中': ['处理中'], '已闭环': ['已闭环'],
                   '未闭环': ['待处理', '处理中']}.get(status, [status])
        conds.append("a.status IN (%s)" % ",".join("?" * len(targets)))
        args += targets
    if conds:
        sql += " WHERE " + " AND ".join(conds)
    sql += (" ORDER BY CASE a.level WHEN '严重' THEN 0 WHEN '警告' THEN 1 ELSE 2 END,"
            " a.id DESC LIMIT ?")
    args.append(int(limit))
    rows = _rows(conn.execute(sql, args))
    conn.close()
    return rows


def handle_alert(alert_id, payload, operator=''):
    """FR-9：预警闭环处置（待处理 → 处理中 → 已闭环）。"""
    status = payload.get('status')
    if status not in ALERT_STATUS:
        return {'success': False, 'error': 'status 需属于 %s' % list(ALERT_STATUS)}
    conn = get_db()
    a = _first(conn.execute("SELECT * FROM plm_alert WHERE id=?", (alert_id,)))
    if not a:
        conn.close()
        return {'success': False, 'error': '预警不存在'}
    conn.execute("""UPDATE plm_alert SET status=?,handler=?,handle_note=?,handle_time=?,
                    updated_at=? WHERE id=?""",
                 (status, operator or payload.get('handler', ''),
                  payload.get('note', a['handle_note']), _now(), _now(), alert_id))
    conn.commit()
    conn.close()
    log_op('alert', alert_id, a['title'], '预警处置',
           {'status': {'before': a['status'], 'after': status},
            'note': payload.get('note', '')}, operator=operator)
    return {'success': True}


# ===================== 驾驶舱总览 =====================

def overview():
    conn = get_db()
    n_proj = conn.execute("SELECT COUNT(*) n FROM plm_project").fetchone()['n']
    n_opp = conn.execute("SELECT COUNT(*) n FROM plm_opportunity").fetchone()['n']
    n_win = conn.execute("SELECT COUNT(*) n FROM plm_opportunity WHERE status='中标'").fetchone()['n']
    n_ct = conn.execute("SELECT COUNT(*) n FROM plm_contract").fetchone()['n']
    n_staff = conn.execute("SELECT COUNT(*) n FROM plm_staff WHERE status<>'离职'").fetchone()['n']
    sign_amount = _f(conn.execute("SELECT COALESCE(SUM(sign_amount),0) a FROM plm_contract")
                     .fetchone()['a'])
    income = _f(conn.execute("SELECT COALESCE(SUM(amount),0) a FROM plm_ledger WHERE kind='income'")
                .fetchone()['a'])
    actual = _f(conn.execute("SELECT COALESCE(SUM(amount),0) a FROM plm_ledger "
                             "WHERE kind='cost' AND plan_or_actual='实际'").fetchone()['a'])
    est_total = _f(conn.execute(
        "SELECT COALESCE(SUM(total_cost),0) a FROM plm_baseline WHERE scope_type='project' "
        "AND stage IN ('estimate_locked','estimate_bid') AND id IN (SELECT MAX(id) FROM plm_baseline "
        "WHERE scope_type='project' AND stage IN ('estimate_locked','estimate_bid') "
        "GROUP BY scope_id)").fetchone()['a'])
    bud_total = _f(conn.execute(
        "SELECT COALESCE(SUM(total_cost),0) a FROM plm_baseline WHERE scope_type='project' "
        "AND stage='budget' AND id IN (SELECT MAX(id) FROM plm_baseline WHERE scope_type='project' "
        "AND stage='budget' GROUP BY scope_id)").fetchone()['a'])
    open_alerts = _rows(conn.execute(
        "SELECT a.*, p.project_no, p.project_name, s.name staff_name FROM plm_alert a "
        "LEFT JOIN plm_project p ON p.id=a.project_id LEFT JOIN plm_staff s ON s.id=a.staff_id "
        "WHERE a.status<>'已闭环' ORDER BY CASE a.level WHEN '严重' THEN 0 WHEN '警告' THEN 1 "
        "ELSE 2 END, a.id DESC LIMIT 20"))
    dim_stat = {r['dim']: r['n'] for r in _rows(conn.execute(
        "SELECT dim, COUNT(*) n FROM plm_alert WHERE status<>'已闭环' GROUP BY dim"))}
    status_stat = {r['status']: r['n'] for r in _rows(
        conn.execute("SELECT status, COUNT(*) n FROM plm_project GROUP BY status"))}
    opp_stat = {r['status']: r['n'] for r in _rows(
        conn.execute("SELECT status, COUNT(*) n FROM plm_opportunity GROUP BY status"))}
    conn.close()
    return {
        'kpi': {
            'projects': n_proj, 'opportunities': n_opp, 'won_opportunities': n_win,
            'contracts': n_ct, 'staff': n_staff,
            'sign_amount': _r(sign_amount), 'income': _r(income),
            'estimate_total': _r(est_total), 'budget_total': _r(bud_total),
            'actual_cost': _r(actual),
            'actual_gross': _r(income - actual),
            'actual_gross_rate': _rate(income - actual, income),
            'budget_usage_rate': _rate(actual, bud_total),
            'open_alerts': len(open_alerts),
        },
        'alert_by_dim': dim_stat,
        'project_by_status': status_stat,
        'opportunity_by_status': opp_stat,
        'projects': list_projects(limit=200),
        'alerts': open_alerts,
        'over_budget_projects': [
            x for x in list_projects(limit=200)
            if x.get('budget_cost') and x.get('actual_cost')
            and _f(x['actual_cost']) > _f(x['budget_cost'])],
    }


# ===================== 模块八：报表导出 =====================

REPORTS = ('panorama', 'project_compare', 'schedule', 'labor', 'cost')


def _xlsx(sheets):
    """sheets: [(sheet_name, [[cell,...], ...]), ...] → xlsx bytes"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        for row in rows:
            ws.append(['' if v is None else v for v in row])
        for col in range(1, min(ws.max_column or 1, 30) + 1):
            width = 10
            for row in range(1, min(ws.max_row or 1, 60) + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    width = max(width, min(38, len(str(v)) + 2))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _pct(v):
    return None if v is None else '%.1f%%' % (float(v) * 100)


def export_report(report, project_id=None):
    """FR-10：返回 (filename, bytes)。无数据时返回表头结构。"""
    if report not in REPORTS:
        raise ValueError('未知报表：%s' % report)
    if report == 'panorama':
        if not project_id:
            raise ValueError('单项目全景报表需提供 project_id')
        pan = project_panorama(project_id)
        if not pan:
            raise ValueError('项目不存在')
        b, ba, pm, hr, fa, aa = (pan['base_info'], pan['baseline_area'], pan['pmo_area'],
                                 pan['hr_area'], pan['finance_area'], pan['alert_area'])
        base_rows = [['字段', '值']] + [[k, v] for k, v in b.items()]
        bl_rows = [['基线', '收入(元)', '成本(元)', '毛利(元)', '毛利率', '状态', '备注'],
                   ['概算', (ba['estimate'] or {}).get('total_income'),
                    (ba['estimate'] or {}).get('total_cost'),
                    (ba['estimate'] or {}).get('gross'), _pct((ba['estimate'] or {}).get('gross_rate')),
                    (ba['estimate'] or {}).get('status'), '顶层管控基线'],
                   ['预算', (ba['budget'] or {}).get('total_income'),
                    (ba['budget'] or {}).get('total_cost'),
                    (ba['budget'] or {}).get('gross'), _pct((ba['budget'] or {}).get('gross_rate')),
                    (ba['budget'] or {}).get('status'), ba['budget_usage_note']],
                   ['核算', None, None, None, None, '预留', RESERVED_NOTE],
                   ['决算', None, None, None, None, '预留', RESERVED_NOTE]]
        ms_rows = [['类型', '名称', '层级', '计划开始', '计划结束', '实际结束',
                    '完成%', '状态', '关键', '负责人', '超期天数']]
        for m in pm['milestones']:
            ms_rows.append(['里程碑', m['name'], m['level'], m['plan_start'], m['plan_end'],
                            m['actual_end'], m['progress'], m['status'],
                            '是' if m['is_key'] else '否', m['owner'],
                            next((n['overdue_days'] for n in pm['schedule']['overdue_nodes']
                                  if n['type'] == 'milestone' and n['id'] == m['id']), '')])
        for t in list_tasks(project_id=project_id):
            ms_rows.append(['任务', t['name'], '', t['plan_start'], t['plan_end'],
                            t['actual_end'], t['progress'], t['status'], '', t['owner'],
                            t['overdue_days'] if t['is_overdue'] else ''])
        hr_rows = [['姓名', '岗位', '本项目工时', '总分配工时', '实际工时', '可用工时',
                    '负荷率', '负荷状态', '并行项目数']]
        for x in hr['participants']:
            hr_rows.append([x['name'], x['role'], x['project_hours'], x['planned_hours'],
                            x['actual_hours'], x['available_hours'], _pct(x['load_rate']),
                            x['load_state'], x['parallel_projects']])
        fin_rows = [['指标', '数值'],
                    ['签单收入', fa['income']['signed']], ['变更收入', fa['income']['change']],
                    ['收入合计', fa['income']['total']],
                    ['概算成本', fa['baseline']['estimate_cost']],
                    ['预算总额', fa['baseline']['budget_total']],
                    ['预估成本(台账)', fa['cost']['estimate']],
                    ['累计实际成本', fa['cost']['actual_cum']],
                    ['其中工时归集人力成本', fa['cost']['labor_auto']],
                    ['累计工时', fa['cost']['hours_total']],
                    ['签单毛利', fa['gross']['signed']],
                    ['签单毛利率', _pct(fa['gross']['signed_rate'])],
                    ['预估毛利', fa['gross']['estimate']],
                    ['实际毛利', fa['gross']['actual']],
                    ['实际毛利率', _pct(fa['gross']['actual_rate'])],
                    ['概算-预算差异', fa['variance']['estimate_vs_budget']],
                    ['预算-实际差异', fa['variance']['budget_vs_actual']],
                    ['预算消耗占比', _pct(fa['variance']['budget_usage_rate'])],
                    ['成本方向', fa['variance']['direction']]]
        al_rows = [['风险维度', '等级', '标题', '当前值', '阈值', '状态', '处置人', '处置说明']]
        for a in aa['items']:
            al_rows.append([a['dim'], a['level'], a['title'], a['metric_value'],
                            a['threshold'], a['status'], a['handler'], a['handle_note']])
        fn = '项目全景_%s.xlsx' % b['project_no']
        return fn, _xlsx([('基础信息', base_rows), ('四算基线', bl_rows), ('进度与任务', ms_rows),
                          ('人力工时', hr_rows), ('成本毛利', fin_rows), ('风险预警', al_rows)])
    if report == 'project_compare':
        rows = [['项目编号', '项目名称', '客户', '项目经理', '状态',
                 '概算成本', '预算总额', '概算-预算差异', '累计实际成本',
                 '预算消耗占比', '收入合计', '实际毛利', '实际毛利率', '未闭环预警']]
        for x in list_projects(limit=1000):
            rows.append([x['project_no'], x['project_name'], x['customer'], x['manager'],
                         x['status'], x['estimate_cost'], x['budget_cost'],
                         _r((x['budget_cost'] or 0) - (x['estimate_cost'] or 0))
                         if (x['budget_cost'] is not None and x['estimate_cost'] is not None)
                         else None,
                         x['actual_cost'], _pct(x['actual_cost'] and x['budget_cost'] and
                                                (x['actual_cost'] / x['budget_cost'])),
                         x['income'], x['actual_gross'], _pct(x['actual_gross_rate']),
                         x['open_alerts']])
        return '多项目概算预算对比.xlsx', _xlsx([('概算预算对比', rows)])
    if report == 'schedule':
        rows = [['类型', '项目编号', '项目名称', '节点', '负责人', '计划开始', '计划结束',
                 '实际结束', '完成%', '状态', '超期天数', '关键节点']]
        for x in list_projects(limit=500):
            pr = project_progress(x['id'])
            rows.append(['项目汇总', x['project_no'], x['project_name'], '整体', x['manager'],
                         x['start_date'], x['end_date'], '',
                         _pct(pr['schedule']['progress_rate']), x['status'],
                         pr['schedule']['max_overdue_days'], ''] )
            for n in pr['schedule']['overdue_nodes']:
                rows.append(['延期' + ('里程碑' if n['type'] == 'milestone' else '任务'),
                             x['project_no'], x['project_name'], n['name'], n['owner'],
                             '', n['plan_end'], '', '', '延期', n['overdue_days'],
                             '是' if n.get('is_key') else ''])
        return 'PMO进度报表.xlsx', _xlsx([('PMO进度', rows)])
    if report == 'labor':
        rows = [['姓名', '岗位', '部门', '人天单价', '可用工时', '已分配工时', '实际工时',
                 '负荷率', '负荷状态', '并行项目数', '并行项目']]
        for x in staff_load():
            rows.append([x['name'], x['role'], x['dept'], x['cost_rate'], x['available_hours'],
                         x['planned_hours'], x['actual_hours'], _pct(x['load_rate']),
                         x['load_state'], x['parallel_projects'],
                         '、'.join(p['project_no'] for p in x['projects'])])
        ts_rows = [['日期', '姓名', '项目编号', '项目名称', '任务', '工时', '备注']]
        for t in list_timesheets(limit=3000):
            ts_rows.append([t['work_date'], t['staff_name'], t['project_no'],
                            t['project_name'], t['task_name'] or '', t['hours'], t['remark']])
        return '人力工时与负荷.xlsx', _xlsx([('人员负荷', rows), ('工时明细', ts_rows)])
    # cost
    rows = [['项目编号', '项目名称', '签单收入', '收入合计', '概算成本', '预算总额',
             '累计实际成本', '预算消耗占比', '实际毛利', '实际毛利率', '成本方向']]
    for x in list_projects(limit=500):
        fa = project_finance(x['id'])
        rows.append([x['project_no'], x['project_name'], fa['income']['signed'],
                     fa['income']['total'], fa['baseline']['estimate_cost'],
                     fa['baseline']['budget_total'], fa['cost']['actual_cum'],
                     _pct(fa['variance']['budget_usage_rate']), fa['gross']['actual'],
                     _pct(fa['gross']['actual_rate']), fa['variance']['direction']])
    return '项目成本毛利报表.xlsx', _xlsx([('成本毛利', rows)])

