"""ETL 任务定义与注册（服务层，跨域共享）。"""
from models import get_db, init_db
import json
import openpyxl
from common.datasource_meta import _ds_latest_path

ETL_JOB_DEFS = [
    {
        'job_key': 'gross-margin',
        'job_name': '签单毛利指标计算',
        'description': '按年份/区域聚合签单毛利、签单毛利率',
        'schedule': '0 2 * * *',
        'calculation_logic': '数据源：总合同表（数据源管理最新版）。字段映射：「统计日期」→年份、「区域」分组、「合同总金额」+「签单毛利」求和。计算：签单毛利率 = 签单毛利 ÷ 合同总金额。产出：indicator_metrics 宽表（dim_type=year 按年份、dim_type=region 按区域×年份）。',
    },
    {
        'job_key': 'payment-cycle',
        'job_name': '回款周期指标计算',
        'description': '按合同聚合回款周期指标',
        'schedule': '0 3 * * *',
        'calculation_logic': '数据源：总合同表 + 项目里程碑表。口径：按合同编号关联回款记录，取最后一笔回款日期；回款周期 = 最后一笔回款日 − 合同签订日（统计日期）；按合同聚合。产出：payment_cycle_metrics 宽表（每合同一行）。',
    },
    {
        'job_key': 'sign-summary',
        'job_name': '签约汇总指标计算',
        'description': '按业务线/区域聚合签约合同额',
        'schedule': '0 4 * * *',
        'calculation_logic': '数据源：总合同表。口径：按「业务线」「区域」分组，聚合当年生效合同额与签约合同数。（骨架阶段，计算逻辑待实现）',
    },
    {
        'job_key': 'fund-occupancy',
        'job_name': '资金占用指标计算',
        'description': 'FIFO 垫资冲抵，计算每个合同的资金占用、加权资金占用、资金成本',
        'schedule': '0 5 * * *',
        'calculation_logic': '数据源：付款明细表 + 收款明细表。口径：按合同编号透视（付款/收款按日期聚合）→ FIFO 先进先出冲抵 → 预收款冲抵后续付款 → 生成垫资片段（SETTLED/OCCUPYING）。计算：当前资金占用=占用中片段金额和；元天合计=片段金额×占用天数；预估资金成本=元天×日利率（年化3%）。产出：fund_metrics 宽表（每合同一行）。',
    },
    {
        'job_key': 'fund-multidim',
        'job_name': '资金占用多维度聚合',
        'description': '按区域/部门/业务线/客户集合/月份聚合资金占用与风险分布',
        'schedule': '30 5 * * *',
        'calculation_logic': '数据源：fund_metrics 宽表（资金占用指标计算产物）。口径：按维度列（region/dept/biz_line/industry/customer_key/project_status/contract_status/sign_year/month）分组聚合合同数、累计付款/收款、当前资金占用、回款率、占用强度、风险等级分布。产出：indicator_metrics 宽表（dim_type=fund_dim）。',
    },
]

def _register_etl_jobs():
    """注册 ETL 任务定义（幂等，UPSERT 保证计算逻辑同步）"""
    from models import init_db
    init_db()
    conn = get_db()
    c = conn.cursor()
    for job in ETL_JOB_DEFS:
        c.execute("""
            INSERT INTO etl_jobs (job_key, job_name, description, calculation_logic, schedule)
            VALUES (?,?,?,?,?)
            ON CONFLICT(job_key) DO UPDATE SET
                job_name=excluded.job_name,
                description=excluded.description,
                calculation_logic=excluded.calculation_logic,
                schedule=excluded.schedule
        """, (job['job_key'], job['job_name'], job['description'], job['calculation_logic'], job['schedule']))
    conn.commit()
    conn.close()


def run_etl_gross_margin():
    """签单毛利 ETL：读总合同表 → 按年份/区域聚合 → 写指标宽表"""
    import openpyxl
    from datetime import datetime
    from collections import defaultdict

    h_fpath = _ds_latest_path('总合同表')
    if not h_fpath:
        return {'success': False, 'error': '总合同表未上传'}

    def safe_float(v):
        if v is None or v == '' or v == '-': return 0.0
        try: return float(v)
        except: return 0.0

    def find_col(headers, keywords):
        for h in headers:
            hl = str(h).lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    wb = openpyxl.load_workbook(h_fpath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) if h is not None else '' for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    col_amt = find_col(headers, ['合同总金额'])
    col_gross = find_col(headers, ['签单毛利'])
    col_region = find_col(headers, ['区域'])
    col_date = find_col(headers, ['统计日期'])
    col_dept = find_col(headers, ['业务类型'])

    def dept_cn(biz):
        """从业务类型映射部门（智能计算与集成事业部 ICID 下属）：
        系统集成业务→系统集成部、运维服务业务→运维服务部、资产运营业务→资产运营部、一体化运维→运维平台；
        其余业务类型不纳入部门维度"""
        biz = str(biz).strip()
        if not biz:
            return ''
        if '一体化运维' in biz:
            return '运维平台'
        if '运维服务' in biz:
            return '运维服务部'
        if '系统集成' in biz:
            return '系统集成部'
        if '资产运营' in biz:
            return '资产运营部'
        return ''

    def parse_year(v):
        if isinstance(v, datetime): return v.year
        if hasattr(v, 'year'): return v.year
        try: return int(str(v)[:4])
        except: return None

    year_agg = defaultdict(lambda: {'amt': 0.0, 'gross': 0.0})
    region_year_agg = defaultdict(lambda: defaultdict(lambda: {'amt': 0.0, 'gross': 0.0}))
    dept_year_agg = defaultdict(lambda: defaultdict(lambda: {'amt': 0.0, 'gross': 0.0}))
    dept_region_year_agg = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'amt': 0.0, 'gross': 0.0})))
    for r in rows[1:]:
        if r is None: continue
        amt = safe_float(r[col_idx[col_amt]]) if col_amt else 0.0
        gross = safe_float(r[col_idx[col_gross]]) if col_gross else 0.0
        region = str(r[col_idx[col_region]]).strip() if col_region and r[col_idx[col_region]] else ''
        dept = str(r[col_idx[col_dept]]).strip() if col_dept and r[col_idx[col_dept]] else ''
        dept = dept_cn(dept) if dept else ''
        y = parse_year(r[col_idx[col_date]]) if col_date else None
        if y is not None:
            year_agg[y]['amt'] += amt
            year_agg[y]['gross'] += gross
            if region:
                region_year_agg[region][y]['amt'] += amt
                region_year_agg[region][y]['gross'] += gross
            if dept:
                dept_year_agg[dept][y]['amt'] += amt
                dept_year_agg[dept][y]['gross'] += gross
            if region and dept:
                dept_region_year_agg[dept][region][y]['amt'] += amt
                dept_region_year_agg[dept][region][y]['gross'] += gross

    def rate(g, a): return round(g / a, 6) if a else 0.0

    conn = get_db()
    c = conn.cursor()
    # 兼容旧数据库：若 indicator_metrics 无 extra_json 列则添加
    cols = [row[1] for row in c.execute("PRAGMA table_info(indicator_metrics)").fetchall()]
    if 'extra_json' not in cols:
        c.execute("ALTER TABLE indicator_metrics ADD COLUMN extra_json TEXT DEFAULT '{}' ")
    c.execute("DELETE FROM indicator_metrics WHERE job_key='gross-margin'")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0
    for y, v in year_agg.items():
        c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, calc_time) VALUES (?,?,?,?,?,?,?,?,?)",
                  ('gross-margin', '签单毛利率', 'year', str(y), str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), now))
        n += 1
    for region, yd in region_year_agg.items():
        for y, v in yd.items():
            c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, calc_time) VALUES (?,?,?,?,?,?,?,?,?)",
                      ('gross-margin', '签单毛利率', 'region', region, str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), now))
            n += 1
    for dept, yd in dept_year_agg.items():
        for y, v in yd.items():
            c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, calc_time) VALUES (?,?,?,?,?,?,?,?,?)",
                      ('gross-margin', '签单毛利率', 'dept', dept, str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), now))
            n += 1
    for dept, rd in dept_region_year_agg.items():
        for region, yd in rd.items():
            for y, v in yd.items():
                extra = json.dumps({'dept': dept, 'region': region})
                c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, extra_json, calc_time) VALUES (?,?,?,?,?,?,?,?,?,?)",
                          ('gross-margin', '签单毛利率', 'dept_region', f"{dept}|{region}", str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), extra, now))
                n += 1
    conn.commit()
    conn.close()
    return {'success': True, 'rows': n}

def run_etl_fund_multidim():
    """资金占用多维度 ETL：读 fund_metrics 宽表 → 按维度×月份聚合 → 写 indicator_metrics

    产出 dim_type='fund_dim'，metric_name 形如 'fund_dim:region'。
    """
    from datetime import datetime
    from collections import defaultdict
    from main import DIM_COLUMN_MAP, DIM_NAME_MAP, _fund_rows_with_dims

    rows = _fund_rows_with_dims()
    if not rows:
        return {'success': False, 'error': 'fund_metrics 为空，请先执行「资金占用指标计算」'}

    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM indicator_metrics WHERE job_key='fund-multidim'")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0

    for dim in DIM_COLUMN_MAP:
        col = DIM_COLUMN_MAP[dim]
        agg = defaultdict(lambda: {
            'count': 0, 'pay': 0.0, 'recv': 0.0, 'occupy': 0.0, 'amt_day': 0.0,
            'cost': 0.0, 'ca': 0.0, 'risk': {'healthy': 0, 'yellow': 0, 'orange': 0, 'red': 0},
        })
        for r in rows:
            name = (r.get(col) or '').strip() or '未知'
            month = (r.get('cycle_start') or '')[:7] or '未知'
            b = agg[(name, month)]
            b['count'] += 1
            b['pay'] += r.get('total_pay') or 0
            b['recv'] += r.get('total_recv') or 0
            b['occupy'] += r.get('current_occupy') or 0
            b['amt_day'] += r.get('amount_day') or 0
            b['cost'] += r.get('est_cost') or 0
            b['ca'] += r.get('contract_amount') or 0
            rl = r.get('risk_level') or 'healthy'
            b['risk'][rl if rl in b['risk'] else 'healthy'] += 1

        for (name, month), b in agg.items():
            ca = b['ca'] or b['pay'] or 0
            extra = {
                'dim': dim, 'dim_name': DIM_NAME_MAP.get(dim, dim), 'month': month,
                'contract_count': b['count'],
                'total_pay': round(b['pay'], 2), 'total_recv': round(b['recv'], 2),
                'current_occupy': round(b['occupy'], 2), 'amount_day': round(b['amt_day'], 2),
                'est_cost': round(b['cost'], 2),
                'recv_rate': round(b['recv'] / ca, 4) if ca > 0 else 0,
                'occupy_intensity': round(b['occupy'] / ca, 4) if ca > 0 else 0,
                'risk_count': b['risk'],
                'risk_level': next((lv for lv in ['red', 'orange', 'yellow', 'healthy']
                                    if b['risk'].get(lv, 0) > 0), 'healthy'),
            }
            c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, extra_json, calc_time) VALUES (?,?,?,?,?,?,?,?,?,?)",
                      ('fund-multidim', f'fund_dim:{dim}', 'fund_dim', name, month,
                       b['count'], b['occupy'], extra['occupy_intensity'],
                       json.dumps(extra, ensure_ascii=False), now))
            n += 1

    conn.commit()
    conn.close()
    return {'success': True, 'rows': n}

def run_etl_payment_cycle():
    """回款周期 ETL：读总合同表 + 项目里程碑表 → 算每合同回款周期 → 写宽表"""
    import openpyxl
    from datetime import datetime, date

    h_fpath = _ds_latest_path('总合同表')
    r_fpath = _ds_latest_path('项目里程碑表')
    if not h_fpath:
        return {'success': False, 'error': '总合同表未上传'}

    def parse_date(v):
        if v is None or v == '' or v == '-': return None
        if isinstance(v, (datetime, date)): return v
        s = str(v).strip()
        if s.startswith('='): return None
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try: return datetime.strptime(s, fmt)
            except: pass
        return None

    def safe_float(v):
        if v is None or v == '' or v == '-': return 0.0
        try: return float(v)
        except: return 0.0

    def find_col(headers, keywords):
        for h in headers:
            hl = str(h).lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    # 读总合同表
    wb = openpyxl.load_workbook(h_fpath, data_only=True)
    ws = wb.active
    hrows = list(ws.iter_rows(values_only=True))
    wb.close()
    h_headers = [str(h) if h else '' for h in hrows[0]]
    h_col_no = find_col(h_headers, ['合同编号'])
    h_col_date = find_col(h_headers, ['统计日期', '签约日期'])
    h_col_amt = find_col(h_headers, ['合同总金额', '合同金额', '合同额'])
    h_idx = {h: i for i, h in enumerate(h_headers)}
    contracts = {}
    for r in hrows[1:]:
        if r is None: continue
        cno = str(r[h_idx[h_col_no]]).strip() if h_col_no and h_idx.get(h_col_no) is not None else ''
        if not cno: continue
        cdate = parse_date(r[h_idx[h_col_date]]) if h_col_date and h_idx.get(h_col_date) is not None else None
        amt = safe_float(r[h_idx[h_col_amt]]) if h_col_amt and h_idx.get(h_col_amt) is not None else 0.0
        contracts[cno] = {'date': cdate, 'amount': amt}

    # 读项目里程碑表，找每合同的最后一笔回款日期
    last_pay = {}
    if r_fpath:
        wb = openpyxl.load_workbook(r_fpath, data_only=True)
        ws = wb.active
        rrows = list(ws.iter_rows(values_only=True))
        wb.close()
        r_headers = [str(h) if h else '' for h in rrows[0]]
        r_col_no = find_col(r_headers, ['合同编号'])
        r_col_pay_date = find_col(r_headers, ['回款时间', '回款日期', '到款时间'])
        r_idx = {h: i for i, h in enumerate(r_headers)}
        for r in rrows[1:]:
            if r is None: continue
            cno = str(r[r_idx[r_col_no]]).strip() if r_col_no and r_idx.get(r_col_no) is not None else ''
            if not cno: continue
            pdate = parse_date(r[r_idx[r_col_pay_date]]) if r_col_pay_date and r_idx.get(r_col_pay_date) is not None else None
            if pdate and (cno not in last_pay or pdate > last_pay[cno]):
                last_pay[cno] = pdate

    # 算回款周期，写宽表
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM payment_cycle_metrics")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0
    for cno, info in contracts.items():
        cdate = info['date']
        pdate = last_pay.get(cno)
        cycle_days = 0
        if cdate and pdate:
            cycle_days = (pdate - cdate).days
            if cycle_days < 0: cycle_days = 0
        c.execute("INSERT INTO payment_cycle_metrics (contract_no, contract_date, last_payment_date, cycle_days, amount, calc_time) VALUES (?,?,?,?,?,?)",
                  (cno, cdate.strftime('%Y-%m-%d') if cdate else '', pdate.strftime('%Y-%m-%d') if pdate else '', cycle_days, info['amount'], now))
        n += 1
    conn.commit()
    conn.close()
    return {'success': True, 'rows': n}
