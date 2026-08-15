"""
合同比对系统 — Excel 导入导出处理
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
import re
from models import get_db, clear_contract
from compare_engine import run_comparison

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')

# 列名映射（支持模糊匹配，持续扩充中）
COLUMN_ALIASES = {
    'device_name': [
        '项目', '项目名称', '产品名称', '设备名称', '名称', '设备名', '品名',
        '货物名称', '物料名称', '标的', '采购内容', '商品名称', '产品', '设备',
        'name', 'product', 'item',
    ],
    'device_model': [
        '型号规格', '型号', '设备型号', '规格型号', '产品型号', '物料型号',
        '规格', 'model', '型号/规格', '规格/型号',
    ],
    'specs_full': [
        '详细参数', '规格参数', '参数', '详细规格', '功能要求', '配置描述',
        '技术参数', '技术规格', '主要参数', '规格描述', '参数规格', '技术指标',
        '性能参数', '功能要求/配置描述', '配置要求', '技术要求',
        'specs', 'spec', 'description', '参数及要求',
    ],
    'qty': [
        '数量', '合同数量', '报价数量', '采购数量', '采购量', '需求数量',
        '供货数量', '计划数量', 'qty', 'quantity',
    ],
    'unit': [
        '单位', '计量单位', 'unit',
    ],
    'unit_price': [
        '单价', '报价单价', '合同单价', '综合单价', '不含税单价',
        '单价（元）', '单价(元)', '报价(元)',
        'price', 'unit_price', '含税单价',
    ],
    'amount': [
        '金额', '报价金额', '合同金额', '总价', '合价', '小计', '合计金额',
        '总金额', '报价金额(元)', '总价（元）', '总价(元)',
        'amount', 'total', '含税金额',
    ],
    'remark': [
        '备注', '说明', 'remark', 'note',
    ],
}


def _clean_header(h: str) -> str:
    """清洗表头：去换行、去长括号注释（保留短的单位括号）"""
    h = str(h or '').strip().replace('\n', '').replace('\\n', '')
    # 去掉4字以上的括号注释，如 "品牌（如指定请填写）" → "品牌"
    # 保留短的如 "单价（元）" "总价（元）"
    h = re.sub(r'（[^）]{4,}）', '', h)
    h = re.sub(r'\([^)]{4,}\)', '', h)
    return h.strip().lower()


def find_column(headers: list, field: str) -> int:
    """多策略列匹配：先匹配完整表头，再匹配清洗后的表头"""
    aliases = COLUMN_ALIASES.get(field, [])
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_clean = _clean_header(h)
        if not h_clean:
            continue
        for alias in aliases:
            a_lower = alias.lower()
            if a_lower in h_clean or h_clean in a_lower:
                return i
    return -1


def _smart_detect_columns(raw_headers: list, clean_headers: list) -> dict:
    """
    智能列匹配 — 四层兜底：
    1. find_column 标准匹配
    2. 兜底：device_name → 第一个非数字/非金额列
    3. 兜底：qty → 第一个全数字列
    4. 兜底：specs_full → 最长文本列
    """
    idx = {}
    for field in ['device_name', 'device_model', 'specs_full', 'qty', 'unit',
                   'unit_price', 'amount', 'remark']:
        idx[field] = find_column(raw_headers, field)

    # 兜底1: device_name → 第一个看起来像名称的列（不是数量/单价/金额/序号）
    if idx['device_name'] < 0:
        skip_kw = ['数量', '单价', '金额', '单位', '序号', '编号', '备注', '合计',
                   'qty', 'price', 'amount', 'no', 'id', '序号']
        for i, h in enumerate(clean_headers):
            if h and not any(kw in h for kw in skip_kw):
                idx['device_name'] = i
                break

    # 兜底2: specs_full → 包含 参数/配置/描述/要求 的列
    # 注意：不含「规格」，否则会误匹配「规格型号」列（型号被当成参数）
    if idx['specs_full'] < 0:
        specs_kw = ['参数', '配置', '描述', '要求', '技术', '功能']
        for i, h in enumerate(clean_headers):
            if h and any(kw in h for kw in specs_kw):
                idx['specs_full'] = i
                break

    # 兜底3: model → 包含 型号/规格/model 的列
    if idx['device_model'] < 0:
        for i, h in enumerate(clean_headers):
            if h and any(kw in h for kw in ['型号', '规格', 'model']):
                idx['device_model'] = i
                break

    return idx


def safe_str(val) -> str:
    """安全转字符串"""
    if val is None:
        return ''
    return str(val).strip()


def safe_float(val):
    """安全转浮点数"""
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def import_contract_excel(contract_id: int, filepath: str) -> dict:
    """导入合同基准Excel — 不限制列名，传什么存什么"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 读取表头（清理换行符）
    raw_headers = []
    for cell in ws[1]:
        val = str(cell.value or '').strip().replace('\n', '').replace('\\n', '')
        raw_headers.append(val)

    clean_headers = [_clean_header(h) for h in raw_headers]

    # 智能匹配列
    idx = _smart_detect_columns(raw_headers, clean_headers)

    # 如果还没找到 device_name，用第一个非空文本列
    if idx['device_name'] < 0:
        for i, h in enumerate(clean_headers):
            if h and not any(kw in h for kw in ['数量', '单价', '金额', '单位']):
                idx['device_name'] = i
                break

    # 清空旧数据并导入
    clear_contract(contract_id)
    conn = get_db()
    c = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        # 构建原始列数据(JSON)
        raw_data = {}
        for i, h in enumerate(raw_headers):
            if h and i < len(row):
                raw_data[h] = safe_str(row[i])

        device_name = safe_str(row[idx['device_name']]) if idx['device_name'] >= 0 and idx['device_name'] < len(row) else ''
        if not device_name:
            continue

        device_model = safe_str(row[idx['device_model']]) if idx['device_model'] >= 0 and idx['device_model'] < len(row) else ''
        specs_full = safe_str(row[idx['specs_full']]) if idx['specs_full'] >= 0 and idx['specs_full'] < len(row) else ''

        if not device_model and device_name:
            parts = device_name.split()
            if len(parts) >= 2:
                device_model = parts[-1]

        from compare_engine import extract_specs
        parsed = extract_specs(specs_full)

        qty = safe_float(row[idx['qty']]) if idx['qty'] >= 0 and idx['qty'] < len(row) else 0
        unit = safe_str(row[idx['unit']]) if idx['unit'] >= 0 and idx['unit'] < len(row) else ''
        unit_price = safe_float(row[idx['unit_price']]) if idx['unit_price'] >= 0 and idx['unit_price'] < len(row) else 0
        amount = safe_float(row[idx['amount']]) if idx['amount'] >= 0 and idx['amount'] < len(row) else 0
        # 金额兜底：公式单元格读不到缓存值时，用 数量×单价 计算
        if amount <= 0 and qty > 0 and unit_price > 0:
            amount = round(qty * unit_price, 2)
        remark = safe_str(row[idx['remark']]) if idx['remark'] >= 0 and idx['remark'] < len(row) else ''

        c.execute("""
            INSERT INTO contract_items
            (contract_id, device_name, device_model, specs_full, specs_cpu, specs_memory, specs_disk,
             specs_other, contract_qty, contract_unit, contract_unit_price, contract_amount, remark, raw_columns)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            contract_id, device_name, device_model, specs_full,
            parsed.get('cpu', ''), parsed.get('memory', ''), parsed.get('disk', ''),
            parsed.get('other', ''),
            qty, unit, unit_price, amount, remark,
            json.dumps(raw_data, ensure_ascii=False)
        ))
        count += 1

    conn.commit()
    conn.close()
    return {'success': True, 'count': count, 'headers': raw_headers,
            'matched_columns': {k: (raw_headers[v] if v >= 0 else None) for k, v in idx.items()}}


def _extract_contract_headers(conn, contract_id: int) -> list:
    """从 contract_items.raw_columns 提取主合同列名（保持顺序去重）"""
    headers = []
    rows = conn.execute(
        "SELECT raw_columns FROM contract_items WHERE contract_id=? ORDER BY id",
        (contract_id,)
    ).fetchall()
    for r in rows:
        try:
            raw = json.loads(r['raw_columns'] or '{}')
            for k in raw:
                if k and k not in headers:
                    headers.append(k)
        except Exception:
            pass
    return headers


def build_column_mapping(contract_headers: list, supplier_headers: list) -> dict:
    """自动对齐：主合同列 ↔ 供应商列。返回 column_mapping dict。"""
    ct_clean = [_clean_header(h) for h in contract_headers]
    sp_clean = [_clean_header(h) for h in supplier_headers]
    ct_idx = _smart_detect_columns(contract_headers, ct_clean)
    sp_idx = _smart_detect_columns(supplier_headers, sp_clean)

    # 语义字段 → 列名（未识别到的字段不记录）
    contract_semantics = {}
    for field, colidx in ct_idx.items():
        if colidx >= 0 and colidx < len(contract_headers):
            contract_semantics[field] = contract_headers[colidx]
    supplier_semantics = {}
    for field, colidx in sp_idx.items():
        if colidx >= 0 and colidx < len(supplier_headers):
            supplier_semantics[field] = supplier_headers[colidx]

    # 主合同每列 → 供应商列
    mapping = {}
    # 非业务列：序号/编号类，仅展示、不参与逐字比对
    _skip_cols = {'序号', '编号', '项号', '序号.', 'No.', 'no.', 'NO', 'NO.', '序号（如有）'}
    for i, ch in enumerate(contract_headers):
        field = None
        for f, colidx in ct_idx.items():
            if colidx == i:
                field = f
                break
        target = ''
        if field is not None:
            si = sp_idx.get(field, -1)
            if si >= 0 and si < len(supplier_headers):
                target = supplier_headers[si]
        # 纯序号列不参与比对（主合同章节号 vs 供应商行号，无业务可比性）
        if _clean_header(ch).strip() in _skip_cols:
            target = ''
        elif not target and ch in supplier_headers:
            target = ch
        mapping[ch] = target

    # 供应商多出来的列（未被任何主合同列映射到的）→ 仅参考、不参与比对
    mapped = {v for v in mapping.values() if v}
    reference_columns = [h for h in supplier_headers if h not in mapped]

    return {
        'contract_headers': contract_headers,
        'supplier_headers': supplier_headers,
        'mapping': mapping,
        'contract_semantics': contract_semantics,
        'supplier_semantics': supplier_semantics,
        'reference_columns': reference_columns,
    }


def reapply_column_mapping(contract_id: int, version_id: int, mapping: dict, contract_semantics: dict) -> dict:
    """按新的列对齐 mapping，从 supplier_items.raw_columns 重提供应商结构化字段，再重新比对。"""
    conn = get_db()
    # 语义字段 → 供应商列名
    semantics_to_sp = {}
    for field, ct_col in (contract_semantics or {}).items():
        sp_col = (mapping or {}).get(ct_col, '')
        if sp_col:
            semantics_to_sp[field] = sp_col

    items = conn.execute(
        "SELECT id, raw_columns FROM supplier_items WHERE version_id=?", (version_id,)
    ).fetchall()
    for it in items:
        try:
            raw = json.loads(it['raw_columns'] or '{}')
        except Exception:
            raw = {}

        def gv(field, default=''):
            col = semantics_to_sp.get(field, '')
            return raw.get(col, '') if col else default

        qty = safe_float(gv('qty', '0'))
        unit_price = safe_float(gv('unit_price', '0'))
        amount = safe_float(gv('amount', '0'))
        if amount <= 0 and qty > 0 and unit_price > 0:
            amount = round(qty * unit_price, 2)

        conn.execute("""UPDATE supplier_items SET device_name=?, device_model=?, specs_full=?,
            quote_qty=?, quote_unit=?, quote_unit_price=?, quote_amount=?, remark=? WHERE id=?""",
            (gv('device_name', ''), gv('device_model', ''), gv('specs_full', ''),
             qty, gv('unit', ''), unit_price, amount, gv('remark', ''), it['id']))

    conn.commit()
    conn.close()
    return run_comparison(contract_id, version_id)


def import_supplier_excel(contract_id: int, filepath: str, supplier_name: str = '') -> dict:
    """导入供应商报价Excel，生成新版本 — 智能列匹配，不预设格式"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 1. 读取并清洗表头
    raw_headers = []
    for cell in ws[1]:
        val = str(cell.value or '').strip().replace('\n', '').replace('\\n', '')
        raw_headers.append(val)
    clean_headers = [_clean_header(h) for h in raw_headers]

    # 2. 智能匹配列
    idx = _smart_detect_columns(raw_headers, clean_headers)

    # 3. 额外兜底：如果第0列是"序号"且 device_name 没找到，尝试用第1列
    if idx['device_name'] < 0 and clean_headers and '序号' in (clean_headers[0] or ''):
        if len(clean_headers) > 1 and clean_headers[1]:
            idx['device_name'] = 1

    conn = get_db()
    c = conn.cursor()

    # 同一供应商的旧版本全部标记为非活跃
    c.execute("""
        UPDATE versions SET is_active = 0
        WHERE contract_id = ? AND supplier_name = ? AND is_active = 1
    """, (contract_id, supplier_name))

    # 创建新版本（带供应商名称，默认 is_active=1）
    c.execute("INSERT INTO versions (contract_id, supplier_name, uploader) VALUES (?, ?, '管理员')",
              (contract_id, supplier_name))
    version_id = c.lastrowid

    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        device_name = safe_str(row[idx['device_name']]) if idx['device_name'] >= 0 and idx['device_name'] < len(row) else ''

        # 跳过汇总行（合计/总计/小计）
        if device_name and any(kw in device_name for kw in ['合计', '总计', '小计', 'sum', 'total']):
            skipped += 1
            continue

        if not device_name:
            skipped += 1
            continue

        specs_full = safe_str(row[idx['specs_full']]) if idx['specs_full'] >= 0 and idx['specs_full'] < len(row) else ''
        from compare_engine import extract_specs
        parsed = extract_specs(specs_full)

        device_model = safe_str(row[idx['device_model']]) if idx['device_model'] >= 0 and idx['device_model'] < len(row) else ''

        # 如果model为空，尝试从device_name分离
        if not device_model and device_name:
            parts = device_name.split()
            if len(parts) >= 2:
                device_model = parts[-1]

        # 构建原始列数据(JSON)
        raw_data = {}
        for i, h in enumerate(raw_headers):
            if h and i < len(row):
                raw_data[h] = safe_str(row[i])

        qty = safe_float(row[idx['qty']]) if idx['qty'] >= 0 and idx['qty'] < len(row) else 0
        unit_price = safe_float(row[idx['unit_price']]) if idx['unit_price'] >= 0 and idx['unit_price'] < len(row) else 0
        amount = safe_float(row[idx['amount']]) if idx['amount'] >= 0 and idx['amount'] < len(row) else 0
        # 金额兜底：公式单元格读不到缓存值时，用 数量×单价 计算
        if amount <= 0 and qty > 0 and unit_price > 0:
            amount = round(qty * unit_price, 2)

        c.execute("""
            INSERT INTO supplier_items
            (contract_id, version_id, device_name, device_model, specs_full, specs_cpu, specs_memory,
             specs_disk, specs_other, quote_qty, quote_unit, quote_unit_price, quote_amount, remark, raw_columns)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            contract_id, version_id,
            device_name,
            device_model,
            specs_full,
            parsed.get('cpu', ''),
            parsed.get('memory', ''),
            parsed.get('disk', ''),
            parsed.get('other', ''),
            qty,
            safe_str(row[idx['unit']]) if idx['unit'] >= 0 and idx['unit'] < len(row) else '',
            unit_price,
            amount,
            safe_str(row[idx['remark']]) if idx['remark'] >= 0 and idx['remark'] < len(row) else '',
            json.dumps(raw_data, ensure_ascii=False)
        ))
        count += 1

    # 生成并存储列对齐关系（主合同列 ↔ 供应商列）
    column_mapping = None
    try:
        contract_headers = _extract_contract_headers(conn, contract_id)
        column_mapping = build_column_mapping(contract_headers, raw_headers)
        c.execute("UPDATE versions SET column_mapping=? WHERE id=?",
                  (json.dumps(column_mapping, ensure_ascii=False), version_id))
    except Exception:
        column_mapping = None

    conn.commit()
    conn.close()

    # 自动执行比对
    result = run_comparison(contract_id, version_id)
    result['version_id'] = version_id
    result['supplier_count'] = count
    result['skipped_rows'] = skipped
    result['raw_headers'] = raw_headers
    result['matched_columns'] = {k: (raw_headers[v] if v >= 0 else '❌未匹配') for k, v in idx.items()}
    result['column_mapping'] = column_mapping or {}
    return result


def export_report(version_id: int) -> str:
    """导出整改报告Excel，返回文件路径"""
    conn = get_db()
    c = conn.cursor()

    v = dict(c.execute("SELECT * FROM versions WHERE id = ?", (version_id,)).fetchone())

    results = c.execute("""
        SELECT
            r.id, r.match_status, r.anomaly_types, r.anomaly_detail,
            r.qty_diff, r.param_diff,
            ct.device_name as ct_name, ct.device_model as ct_model,
            ct.specs_full as ct_specs, ct.contract_qty, ct.contract_unit,
            ct.contract_unit_price, ct.contract_amount,
            sp.device_name as sp_name, sp.device_model as sp_model,
            sp.specs_full as sp_specs, sp.quote_qty, sp.quote_unit,
            sp.quote_unit_price, sp.quote_amount
        FROM comparison_results r
        LEFT JOIN contract_items ct ON r.contract_item_id = ct.id
        LEFT JOIN supplier_items sp ON r.supplier_item_id = sp.id
        WHERE r.version_id = ?
        ORDER BY r.match_status, ct.device_name
    """, (version_id,)).fetchall()

    conn.close()

    wb = openpyxl.Workbook()

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='center', wrap_text=True)
    red_font = Font(name='微软雅黑', color='FF4444', size=10)
    green_font = Font(name='微软雅黑', color='00CC88', size=10)
    orange_font = Font(name='微软雅黑', color='FF8800', size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Sheet1: 整改报告总览
    ws1 = wb.active
    ws1.title = "整改报告总览"

    ws1.merge_cells('A1:H1')
    ws1['A1'] = '供应商报价整改报告'
    ws1['A1'].font = Font(name='微软雅黑', bold=True, size=16, color='1a1a2e')

    overview_data = [
        ('版本号', v['id']),
        ('上传时间', v['upload_time']),
        ('总条目数', v['total_items']),
        ('成功匹配', v['matched_count']),
        ('匹配异常', v['anomaly_count']),
        ('待采购漏报', v['pending_count']),
        ('供应商增项', v['extra_count']),
        ('整体采购进度', f"{v['progress']}%"),
    ]
    for i, (label, value) in enumerate(overview_data, 3):
        ws1[f'A{i}'] = label
        ws1[f'A{i}'].font = Font(name='微软雅黑', bold=True, size=10)
        ws1[f'B{i}'] = value
        ws1[f'B{i}'].font = Font(name='微软雅黑', size=10)

    row = 12
    ws1[f'A{row}'] = '整改提示：'
    ws1[f'A{row}'].font = Font(name='微软雅黑', bold=True, size=11, color='CC0000')
    ws1.merge_cells(f'A{row+1}:H{row+1}')
    ws1[f'A{row+1}'] = '请根据以下明细逐项核实修改：\n1. 匹配异常条目 → 修正型号、参数或数量后重新报价\n2. 待采购漏报条目 → 补充报价\n3. 供应商增项条目 → 确认是否需要保留或删除\n修改完成后请重新提交报价文件。'
    ws1[f'A{row+1}'].alignment = Alignment(wrap_text=True, vertical='top')

    # Sheet2: 异常明细
    ws2 = wb.create_sheet("异常明细")

    headers2 = ['设备名称', '型号规格', '合同数量', '报价数量', '合同参数', '报价参数', '异常类型', '异常详情']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for r in results:
        ws2.cell(row=row, column=1, value=r['ct_name'] or r['sp_name'] or '').border = thin_border
        ws2.cell(row=row, column=2, value=r['ct_model'] or r['sp_model'] or '').border = thin_border
        ws2.cell(row=row, column=3, value=r['contract_qty'] or '').border = thin_border
        ws2.cell(row=row, column=4, value=r['quote_qty'] or '').border = thin_border
        ws2.cell(row=row, column=5, value=r['ct_specs'] or '').border = thin_border
        ws2.cell(row=row, column=6, value=r['sp_specs'] or '').border = thin_border

        try:
            types = json.loads(r['anomaly_types'])
        except:
            types = [r['anomaly_types']]
        ws2.cell(row=row, column=7, value=', '.join(types)).border = thin_border
        ws2.cell(row=row, column=8, value=r['anomaly_detail']).border = thin_border

        status = r['match_status']
        if status == '匹配异常':
            for c_idx in range(1, 9):
                ws2.cell(row=row, column=c_idx).font = orange_font
        elif status == '待采购':
            for c_idx in range(1, 9):
                ws2.cell(row=row, column=c_idx).font = red_font
        elif status == '供应商增项':
            for c_idx in range(1, 9):
                ws2.cell(row=row, column=c_idx).font = Font(name='微软雅黑', color='9944FF', size=10)

        row += 1

    for ws in [ws1, ws2]:
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 16

    filepath = os.path.join(UPLOAD_DIR, f'整改报告_v{version_id}.xlsx')
    wb.save(filepath)
    return filepath
