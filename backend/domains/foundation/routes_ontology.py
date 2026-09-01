"""基础支撑域 · 原子本体 MCP（mcp/ontology）只读网关（R2 split from main.py）。"""
import os

from fastapi import APIRouter

from common.privacy import is_privacy_header, filter_privacy_headers
from common.datasource_meta import _load_ds_meta, _ds_latest_path

router = APIRouter(prefix="/api/mcp/ontology", tags=["foundation-ontology"])


@router.get("/tables")
def mcp_ontology_tables():
    """列出原始本体表（只读）"""
    meta = _load_ds_meta()
    return {'tables': [{'name': k, 'version_count': len(v.get('versions', []))} for k, v in meta.items()]}


@router.get("/schema")
def mcp_ontology_schema(table_name: str):
    """获取数据表结构（原子只读）：返回所有列名及示例值，供智能体了解字段含义。"""
    import openpyxl
    fpath = _ds_latest_path(table_name)
    if not fpath:
        return {'success': False, 'error': f'表「{table_name}」未上传'}
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) if h is not None else '' for h in rows[0]]
    keep_idx = filter_privacy_headers(headers)
    s1 = rows[1] if len(rows) > 1 else None
    s2 = rows[2] if len(rows) > 2 else None
    columns = []
    for i in keep_idx:
        h = headers[i]
        ex = []
        if s1 is not None and i < len(s1) and s1[i] is not None:
            ex.append(str(s1[i])[:24])
        if s2 is not None and i < len(s2) and s2[i] is not None:
            ex.append(str(s2[i])[:24])
        columns.append({'name': h, 'example': ex})
    return {'table_name': table_name, 'columns': columns, 'column_count': len(columns)}


@router.get("/query")
def mcp_ontology_query(table_name: str, keyword: str = '', time_column: str = '',
                       start_date: str = '', end_date: str = '', columns: str = '', limit: int = 100):
    """查询原始明细（原子只读）：关键词模糊匹配、时间范围过滤、列投影。columns 为逗号分隔列名，空则返回全部列。"""
    import openpyxl
    from datetime import datetime, timedelta
    fpath = _ds_latest_path(table_name)
    if not fpath:
        return {'success': False, 'error': f'表「{table_name}」未上传'}
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) if h is not None else '' for h in rows[0]]
    keep_idx = filter_privacy_headers(headers)
    col_idx = {h: i for i, h in enumerate(headers)}

    selected = [c.strip() for c in columns.split(',') if c.strip() and not is_privacy_header(c.strip())] if columns else []

    def parse_date(v):
        if v is None or v == '': return None
        if isinstance(v, datetime): return v
        s = str(v).strip()
        try: return datetime.strptime(s[:10], '%Y-%m-%d')
        except Exception: pass
        try:
            n = float(s)
            if 20000 < n < 80000:
                return datetime(1899, 12, 30) + timedelta(days=int(n))
        except Exception: pass
        return None

    start = parse_date(start_date)
    end = parse_date(end_date)
    tc = col_idx.get(time_column) if time_column else None

    kw = keyword.strip().lower()
    data_rows = []
    for r in rows[1:]:
        if r is None: continue
        if kw:
            hit = any(kw in str(v).lower() for v in r if v is not None)
            if not hit: continue
        if tc is not None:
            dt = parse_date(r[tc] if tc < len(r) else None)
            if dt is None: continue
            if start and dt < start: continue
            if end and dt > end: continue
        data_rows.append([str(v) if v is not None else '' for v in r])
        if len(data_rows) >= limit: break

    if selected:
        sel_idx = [col_idx[c] for c in selected if c in col_idx]
        out_headers = [headers[i] for i in sel_idx]
        out_rows = [[row[i] if i < len(row) else '' for i in sel_idx] for row in data_rows]
    else:
        out_headers = [headers[i] for i in keep_idx]
        out_rows = [[row[i] if i < len(row) else '' for i in keep_idx] for row in data_rows]

    return {'table_name': table_name, 'headers': out_headers, 'rows': out_rows, 'count': len(out_rows)}