"""基础支撑域 · 数据源路由（原 main.py /api/datasource/*）。"""
import os

from fastapi import APIRouter, Query, File, UploadFile
from fastapi.responses import JSONResponse

from common.paths import DATASOURCE_DIR
from common.privacy import is_privacy_header, filter_privacy_headers, sanitize_excel_file
from common.datasource_meta import _load_ds_meta, _save_ds_meta, _ensure_table

router = APIRouter(prefix="/api/datasource", tags=["foundation-datasource"])


@router.get("/tables")
def datasource_tables():
    """列出所有数据表及其最新版本摘要"""
    meta = _load_ds_meta()
    tables = []
    for tname, tdata in meta.items():
        vers = tdata.get('versions', [])
        latest = vers[0] if vers else None
        tables.append({
            'name': tname,
            'version_count': len(vers),
            'latest_id': latest['id'] if latest else None,
            'latest_time': latest['upload_time'] if latest else None,
            'latest_rows': latest['row_count'] if latest else 0,
            'latest_columns': [c for c in (latest['columns'] or []) if not is_privacy_header(c)],
        })
    return {'tables': tables}


@router.post("/upload")
async def datasource_upload(file: UploadFile = File(...), table_name: str = Query(...)):
    """上传原始数据表（Excel），按表名隔离版本"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        return JSONResponse({'success': False, 'error': '仅支持 .xlsx / .xls 文件'})
    meta = _load_ds_meta()
    tdata = _ensure_table(meta, table_name)
    vid = tdata['next_id']
    fname = f'{table_name}_v{vid}.xlsx'
    fpath = os.path.join(DATASOURCE_DIR, fname)
    content = await file.read()
    with open(fpath, 'wb') as f:
        f.write(content)
    # 客户名/客户简称/项目名等敏感列一律删除，不进入系统
    sanitize_excel_file(fpath)
    import openpyxl
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]] if rows else []
    row_count = len(rows) - 1
    wb.close()
    from datetime import datetime
    ver = {
        'id': vid,
        'filename': file.filename,
        'upload_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'row_count': row_count,
        'columns': headers,
        'file': fname,
    }
    tdata['versions'].insert(0, ver)
    tdata['next_id'] = vid + 1
    _save_ds_meta(meta)
    return {'success': True, 'version': ver, 'table_name': table_name}


@router.get("/versions")
def datasource_versions(table_name: str = Query(...)):
    """获取指定表的所有版本"""
    meta = _load_ds_meta()
    tdata = meta.get(table_name, {})
    vers = tdata.get('versions', [])
    return {'table_name': table_name, 'versions': vers, 'latest_id': vers[0]['id'] if vers else None}


@router.get("/latest")
def datasource_latest(table_name: str = Query(...)):
    """获取指定表最新版本的数据预览"""
    meta = _load_ds_meta()
    tdata = meta.get(table_name, {})
    vers = tdata.get('versions', [])
    if not vers:
        return {'version': None, 'headers': [], 'rows': [], 'row_count': 0}
    latest = vers[0]
    fpath = os.path.join(DATASOURCE_DIR, latest['file'])
    if not os.path.exists(fpath):
        return {'version': latest, 'headers': [], 'rows': [], 'row_count': 0, 'error': '文件丢失'}
    import openpyxl
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows_iter, [])]
    keep_idx = filter_privacy_headers(headers)
    headers = [headers[i] for i in keep_idx]
    preview_rows = []
    for i, row in enumerate(rows_iter):
        if i >= 20:
            break
        preview_rows.append([str(v) if v is not None else '' for v in [row[i] for i in keep_idx if i < len(row)]])
    wb.close()
    return {'version': latest, 'headers': headers, 'rows': preview_rows, 'row_count': latest['row_count']}


@router.delete("/version/{table_name}/{version_id}")
def datasource_delete_version(table_name: str, version_id: int):
    """删除指定表的指定版本"""
    meta = _load_ds_meta()
    tdata = meta.get(table_name)
    if not tdata:
        return JSONResponse({'success': False, 'error': '表不存在'}, status_code=404)
    target = None
    for v in tdata.get('versions', []):
        if v['id'] == version_id:
            target = v
            break
    if not target:
        return JSONResponse({'success': False, 'error': '版本不存在'}, status_code=404)
    fpath = os.path.join(DATASOURCE_DIR, target['file'])
    if os.path.exists(fpath):
        os.remove(fpath)
    tdata['versions'] = [v for v in tdata['versions'] if v['id'] != version_id]
    if not tdata['versions']:
        del meta[table_name]
    _save_ds_meta(meta)
    return {'success': True}