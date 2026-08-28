"""生命周期域 · PLM /api/plm (R2 split).
注：放宽迁移，个别依赖 main 的内部符号若缺失属可接受范围。"""
from typing import Optional, Any, List, Dict, Union
import io, os, json, re
from fastapi import APIRouter, Query, File, UploadFile, Form, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, Response
from fastapi import File as FFile
from fastapi.responses import Response as _RawResponse
import urllib.parse
from plm_models import *
import plm_models as plm
from models import get_db

router = APIRouter(prefix="", tags=["lifecycle-plm"])

def _plm_op(payload=None):
    """取操作人（本期无鉴权，仅留痕）。"""
    if isinstance(payload, dict) and payload.get('operator'):
        return str(payload['operator'])[:64]
    return 'admin'

def _plm_ret(res):
    """统一响应包装：读接口 → {success, data}；写接口透出业务结果。"""
    if isinstance(res, dict) and 'success' in res:
        return res
    return {'success': True, 'data': res}

@router.get("/api/plm/overview")
def api_plm_overview():
    return _plm_ret(plm.overview())

@router.get("/api/plm/dict")
def api_plm_dict(category: Optional[str] = None):
    return _plm_ret(plm.list_dict(category))

@router.post("/api/plm/dict")
def api_plm_dict_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_dict(payload.get('category', ''), payload.get('key', ''),
                                    payload.get('label', ''), payload.get('sort', 0),
                                    payload.get('remark', ''), _plm_op(payload)))

@router.delete("/api/plm/dict/{dict_id}")
def api_plm_dict_delete(dict_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_dict(dict_id, operator))

@router.get("/api/plm/config")
def api_plm_config():
    return _plm_ret(plm.list_config())

@router.put("/api/plm/config")
def api_plm_config_update(payload: Dict[str, Any]):
    key = payload.get('key')
    if not key:
        return {'success': False, 'error': 'key 必填'}
    return _plm_ret({'key': key,
                     'value': plm.set_config(key, payload.get('value', ''),
                                             payload.get('description', ''), _plm_op(payload))})

@router.get("/api/plm/logs")
def api_plm_logs(target_type: Optional[str] = None, target_id: Optional[str] = None,
                 limit: int = Query(200, le=1000)):
    return _plm_ret(plm.list_logs(target_type, target_id, limit))

@router.get("/api/plm/opportunities")
def api_plm_opp_list(keyword: Optional[str] = None, status: Optional[str] = None):
    return _plm_ret(plm.list_opportunities(keyword, status))

@router.post("/api/plm/opportunities")
def api_plm_opp_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_opportunity(payload, _plm_op(payload)))

@router.get("/api/plm/opportunities/{opp_id}")
def api_plm_opp_get(opp_id: int):
    r = plm.get_opportunity(opp_id)
    if not r:
        return JSONResponse({'success': False, 'error': '商机不存在'}, status_code=404)
    return _plm_ret(r)

@router.put("/api/plm/opportunities/{opp_id}")
def api_plm_opp_update(opp_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_opportunity(opp_id, payload, _plm_op(payload)))

@router.delete("/api/plm/opportunities/{opp_id}")
def api_plm_opp_delete(opp_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_opportunity(opp_id, operator))

@router.post("/api/plm/opportunities/{opp_id}/follow")
def api_plm_opp_follow(opp_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.add_follow_record(opp_id, payload.get('content', ''),
                                          _plm_op(payload), payload.get('time')))

@router.get("/api/plm/opportunities/{opp_id}/estimate")
def api_plm_opp_estimate(opp_id: int):
    return _plm_ret(plm.get_opportunity_estimate(opp_id))

@router.post("/api/plm/opportunities/{opp_id}/estimate")
def api_plm_opp_estimate_save(opp_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.save_opportunity_estimate(opp_id, payload, _plm_op(payload)))

@router.get("/api/plm/opportunities/{opp_id}/docs")
def api_plm_opp_docs(opp_id: int):
    return _plm_ret(plm.list_presale_docs(opp_id))

@router.post("/api/plm/opportunities/{opp_id}/docs")
def api_plm_opp_doc_create(opp_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p['opportunity_id'] = opp_id
    return _plm_ret(plm.create_presale_doc(p, _plm_op(payload)))

@router.delete("/api/plm/presale-docs/{doc_id}")
def api_plm_opp_doc_delete(doc_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_presale_doc(doc_id, operator))

@router.post("/api/plm/opportunities/convert")
def api_plm_convert(payload: Dict[str, Any]):
    return _plm_ret(plm.convert_opportunity(payload, _plm_op(payload)))

@router.get("/api/plm/contracts")
def api_plm_ct_list(keyword: Optional[str] = None):
    return _plm_ret(plm.list_contracts(keyword))

@router.post("/api/plm/contracts")
def api_plm_ct_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_contract(payload, _plm_op(payload)))

@router.get("/api/plm/contracts/{contract_id}")
def api_plm_ct_get(contract_id: int):
    r = plm.get_contract(contract_id)
    if not r:
        return JSONResponse({'success': False, 'error': '合同不存在'}, status_code=404)
    return _plm_ret(r)

@router.put("/api/plm/contracts/{contract_id}")
def api_plm_ct_update(contract_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_contract(contract_id, payload, _plm_op(payload)))

@router.delete("/api/plm/contracts/{contract_id}")
def api_plm_ct_delete(contract_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_contract(contract_id, operator))

@router.get("/api/plm/projects")
def api_plm_proj_list(keyword: Optional[str] = None, status: Optional[str] = None):
    return _plm_ret(plm.list_projects(keyword, status))

@router.post("/api/plm/projects")
def api_plm_proj_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_project(payload, _plm_op(payload)))

@router.get("/api/plm/projects/{project_id}")
def api_plm_proj_get(project_id: int):
    r = plm.get_project(project_id)
    if not r:
        return JSONResponse({'success': False, 'error': '项目不存在'}, status_code=404)
    return _plm_ret(r)

@router.put("/api/plm/projects/{project_id}")
def api_plm_proj_update(project_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_project(project_id, payload, _plm_op(payload)))

@router.delete("/api/plm/projects/{project_id}")
def api_plm_proj_delete(project_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_project(project_id, operator))

@router.get("/api/plm/projects/{project_id}/panorama")
def api_plm_proj_panorama(project_id: int):
    r = plm.project_panorama(project_id)
    if not r:
        return JSONResponse({'success': False, 'error': '项目不存在'}, status_code=404)
    return _plm_ret(r)

@router.get("/api/plm/projects/{project_id}/progress")
def api_plm_proj_progress(project_id: int):
    return _plm_ret(plm.project_progress(project_id))

@router.get("/api/plm/projects/{project_id}/finance")
def api_plm_proj_finance(project_id: int):
    return _plm_ret(plm.project_finance(project_id))

@router.get("/api/plm/projects/{project_id}/baselines")
def api_plm_baseline_list(project_id: int):
    return _plm_ret(plm.list_baselines(project_id=project_id))

@router.get("/api/plm/projects/{project_id}/baseline-compare")
def api_plm_baseline_compare(project_id: int):
    return _plm_ret(plm.compare_baselines(project_id))

@router.post("/api/plm/projects/{project_id}/baselines")
def api_plm_baseline_save(project_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p.setdefault('project_id', project_id)
    p.setdefault('scope_type', 'project')
    p.setdefault('scope_id', project_id)
    return _plm_ret(plm.save_baseline(p, _plm_op(payload)))

@router.get("/api/plm/baselines/{baseline_id}")
def api_plm_baseline_get(baseline_id: int):
    r = plm.get_baseline(baseline_id)
    if not r:
        return JSONResponse({'success': False, 'error': '基线不存在'}, status_code=404)
    return _plm_ret(r)

@router.put("/api/plm/baselines/{baseline_id}")
def api_plm_baseline_update(baseline_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p['id'] = baseline_id
    return _plm_ret(plm.save_baseline(p, _plm_op(payload)))

@router.post("/api/plm/baselines/{baseline_id}/confirm")
def api_plm_baseline_confirm(baseline_id: int, payload: Dict[str, Any] = None):
    return _plm_ret(plm.confirm_baseline(baseline_id, _plm_op(payload)))

@router.post("/api/plm/baselines/{baseline_id}/lock")
def api_plm_baseline_lock(baseline_id: int, payload: Dict[str, Any] = None):
    return _plm_ret(plm.lock_baseline(baseline_id, _plm_op(payload)))

@router.delete("/api/plm/baselines/{baseline_id}")
def api_plm_baseline_delete(baseline_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_baseline(baseline_id, operator))

@router.get("/api/plm/projects/{project_id}/milestones")
def api_plm_ms_list(project_id: int):
    return _plm_ret(plm.list_milestones(project_id))

@router.get("/api/plm/milestones")
def api_plm_ms_all(keyword: Optional[str] = None):
    """跨项目里程碑全量列表（供「项目管理·里程碑」视图）。"""
    return _plm_ret(plm.list_all_milestones(keyword))

@router.post("/api/plm/milestones/import")
async def api_plm_ms_import(file: UploadFile = File(...), operator: str = Query('admin')):
    """上传「项目里程碑表」xlsx，按列映射写入 plm_milestone（项目号/合同号关联）。"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception as e:
        return {'success': False, 'error': '文件解析失败：%s' % e}
    ws = wb.active
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or '').strip())
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == '' for v in r):
            continue
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(r):
                d[h] = r[i]
        rows.append(d)
    res = plm.import_milestones(rows, operator)
    res['headers'] = headers
    return _plm_ret(res)

@router.post("/api/plm/projects/{project_id}/milestones")
def api_plm_ms_create(project_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p.setdefault('project_id', project_id)
    return _plm_ret(plm.create_milestone(p, _plm_op(payload)))

@router.put("/api/plm/milestones/{milestone_id}")
def api_plm_ms_update(milestone_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_milestone(milestone_id, payload, _plm_op(payload)))

@router.delete("/api/plm/milestones/{milestone_id}")
def api_plm_ms_delete(milestone_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_milestone(milestone_id, operator))

@router.get("/api/plm/projects/{project_id}/tasks")
def api_plm_task_list(project_id: int, milestone_id: Optional[int] = None):
    return _plm_ret(plm.list_tasks(project_id, milestone_id))

@router.post("/api/plm/tasks")
def api_plm_task_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_task(payload, _plm_op(payload)))

@router.get("/api/plm/tasks/{task_id}")
def api_plm_task_get(task_id: int):
    r = plm.get_task(task_id)
    if not r:
        return JSONResponse({'success': False, 'error': '任务不存在'}, status_code=404)
    return _plm_ret(r)

@router.put("/api/plm/tasks/{task_id}")
def api_plm_task_update(task_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_task(task_id, payload, _plm_op(payload)))

@router.delete("/api/plm/tasks/{task_id}")
def api_plm_task_delete(task_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_task(task_id, operator))

@router.get("/api/plm/staff")
def api_plm_staff_list(keyword: Optional[str] = None, status: Optional[str] = None):
    return _plm_ret(plm.list_staff(keyword, status))

@router.get("/api/plm/staff/load")
def api_plm_staff_load():
    return _plm_ret(plm.staff_load())

@router.post("/api/plm/staff")
def api_plm_staff_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_staff(payload, _plm_op(payload)))

@router.get("/api/plm/staff/{staff_id}")
def api_plm_staff_get(staff_id: int):
    r = plm.get_staff(staff_id)
    if not r:
        return JSONResponse({'success': False, 'error': '人员不存在'}, status_code=404)
    return _plm_ret(r)

@router.put("/api/plm/staff/{staff_id}")
def api_plm_staff_update(staff_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_staff(staff_id, payload, _plm_op(payload)))

@router.delete("/api/plm/staff/{staff_id}")
def api_plm_staff_delete(staff_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_staff(staff_id, operator))

@router.get("/api/plm/assignments")
def api_plm_asg_list(project_id: Optional[int] = None, staff_id: Optional[int] = None,
                     status: Optional[str] = None):
    return _plm_ret(plm.list_assignments(project_id, staff_id, status))

@router.post("/api/plm/assignments")
def api_plm_asg_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_assignment(payload, _plm_op(payload)))

@router.put("/api/plm/assignments/{assign_id}")
def api_plm_asg_update(assign_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_assignment(assign_id, payload, _plm_op(payload)))

@router.delete("/api/plm/assignments/{assign_id}")
def api_plm_asg_delete(assign_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_assignment(assign_id, operator))

@router.get("/api/plm/timesheets")
def api_plm_ts_list(project_id: Optional[int] = None, staff_id: Optional[int] = None):
    return _plm_ret(plm.list_timesheets(project_id, staff_id))

@router.post("/api/plm/timesheets")
def api_plm_ts_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_timesheet(payload, _plm_op(payload)))

@router.put("/api/plm/timesheets/{ts_id}")
def api_plm_ts_update(ts_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_timesheet(ts_id, payload, _plm_op(payload)))

@router.delete("/api/plm/timesheets/{ts_id}")
def api_plm_ts_delete(ts_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_timesheet(ts_id, operator))

@router.post("/api/plm/timesheets/sync")
def api_plm_ts_sync(payload: Dict[str, Any] = None):
    p = payload or {}
    return _plm_ret(plm.sync_labor_cost(p.get('project_id'), p.get('staff_id')))

@router.get("/api/plm/ledger")
def api_plm_ledger_list(project_id: Optional[int] = None, kind: Optional[str] = None,
                        category: Optional[str] = None, source: Optional[str] = None):
    return _plm_ret(plm.list_ledger(project_id, kind, category, source))

@router.post("/api/plm/ledger")
def api_plm_ledger_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_ledger(payload, _plm_op(payload)))

@router.put("/api/plm/ledger/{ledger_id}")
def api_plm_ledger_update(ledger_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_ledger(ledger_id, payload, _plm_op(payload)))

@router.delete("/api/plm/ledger/{ledger_id}")
def api_plm_ledger_delete(ledger_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_ledger(ledger_id, operator))

@router.get("/api/plm/alert-rules")
def api_plm_rule_list():
    return _plm_ret(plm.list_alert_rules())

@router.put("/api/plm/alert-rules/{rule_key}")
def api_plm_rule_update(rule_key: str, payload: Dict[str, Any]):
    return _plm_ret(plm.update_alert_rule(rule_key, payload, _plm_op(payload)))

@router.get("/api/plm/alerts")
def api_plm_alert_list(project_id: Optional[int] = None, dim: Optional[str] = None,
                       status: Optional[str] = None, level: Optional[str] = None):
    return _plm_ret(plm.list_alerts(project_id, dim, status, level))

@router.post("/api/plm/alerts/scan")
def api_plm_alert_scan(payload: Dict[str, Any] = None):
    return _plm_ret(plm.scan_alerts(_plm_op(payload)))

@router.put("/api/plm/alerts/{alert_id}/handle")
def api_plm_alert_handle(alert_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.handle_alert(alert_id, payload, _plm_op(payload)))

@router.get("/api/plm/export/{report}")
def api_plm_export(report: str, project_id: Optional[int] = None):
    try:
        filename, data = plm.export_report(report, project_id)
    except ValueError as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)
    return _RawResponse(
        content=data,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': "attachment; filename*=UTF-8''%s"
                                        % urllib.parse.quote(filename)})
