"""
合同比对系统 — FastAPI 主应用（多合同版）
端口: 9006
"""

from fastapi import FastAPI, UploadFile, File, Query, Form, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, Response
import io
import os
import json
import re
import shutil
import urllib.parse
from openpyxl.utils import get_column_letter
import httpx

from models import init_db, get_db, create_contract, delete_contract, update_contract_status
from compare_engine import run_comparison
from excel_handler import import_contract_excel, import_supplier_excel, export_report, reapply_column_mapping
from procurement_models import init_procurement_db, seed_procurement_master
from services.etl import ETL_JOB_DEFS, _register_etl_jobs
app = FastAPI(title="合同比对系统（多合同版）", version="2.0")
from fastapi.middleware.cors import CORSMiddleware
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── R2 逐域路由拆分（behavior 不变，仅 registrar include）──
from domains.foundation.routes_datasource import router as foundation_ds_router
from domains.foundation.routes_ontology import router as foundation_ontology_router
from domains.finance.routes_gross import router as finance_gross_router
from domains.finance.routes_payment import router as finance_payment_router
app.include_router(foundation_ds_router)
app.include_router(foundation_ontology_router)
app.include_router(finance_gross_router)
app.include_router(finance_payment_router)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, '..', 'uploads')
FRONTEND_DIR = os.path.join(BASE_DIR, '..', 'frontend')
DATASOURCE_DIR = os.path.join(BASE_DIR, '..', 'datasource')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(DATASOURCE_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────
# CC-011 前端资源缓存策略
# 页面与静态资源此前是裸引用（/common.css、/plm.app.js …），既无版本位也无
# Cache-Control，浏览器会按启发式新鲜度直接复用旧副本、根本不回访 —— 导致
# 每次前端发版都要用户强制刷新，且 HTML 与 CSS 新旧不同步时出现"结构对、
# 样式没套上"的半新状态。
# 这里用 no-cache（每次再校验）而不是 no-store（禁用缓存、每次全量重传）：
# FileResponse 本就输出 ETag / Last-Modified，未变更时浏览器拿 304（数百字节）。
# 命中范围 = 显式清单 + 后缀兜底，后续新增 .css/.js 无需再改这里。
# ─────────────────────────────────────────────────────────────
NO_CACHE_PATHS = frozenset((
    '/', '/gross', '/plm', '/procurement',
    '/common.css', '/plm.app.js', '/procurement.app.js', '/china.json',
))
NO_CACHE_SUFFIXES = ('.css', '.js')


def _etag_match(header_value, etag):
    """按 RFC 7232 语义比对 If-None-Match：支持 `*`、逗号列表与 W/ 弱标记。"""
    if header_value.strip() == '*':
        return True
    want = etag.removeprefix('W/').strip()
    return any(tag.strip().removeprefix('W/') == want for tag in header_value.split(','))


@app.middleware("http")
async def no_cache_static_assets(request: Request, call_next):
    """页面与静态资源每次再校验；命中指纹则回 304，避免全量重传。动态接口原样透传。"""
    response = await call_next(request)
    path = request.url.path
    if not (path in NO_CACHE_PATHS or path.endswith(NO_CACHE_SUFFIXES)):
        return response
    response.headers['Cache-Control'] = 'no-cache'
    # FileResponse 只发 ETag、不处理 If-None-Match（只有 StaticFiles 会），
    # 所以这里自己短路，否则 no-cache 会让每次导航都全量重下 170KB 的 index.html。
    etag = response.headers.get('etag')
    inm = request.headers.get('if-none-match')
    if etag and inm and response.status_code == 200 and _etag_match(inm, etag):
        headers = {'Cache-Control': 'no-cache', 'ETag': etag}
        if response.headers.get('last-modified'):
            headers['Last-Modified'] = response.headers['last-modified']
        return Response(status_code=304, headers=headers)
    return response

# ── neuops 智能体网关（emp-008 采购询比价）──
NEUOPS_BASE = os.getenv("NEUOPS_BASE", "http://127.0.0.1:9007")


def trigger_neuops(path: str, payload: dict, timeout: float = 15.0) -> dict:
    """调用 neuops 智能体 trigger API。失败不阻断主流程，返回 trigger 结果。"""
    import copy
    p = copy.deepcopy(payload)
    # 清理空 dict 字段（neuops Pydantic Optional[Model] 遇到 {} 会报 required）
    if isinstance(p, dict) and isinstance(p.get("selected_supplier"), dict) and not p["selected_supplier"]:
        p["selected_supplier"] = None
    try:
        r = httpx.post(f"{NEUOPS_BASE}/api/procurement-agent/{path}",
                       json=p, timeout=timeout)
        return r.json()
    except Exception as e:
        return {"success": False, "error": f"neuops trigger 失败: {type(e).__name__}: {e}"}

# ─────────────────────────────────────────────────────────────
# 客户/项目敏感信息：一律丢弃，不导入数据库、不展示（只保留合同编号）
# 匹配列名：甲方名称/客户名称/客户简称/客户分类/最终用户/项目名称/项目描述等
# 例外：客户标识（脱敏键，如 QDHEKJ）保留，用于多维度聚合，不视为敏感列
# ─────────────────────────────────────────────────────────────
PRIVACY_HEADER_PATTERN = re.compile(
    r"(甲方|客户|业主|招标人|采购人|建设单位|使用单位|最终用户)"
    r"(名称|简称|全称|分类|编号)?$|"
    r"^(项目名称|项目描述|项目简介|合同名称|合同名|合同标题)$|"
    r"主合同客户名称|关键客户"
)


def is_privacy_header(h):
    """判断列名是否为客户/项目敏感信息"""
    return bool(h) and bool(PRIVACY_HEADER_PATTERN.search(str(h)))


def filter_privacy_headers(headers):
    """过滤掉敏感列，返回保留的列索引（数据列下标）"""
    keep = []
    for i, h in enumerate(headers):
        if not is_privacy_header(h):
            keep.append(i)
    return keep


def sanitize_excel_file(path):
    """就地删除 Excel 文件中所有 sheet 的敏感列（客户名/客户简称/项目名等）。

    返回 (changed, dropped_cols)。仅支持 .xlsx；.xls 无法就地改写时返回 changed=False。
    """
    dropped_cols = set()
    if not str(path).lower().endswith('.xlsx'):
        return False, []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
    except Exception:
        return False, []
    modified = False
    for ws in wb.worksheets:
        headers = [str(c.value) if c.value is not None else '' for c in ws[1]]
        drop_idx = [i for i, h in enumerate(headers) if is_privacy_header(h)]
        if not drop_idx:
            continue
        dropped_cols.update(headers[i] for i in drop_idx)
        # 从后往前删除列，避免索引错位
        for i in sorted(drop_idx, reverse=True):
            ws.delete_cols(i + 1, 1)
        modified = True
    if modified:
        wb.save(path)
    wb.close()
    return modified, sorted(dropped_cols)

# 数据源版本元数据文件
DS_META_FILE = os.path.join(DATASOURCE_DIR, 'versions.json')

def _load_ds_meta():
    if os.path.exists(DS_META_FILE):
        with open(DS_META_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def _save_ds_meta(meta):
    with open(DS_META_FILE, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def _ensure_table(meta, table_name):
    if table_name not in meta:
        meta[table_name] = {'versions': [], 'next_id': 1}
    return meta[table_name]


# ===================== 数据源 API =====================

# ===================== 回款周期分析 API =====================

@app.get("/api/analysis/payment-cycle")
def analysis_payment_cycle():
    """回款周期分析：H表∪R表补充→按月累计计算"""
    import openpyxl
    from datetime import datetime, date
    meta = _load_ds_meta()
    
    # 读取 H 表（总合同表）
    htdata = meta.get('总合同表', {})
    hvers = htdata.get('versions', [])
    if not hvers:
        return {'success': True, 'data': {'source_version': 0, 'months': [], 'icid': {}, 'department': {}, 'zones': [], 'enriched_rows': []}}
    
    hfpath = os.path.join(DATASOURCE_DIR, hvers[0]['file'])
    if not os.path.exists(hfpath):
        return {'success': False, 'error': '总合同表文件丢失'}
    wb = openpyxl.load_workbook(hfpath, read_only=True)
    ws = wb.active
    hrows = list(ws.iter_rows(values_only=True))
    wb.close()
    h_headers = [str(h) if h else '' for h in hrows[0]]
    h_data = []
    for r in hrows[1:]:
        row = {}
        for i, h in enumerate(h_headers):
            row[h] = r[i] if i < len(r) else None
        h_data.append(row)
    
    # 工具函数
    def parse_date(v):
        if v is None or v == '' or v == '-': return None
        if isinstance(v, (datetime, date)): return v
        s = str(v).strip()
        if s.startswith('='): return None
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try: return datetime.strptime(s, fmt)
            except: pass
        return None
    
    def safe_float(v):
        if v is None or v == '' or v == '-': return 0
        s = str(v).strip()
        if s.startswith('='): return 0
        try: return float(v)
        except: return 0
    
    def safe_str(v):
        if v is None: return ''
        s = str(v).strip()
        return '' if s.startswith('=') else s
    
    # 列名匹配
    def find_col(headers, keywords):
        for h in headers:
            hl = h.lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None
    
    h_col_no = find_col(h_headers, ['合同编号', '编号'])
    h_col_date = find_col(h_headers, ['统计日期', '签约日期', '日期'])
    h_col_dept = find_col(h_headers, ['部门', '责任部门'])
    # 修正：'部门' 可能被 '签定部门' 抢先匹配，确认真实的部门列
    if h_col_dept and '签定' in str(h_col_dept):
        for h in h_headers:
            if h == '部门':
                h_col_dept = h
                break
    h_col_amount = find_col(h_headers, ['合同金额', '金额', '合同额', '合同总金额'])
    h_col_region = find_col(h_headers, ['区域', '大区', '片区'])
    # 省份列：优先精确匹配“省”（避免被“省分”抢占）
    h_col_province = None
    for h in h_headers:
        if str(h).strip() in ('省', '省份'):
            h_col_province = h
            break
    if h_col_province is None:
        h_col_province = find_col(h_headers, ['省份'])
    
    # 读取 R 表（项目里程碑表），按合同编号建索引
    rtdata = meta.get('项目里程碑表', {})
    rvers = rtdata.get('versions', [])
    r_index = {}  # 合同编号 → [里程碑行列表]
    if rvers:
        rfpath = os.path.join(DATASOURCE_DIR, rvers[0]['file'])
        if os.path.exists(rfpath):
            wb = openpyxl.load_workbook(rfpath, read_only=True)
            ws = wb.active
            rrows = list(ws.iter_rows(values_only=True))
            wb.close()
            r_headers = [str(h) if h else '' for h in rrows[0]]
            r_col_contract = find_col(r_headers, ['合同编号'])
            for r in rrows[1:]:
                row = {}
                for i, h in enumerate(r_headers):
                    row[h] = r[i] if i < len(r) else None
                cid = str(row.get(r_col_contract, '')).strip() if r_col_contract else ''
                if cid:
                    r_index.setdefault(cid, []).append(row)
    
    # 通用：为指定年份的H表行补充回款周期数据
    def enrich_for_year(h_rows, target_year):
        """返回 enriched 列表"""
        result = []
        for row in h_rows:
            cno = safe_str(row.get(h_col_no, ''))
            sdate = parse_date(row.get(h_col_date))
            dept = safe_str(row.get(h_col_dept, ''))
            region = safe_str(row.get(h_col_region, '')) if h_col_region else ''
            province = safe_str(row.get(h_col_province, '')) if h_col_province else ''
            amount = safe_float(row.get(h_col_amount)) if h_col_amount else 0
            
            if not sdate or sdate.year != target_year:
                continue
            
            matches = r_index.get(cno, [])
            valid = []
            for m in matches:
                pdate = parse_date(m.get('计划回款时间'))
                pval = safe_float(m.get('计划产值(元)'))
                if pdate and pval != 0:
                    valid.append((pdate, pval, m))
            valid.sort(key=lambda x: x[0], reverse=True)
            
            last_payback = valid[0][0] if valid else None
            cycle_days = (last_payback - sdate).days if last_payback and sdate else 0
            years = round(cycle_days / 365, 4) if cycle_days else 0
            
            if years < 0.5: zone = '0.5以内'
            elif years < 1: zone = '0.5-1年'
            elif years < 2: zone = '1年以上'
            elif years < 3: zone = '2年以上'
            else: zone = '3年以上'
            
            result.append({
                'contract_no': cno,
                'sign_date': sdate.strftime('%Y-%m-%d') if sdate else '',
                'dept': dept,
                'region': region,
                'province': province,
                'amount': amount,
                'last_payback_date': last_payback.strftime('%Y-%m-%d') if last_payback else '',
                'cycle_days': cycle_days,
                'years': years,
                'zone': zone,
            })
        return result
    
    enriched_2026 = enrich_for_year(h_data, 2026)
    enriched_2025 = enrich_for_year(h_data, 2025)
    
    # 区域聚合（2026全年）— 含平均周期 / 有回款 / 无回款 / 金额
    region_stats = {}
    for r in enriched_2026:
        reg = r.get('region', '') or '未知'
        if reg not in region_stats:
            region_stats[reg] = {'count': 0, 'total_days': 0, 'with_payment': 0, 'no_payment': 0, 'total_amount': 0.0}
        cd = r.get('cycle_days', 0) or 0
        amt = r.get('amount', 0) or 0
        region_stats[reg]['count'] += 1
        region_stats[reg]['total_days'] += cd
        region_stats[reg]['total_amount'] += amt
        if cd > 0:
            region_stats[reg]['with_payment'] += 1
        else:
            region_stats[reg]['no_payment'] += 1
    region_agg = [{'region': k, 'count': v['count'], 'avg_days': round(v['total_days']/v['count']) if v['count']>0 else 0,
                   'with_payment': v['with_payment'], 'no_payment': v['no_payment'],
                   'amount': round(v['total_amount'])} 
                  for k, v in sorted(region_stats.items(), key=lambda x: -x[1]['count'])]

    # 省份聚合（2026全年）— 供省份着色地图
    prov_stats = {}
    for r in enriched_2026:
        pv = r.get('province', '') or '未知'
        if pv not in prov_stats:
            prov_stats[pv] = {'count': 0, 'total_days': 0, 'with_payment': 0, 'no_payment': 0, 'total_amount': 0.0}
        cd = r.get('cycle_days', 0) or 0
        amt = r.get('amount', 0) or 0
        prov_stats[pv]['count'] += 1
        prov_stats[pv]['total_days'] += cd
        prov_stats[pv]['total_amount'] += amt
        if cd > 0:
            prov_stats[pv]['with_payment'] += 1
        else:
            prov_stats[pv]['no_payment'] += 1
    province_agg = [{'province': k, 'count': v['count'], 'avg_days': round(v['total_days']/v['count']) if v['count']>0 else 0,
                     'with_payment': v['with_payment'], 'no_payment': v['no_payment'],
                     'amount': round(v['total_amount'])} 
                    for k, v in sorted(prov_stats.items(), key=lambda x: -x[1]['count'])]
    
    # 按月累计计算（6-8月）
    target_months = [(2026, 6), (2026, 7), (2026, 8)]
    
    def calc_metrics(rows, y, m):
        count = len(rows)
        total_amount = sum(r['amount'] for r in rows)
        total_cycle = sum(r['cycle_days'] for r in rows)
        avg_cycle = round(total_cycle / count, 1) if count > 0 else 0
        avg_years = round(avg_cycle / 365, 2) if avg_cycle > 0 else 0
        zones = [0, 0, 0, 0, 0]
        for r in rows:
            yrs = r['years']
            if yrs <= 0: continue
            if yrs < 0.5: zones[0] += 1
            elif yrs < 1: zones[1] += 1
            elif yrs < 2: zones[2] += 1
            elif yrs < 3: zones[3] += 1
            else: zones[4] += 1
        return {'project_count': count, 'contract_amount': round(total_amount/10000,2), 'cumulative_days': total_cycle, 'avg_days': avg_cycle, 'avg_years': avg_years, 'zones': zones}
    
    def in_month(r, yr, mo):
        if not r['sign_date']: return False
        try:
            sd = datetime.strptime(r['sign_date'], '%Y-%m-%d')
            return sd.year < yr or (sd.year == yr and sd.month <= mo)
        except: return False
    
    def filter_rows(enriched, year, month):
        return [r for r in enriched if in_month(r, year, month)]
    
    months_result = []
    icid_result = {'project_count': {}, 'cumulative_days': {}, 'avg_days': {}, 'avg_years': {}, 'contract_amount': {}}
    dept_result = {'project_count': {}, 'cumulative_days': {}, 'avg_days': {}, 'avg_years': {}, 'contract_amount': {}}
    zones_result = []
    
    for y, m in target_months:
        m_key = f'{y}-{m:02d}'
        py = 2025  # previous year
        
        cur_rows = filter_rows(enriched_2026, y, m)
        prev_rows = filter_rows(enriched_2025, py, m)
        
        cur_icid = calc_metrics(cur_rows, y, m)
        prev_icid = calc_metrics(prev_rows, py, m)
        cur_dept_rows = [r for r in cur_rows if '系统集成' in r.get('dept','')]
        prev_dept_rows = [r for r in prev_rows if '系统集成' in r.get('dept','')]
        cur_dept = calc_metrics(cur_dept_rows, y, m)
        prev_dept = calc_metrics(prev_dept_rows, py, m)
        
        months_result.append({'key': m_key, 'label': f'{y}年{m}月', 'current': f'{y}年{m}月', 'last_year': f'{py}年{m}月'})
        
        for metric in ['project_count', 'cumulative_days', 'avg_days', 'avg_years', 'contract_amount']:
            cv, pv = cur_icid[metric], prev_icid[metric]
            icid_result[metric][m_key] = {'current': cv, 'previous': pv, 'diff': round(cv-pv,2) if isinstance(cv,(int,float)) and isinstance(pv,(int,float)) else None}
        
        for metric in ['project_count', 'cumulative_days', 'avg_days', 'avg_years', 'contract_amount']:
            cv, pv = cur_dept[metric], prev_dept[metric]
            dept_result[metric][m_key] = {'current': cv, 'previous': pv, 'diff': round(cv-pv,2) if isinstance(cv,(int,float)) and isinstance(pv,(int,float)) else None}
        
        for zi in range(5):
            if len(zones_result) <= zi: zones_result.append({})
            zones_result[zi][m_key] = {'current': cur_icid['zones'][zi], 'previous': prev_icid['zones'][zi], 'diff': cur_icid['zones'][zi] - prev_icid['zones'][zi]}
    
    total_row = {}
    for m in months_result:
        mk = m['key']
        tc = sum(zones_result[zi][mk]['current'] for zi in range(5))
        tp = sum(zones_result[zi][mk]['previous'] for zi in range(5))
        total_row[mk] = {'current': tc, 'previous': tp, 'diff': tc - tp}
    zones_result.append(total_row)
    
    return {
        'success': True,
        'data': {
            'source_version': hvers[0]['id'],
            'months': months_result,
            'icid': icid_result,
            'department': dept_result,
            'zones': [{'data': z} for z in zones_result],
            'enriched_rows': enriched_2026[:500],
            'enriched_total': len(enriched_2026),
            'regions': region_agg,
            'province_stats': province_agg,
        }
    }


@app.get("/api/analysis/payment-cycle/export")
def export_payment_cycle():
    """回款周期分析结果导出 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, date
    import io

    # --- 复用分析逻辑 ---
    meta = _load_ds_meta()

    htdata = meta.get('总合同表', {})
    hvers = htdata.get('versions', [])
    if not hvers:
        return JSONResponse({'success': False, 'error': '无数据源'}, status_code=400)

    hfpath = os.path.join(DATASOURCE_DIR, hvers[0]['file'])
    if not os.path.exists(hfpath):
        return JSONResponse({'success': False, 'error': '总合同表文件丢失'}, status_code=400)
    wb = openpyxl.load_workbook(hfpath, read_only=True)
    ws = wb.active
    hrows = list(ws.iter_rows(values_only=True))
    wb.close()
    h_headers = [str(h) if h else '' for h in hrows[0]]
    h_data = []
    for r in hrows[1:]:
        row = {}
        for i, h in enumerate(h_headers):
            row[h] = r[i] if i < len(r) else None
        h_data.append(row)

    def parse_date(v):
        if v is None or v == '' or v == '-': return None
        if isinstance(v, (datetime, date)): return v
        s = str(v).strip()
        if s.startswith('='): return None
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try: return datetime.strptime(s, fmt)
            except: pass
        return None

    def safe_float(v):
        if v is None or v == '' or v == '-': return 0
        try: return float(v)
        except: return 0

    def safe_str(v):
        if v is None: return ''
        return str(v).strip()

    def find_col(headers, keywords):
        for h in headers:
            hl = h.lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    h_col_no = find_col(h_headers, ['合同编号', '编号'])
    h_col_date = find_col(h_headers, ['统计日期', '签约日期', '日期'])
    h_col_dept = find_col(h_headers, ['部门', '责任部门'])
    if h_col_dept and '签定' in str(h_col_dept):
        for h in h_headers:
            if h == '部门':
                h_col_dept = h
                break
    h_col_amount = find_col(h_headers, ['合同金额', '金额', '合同额', '合同总金额'])
    h_col_region = find_col(h_headers, ['区域', '大区', '片区'])

    rtdata = meta.get('项目里程碑表', {})
    rvers = rtdata.get('versions', [])
    r_index = {}
    if rvers:
        rfpath = os.path.join(DATASOURCE_DIR, rvers[0]['file'])
        if os.path.exists(rfpath):
            wb = openpyxl.load_workbook(rfpath, read_only=True)
            ws = wb.active
            rrows = list(ws.iter_rows(values_only=True))
            wb.close()
            r_headers = [str(h) if h else '' for h in rrows[0]]
            r_col_contract = find_col(r_headers, ['合同编号'])
            for r in rrows[1:]:
                row = {}
                for i, h in enumerate(r_headers):
                    row[h] = r[i] if i < len(r) else None
                cid = str(row.get(r_col_contract, '')).strip() if r_col_contract else ''
                if cid:
                    r_index.setdefault(cid, []).append(row)

    def enrich_for_year(h_rows, target_year):
        result = []
        for row in h_rows:
            cno = safe_str(row.get(h_col_no, ''))
            sdate = parse_date(row.get(h_col_date))
            dept = safe_str(row.get(h_col_dept, ''))
            region = safe_str(row.get(h_col_region, '')) if h_col_region else ''
            amount = safe_float(row.get(h_col_amount)) if h_col_amount else 0
            if not sdate or sdate.year != target_year:
                continue
            matches = r_index.get(cno, [])
            valid = []
            for m in matches:
                pdate = parse_date(m.get('计划回款时间'))
                pval = safe_float(m.get('计划产值(元)'))
                if pdate and pval != 0:
                    valid.append((pdate, pval, m))
            valid.sort(key=lambda x: x[0], reverse=True)
            last_payback = valid[0][0] if valid else None
            cycle_days = (last_payback - sdate).days if last_payback and sdate else 0
            years = round(cycle_days / 365, 4) if cycle_days else 0
            if years < 0.5: zone = '0.5以内'
            elif years < 1: zone = '0.5-1年'
            elif years < 2: zone = '1年以上'
            elif years < 3: zone = '2年以上'
            else: zone = '3年以上'
            result.append({
                'contract_no': cno, 'sign_date': sdate.strftime('%Y-%m-%d') if sdate else '',
                'dept': dept, 'amount': amount,
                'last_payback_date': last_payback.strftime('%Y-%m-%d') if last_payback else '',
                'cycle_days': cycle_days, 'years': years, 'zone': zone,
            })
        return result

    enriched_2026 = enrich_for_year(h_data, 2026)
    enriched_2025 = enrich_for_year(h_data, 2025)

    target_months = [(2026, 6), (2026, 7), (2026, 8)]

    def calc_metrics(rows, y, m):
        count = len(rows)
        total_amount = sum(r['amount'] for r in rows)
        total_cycle = sum(r['cycle_days'] for r in rows)
        avg_cycle = round(total_cycle / count, 1) if count > 0 else 0
        avg_years = round(avg_cycle / 365, 2) if avg_cycle > 0 else 0
        zones = [0, 0, 0, 0, 0]
        for r in rows:
            yrs = r['years']
            if yrs <= 0: continue
            if yrs < 0.5: zones[0] += 1
            elif yrs < 1: zones[1] += 1
            elif yrs < 2: zones[2] += 1
            elif yrs < 3: zones[3] += 1
            else: zones[4] += 1
        return {'project_count': count, 'contract_amount': round(total_amount/10000,2),
                'cumulative_days': total_cycle, 'avg_days': avg_cycle, 'avg_years': avg_years, 'zones': zones}

    def in_month(r, yr, mo):
        if not r['sign_date']: return False
        try:
            sd = datetime.strptime(r['sign_date'], '%Y-%m-%d')
            return sd.year < yr or (sd.year == yr and sd.month <= mo)
        except: return False

    def filter_rows(enriched, year, month):
        return [r for r in enriched if in_month(r, year, month)]

    # --- 构建 Excel ---
    out = io.BytesIO()
    xwb = openpyxl.Workbook()
    xwb.remove(xwb.active)

    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(name='Microsoft YaHei', size=10, color='00e5ff', bold=True)
    normal_font = Font(name='Microsoft YaHei', size=10, color='e0e0e0')
    cell_fill = PatternFill(start_color='0d1530', end_color='0d1530', fill_type='solid')
    alt_fill = PatternFill(start_color='111b33', end_color='111b33', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='1a2540'),
        right=Side(style='thin', color='1a2540'),
        top=Side(style='thin', color='1a2540'),
        bottom=Side(style='thin', color='1a2540'),
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    def style_cell(cell, is_header=False, is_left=False, fill_idx=0):
        cell.font = header_font if is_header else normal_font
        cell.fill = header_fill if is_header else (cell_fill if fill_idx % 2 == 0 else alt_fill)
        cell.border = thin_border
        cell.alignment = left_align if is_left else center_align

    # ===== Sheet 1: ICID整体汇总 =====
    ws1 = xwb.create_sheet('ICID整体汇总')
    metrics_labels = ['合同额（万元）', '项目个数', '当年累计回款周期（天）', '平均合同回款周期（天）', '平均合同回款周期（年）']
    metrics_keys = ['contract_amount', 'project_count', 'cumulative_days', 'avg_days', 'avg_years']

    col = 1
    ws1.cell(row=1, column=col, value='指标名称')
    style_cell(ws1.cell(row=1, column=col), is_header=True, is_left=True)
    col += 1
    for y, m in target_months:
        m_key = f'{y}-{m:02d}'
        py = 2025
        for label in [f'{y}年{m}月', f'{py}年{m}月', '增长额']:
            ws1.cell(row=1, column=col, value=label)
            style_cell(ws1.cell(row=1, column=col), is_header=True)
            col += 1

    for ri, label in enumerate(metrics_labels):
        r = ri + 2
        col = 1
        ws1.cell(row=r, column=col, value=label)
        style_cell(ws1.cell(row=r, column=col), is_left=True, fill_idx=ri)
        col += 1
        for y, m in target_months:
            m_key = f'{y}-{m:02d}'
            cur_rows = filter_rows(enriched_2026, y, m)
            prev_rows = filter_rows(enriched_2025, 2025, m)
            cur_m = calc_metrics(cur_rows, y, m)
            prev_m = calc_metrics(prev_rows, 2025, m)
            cv, pv = cur_m[metrics_keys[ri]], prev_m[metrics_keys[ri]]
            diff = round(cv - pv, 2) if isinstance(cv, (int, float)) and isinstance(pv, (int, float)) else '-'
            for v in [cv, pv, diff]:
                ws1.cell(row=r, column=col, value=v)
                style_cell(ws1.cell(row=r, column=col), fill_idx=ri)
                col += 1

    # 列宽
    ws1.column_dimensions['A'].width = 28
    for i in range(2, col):
        ws1.column_dimensions[get_column_letter(i)].width = 16

    # ===== Sheet 2: 部门汇总 =====
    ws2 = xwb.create_sheet('系统集成部门汇总')
    col = 1
    ws2.cell(row=1, column=col, value='指标名称')
    style_cell(ws2.cell(row=1, column=col), is_header=True, is_left=True)
    col += 1
    for y, m in target_months:
        py = 2025
        for label in [f'{y}年{m}月', f'{py}年{m}月', '增长额']:
            ws2.cell(row=1, column=col, value=label)
            style_cell(ws2.cell(row=1, column=col), is_header=True)
            col += 1

    for ri, label in enumerate(metrics_labels):
        r = ri + 2
        col = 1
        ws2.cell(row=r, column=col, value=label)
        style_cell(ws2.cell(row=r, column=col), is_left=True, fill_idx=ri)
        col += 1
        for y, m in target_months:
            m_key = f'{y}-{m:02d}'
            cur_rows = filter_rows(enriched_2026, y, m)
            prev_rows = filter_rows(enriched_2025, 2025, m)
            cur_dept_rows = [x for x in cur_rows if '系统集成' in x.get('dept', '')]
            prev_dept_rows = [x for x in prev_rows if '系统集成' in x.get('dept', '')]
            cur_m = calc_metrics(cur_dept_rows, y, m)
            prev_m = calc_metrics(prev_dept_rows, 2025, m)
            cv, pv = cur_m[metrics_keys[ri]], prev_m[metrics_keys[ri]]
            diff = round(cv - pv, 2) if isinstance(cv, (int, float)) and isinstance(pv, (int, float)) else '-'
            for v in [cv, pv, diff]:
                ws2.cell(row=r, column=col, value=v)
                style_cell(ws2.cell(row=r, column=col), fill_idx=ri)
                col += 1

    ws2.column_dimensions['A'].width = 28
    for i in range(2, col):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    # ===== Sheet 3: 分区明细 =====
    ws3 = xwb.create_sheet('回款周期分区明细')
    zone_labels = ['0.5以内', '0.5-1年', '1年以上', '2年以上', '3年以上', '总计']
    col = 1
    ws3.cell(row=1, column=col, value='回款周期分区')
    style_cell(ws3.cell(row=1, column=col), is_header=True, is_left=True)
    col += 1
    for y, m in target_months:
        py = 2025
        for label in [f'{y}年{m}月', f'{py}年{m}月', '差值']:
            ws3.cell(row=1, column=col, value=label)
            style_cell(ws3.cell(row=1, column=col), is_header=True)
            col += 1

    # 先算各区数据
    zones_data = []
    for y, m in target_months:
        cur_rows = filter_rows(enriched_2026, y, m)
        prev_rows = filter_rows(enriched_2025, 2025, m)
        cur_m = calc_metrics(cur_rows, y, m)
        prev_m = calc_metrics(prev_rows, 2025, m)
        zones_data.append((cur_m['zones'], prev_m['zones']))
    # 总计
    totals = []
    for y, m in target_months:
        cur_rows = filter_rows(enriched_2026, y, m)
        prev_rows = filter_rows(enriched_2025, 2025, m)
        totals.append((len(cur_rows), len(prev_rows)))

    for zi, zlabel in enumerate(zone_labels):
        r = zi + 2
        col = 1
        ws3.cell(row=r, column=col, value=zlabel)
        style_cell(ws3.cell(row=r, column=col), is_left=True, fill_idx=zi)
        for mi in range(len(target_months)):
            if zi < 5:
                cv, pv = zones_data[mi][0][zi], zones_data[mi][1][zi]
            else:
                cv, pv = totals[mi]
            diff = cv - pv
            for v in [cv, pv, diff]:
                ws3.cell(row=r, column=col, value=v)
                style_cell(ws3.cell(row=r, column=col), fill_idx=zi)
                col += 1

    ws3.column_dimensions['A'].width = 20
    for i in range(2, col):
        ws3.column_dimensions[get_column_letter(i)].width = 14

    # ===== Sheet 4: 补充数据表 =====
    ws4 = xwb.create_sheet('补充数据表(2026年)')
    # 按回款周期降序
    enriched_2026.sort(key=lambda x: x['cycle_days'] or 0, reverse=True)
    headers4 = ['合同编号', '签约日期', '部门', '最后一笔回款日期', '回款周期(天)', '年', '分区']
    for ci, h in enumerate(headers4):
        ws4.cell(row=1, column=ci + 1, value=h)
        style_cell(ws4.cell(row=1, column=ci + 1), is_header=True)

    red_font = Font(name='Microsoft YaHei', size=10, color='ff5252', bold=True)
    for ri, row in enumerate(enriched_2026):
        r = ri + 2
        vals = [row['contract_no'], row['sign_date'], row['dept'],
                row['last_payback_date'], row['cycle_days'], row['years'], row['zone']]
        for ci, v in enumerate(vals):
            cell = ws4.cell(row=r, column=ci + 1, value=v if v != '' else '-')
            style_cell(cell, fill_idx=ri)
            if ci == 4 and isinstance(v, (int, float)) and v > 500:
                cell.font = red_font

    col_widths4 = [22, 14, 14, 20, 16, 10, 14]
    for ci, w in enumerate(col_widths4):
        ws4.column_dimensions[get_column_letter(ci + 1)].width = w

    # 保存到 bytes
    xwb.save(out)
    out.seek(0)

    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    filename = quote('回款周期分析报告.xlsx')
    return StreamingResponse(
        out,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"}
    )


@app.on_event("startup")
def startup():
    init_db()
    # 备品备件采购询比价智能体 — 数据层初始化
    init_procurement_db()
    seed_procurement_master()
    # 项目全生命周期管理（CC-010）— 数据层初始化（幂等建表 + 字典/预警规则/参数预置）
    import plm_models as _plm
    _plm.init_plm_db()
    _plm.seed_plm_master()


# ===================== 供应商列名 =====================
# 供应商列不再做删减过滤，比对结果/版本详情均展示全部原始列


@app.get("/")
def index():
    return FileResponse(os.path.join(FRONTEND_DIR, 'index.html'))


@app.get("/common.css")
def common_css():
    return FileResponse(os.path.join(FRONTEND_DIR, 'common.css'))


@app.get("/gross")
def gross_page():
    return FileResponse(os.path.join(FRONTEND_DIR, 'gross.html'))

@app.get("/china.json")
def china_map_data():
    return FileResponse(os.path.join(FRONTEND_DIR, 'china.json'))


# ===================== 备品备件采购询比价 =====================
from procurement_models import (
    create_task, get_task, list_tasks, confirm_selection, input_test_result, cancel_task,
    list_op_logs,
    # ---- 供应商主数据 CRUD（资源池，不绑定合同）----
    list_suppliers, get_supplier, create_supplier, update_supplier, delete_supplier,
    # ---- 台账增强查询 ----
    list_ledger_advanced,
    # ---- 合同主数据表 CRUD（5 个）----
    # create_contract / delete_contract 与 models.py 同名，必须别名导入：
    # 此前它们遮蔽了 models 的版本，导致 POST/DELETE /api/contracts 只能靠
    # `import contract_models`（该模块不存在）兜底，前端「新建合同」直接 500。
    list_contracts, get_contract,
    create_contract as proc_create_contract, update_contract,
    delete_contract as proc_delete_contract,
    # ---- 全局邮件抄送配置（list/create/delete + 给 neuops 拿 CC 列表）----
    list_mail_cc, create_mail_cc, delete_mail_cc, get_all_cc_emails,
    # ---- 前端人工修改报价 ----
    manual_update_supplier_quote,
    # ---- 备品备件主数据 CRUD ----
    list_spare_parts, get_spare_part, create_spare_part, update_spare_part,
    delete_spare_part, list_spare_part_categories,
)
from pydantic import BaseModel
from typing import List, Optional, Dict, Any  # Any：CC-010 /api/plm/* 路由体使用


class SupplierItem(BaseModel):
    """询价供应商条目（前端页面 -> 9006 API -> DB）。
    【修复 2026-08-24】显式声明 id 字段（资源池供应商有id，临时供应商无id）。
    之前未声明时，Pydantic 默认 extra='ignore' 会静默丢弃前端传入的 data-pool-id，
    导致 flow-02 中 s.id 恒为 None，全部被误标记为 _is_temp=True。"""
    model_config = {"extra": "allow"}  # 允许额外字段（如 flow-02 回写的下划线字段透传）
    id: int | None = None
    name: str
    email: str


class NewTaskBody(BaseModel):
    """新建询价任务（页面入口）：合同号 + 备件 + 数量 + 紧急等级 + 询价供应商（空则自动带池子）"""
    contract_no: str
    spare_part_model: str
    purchase_qty: float
    emergency_level: str
    inquiry_supplier_list: List[SupplierItem] = []


class AgentNewTaskBody(BaseModel):
    """智能体创建任务（对话入口）：直接传业务字段，走标准 create_task + trigger_neuops"""
    contract_no: str
    spare_part_model: str
    purchase_qty: float
    emergency_level: str
    inquiry_supplier_list: List[Dict[str, str]] = []  # 可空，空则自动带池子
    creator: str = 'agent'


class SelectBody(BaseModel):
    selected_supplier: SupplierItem
    deal_unit_price: float
    # source 标记：card_callback 表示从飞书卡片按钮触发；web(默认) 表示从前端页面手动选型
    source: str = 'web'


class TestResultBody(BaseModel):
    test_result: str
    remark: str = ''
    source: str = 'web'


class CancelBody(BaseModel):
    cancel_reason: str
    source: str = 'web'


@app.get("/procurement")
def procurement_page():
    return FileResponse(os.path.join(FRONTEND_DIR, 'procurement.html'))


@app.get("/procurement.app.js")
def procurement_app_js():
    return FileResponse(os.path.join(FRONTEND_DIR, 'procurement.app.js'))


# ---- 任务 ----
@app.get("/api/procurement/tasks")
def api_proc_task_list(status: Optional[str] = None):
    return {"success": True, "data": list_tasks(status=status)}


@app.get("/api/procurement/tasks/{task_id}")
def api_proc_task_get(task_id: str):
    t = get_task(task_id)
    if not t:
        return JSONResponse({"success": False, "error": "任务不存在"}, status_code=404)
    t['_op_logs'] = list_op_logs(task_id)
    return {"success": True, "data": t}


@app.post("/api/procurement/tasks")
def api_proc_task_create(body: NewTaskBody):
    """新建询价任务：落库 + 操作日志 + 触发 neuops 智能体发询价邮件+飞书通知
    若未传 inquiry_supplier_list，create_task 会自动从供应商资源池全量带出
    """
    try:
        t = create_task(
            contract_no=body.contract_no, spare_part_model=body.spare_part_model,
            purchase_qty=body.purchase_qty, emergency_level=body.emergency_level,
            inquiry_supplier_list=[s.dict() for s in body.inquiry_supplier_list] if body.inquiry_supplier_list else None,
            creator='pm',
        )
        # 触发 neuops emp-008：flow-proc-01(已落库) + flow-proc-02(发询价邮件+飞书通知)
        agent_r = trigger_neuops("trigger/task-created", t)
        return {"success": True, "data": t, "agent_trigger": agent_r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.post("/api/procurement/tasks/agent")
def api_proc_task_create_agent(body: AgentNewTaskBody):
    """智能体创建任务（对话入口）：直接传业务字段，走标准 create_task + trigger_neuops。
    保证 task_id 格式、reply_deadline 自动计算、操作日志、flow-proc-01/02 触发。"""
    try:
        t = create_task(
            contract_no=body.contract_no, spare_part_model=body.spare_part_model,
            purchase_qty=body.purchase_qty, emergency_level=body.emergency_level,
            inquiry_supplier_list=body.inquiry_supplier_list or None,
            creator=body.creator,
        )
        # 触发 neuops emp-008：flow-proc-01(已落库) + flow-proc-02(发询价邮件+飞书通知)
        agent_r = trigger_neuops("trigger/task-created", t)
        return {"success": True, "data": t, "agent_trigger": agent_r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.post("/api/procurement/tasks/{task_id}/select")
def api_proc_task_select(task_id: str, body: SelectBody):
    """选型确认：落库 + 触发 neuops 发采购确认邮件+飞书通知"""
    try:
        t = confirm_selection(
            task_id=task_id,
            selected_supplier=body.selected_supplier.dict(),
            deal_unit_price=body.deal_unit_price,
            operator='pm',
        )
        # 触发 neuops emp-008：flow-proc-05(发采购确认邮件+飞书通知)
        # 透传 source：card_callback 场景下 flow-proc-05 会跳过 confirm_purchase 新卡片通知，
        # 避免与 card-callback 返回的就地替换置灰卡片造成双卡片
        agent_r = trigger_neuops("trigger/task-selected", {
            "task": t, "selected_supplier": body.selected_supplier.dict(),
            "deal_unit_price": body.deal_unit_price,
            "source": body.source or "web",
        })
        return {"success": True, "data": t, "agent_trigger": agent_r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.post("/api/procurement/tasks/{task_id}/test")
def api_proc_task_test(task_id: str, body: TestResultBody):
    """测试结果录入：落库 + 触发 neuops 闭环/告警+飞书通知"""
    try:
        t = input_test_result(
            task_id=task_id, test_result=body.test_result,
            remark=body.remark, operator='pm',
        )
        agent_r = trigger_neuops("trigger/test-result", {
            "task": t, "test_result": body.test_result, "remark": body.remark,
            "source": body.source or "web",
        })
        return {"success": True, "data": t, "agent_trigger": agent_r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.post("/api/procurement/tasks/{task_id}/cancel")
def api_proc_task_cancel(task_id: str, body: CancelBody):
    """任务取消：落库 + 触发 neuops 飞书通知取消"""
    try:
        t = cancel_task(task_id=task_id, cancel_reason=body.cancel_reason, operator='pm')
        agent_r = trigger_neuops("trigger/task-canceled", {
            "task": t, "cancel_reason": body.cancel_reason,
            "source": body.source or "web",
        })
        return {"success": True, "data": t, "agent_trigger": agent_r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.get("/api/procurement/tasks/{task_id}/logs")
def api_proc_task_logs(task_id: str):
    return {"success": True, "data": list_op_logs(task_id)}


class ManualQuoteBody(BaseModel):
    reply_index: int
    unit_price: Optional[float] = None
    total_price: Optional[float] = None
    lead_time: Optional[str] = None
    brand: Optional[str] = None
    model: Optional[str] = None
    note: Optional[str] = None


@app.patch("/api/procurement/tasks/{task_id}/quote/manual")
def api_proc_task_quote_manual(task_id: str, body: ManualQuoteBody):
    """前端铅笔按钮：人工录入/修改某供应商报价。保存后 is_manual=True，后续 IMAP 复解析不会覆盖。"""
    try:
        t = manual_update_supplier_quote(
            task_id=task_id, reply_index=body.reply_index,
            payload={
                "unit_price": body.unit_price, "total_price": body.total_price,
                "lead_time": body.lead_time, "brand": body.brand,
                "model": body.model, "note": body.note,
            },
            operator="frontend:user",
        )
        return {"success": True, "data": t}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


# ---- 台账 ----
@app.get("/api/procurement/ledger")
def api_proc_ledger_list(contract_no: Optional[str] = None,
                         supplier_name: Optional[str] = None,
                         from_date: Optional[str] = None,
                         to_date: Optional[str] = None,
                         limit: int = 500):
    """采购业务台账（增强查询：合同 / 供应商 / 日期范围 / 条数上限）"""
    rows = list_ledger_advanced(contract_no=contract_no,
                                supplier_name=supplier_name,
                                from_date=from_date, to_date=to_date, limit=limit)
    return {"success": True, "data": rows}


# ============================================================
# 【新增 A】供应商主数据 CRUD（5 个 REST 路由）
# ============================================================

class SupplierBody(BaseModel):
    name: str
    email: str
    capability: Optional[str] = ''


class SupplierUpdateBody(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    capability: Optional[str] = None


@app.get("/api/procurement/suppliers")
def api_proc_suppliers_list(keyword: Optional[str] = None, limit: int = 500):
    """供应商主数据列表（支持 名称/邮箱/供货能力 关键词模糊搜索）"""
    return {"success": True, "data": list_suppliers(keyword=keyword, limit=limit)}


@app.get("/api/procurement/suppliers/{supplier_id}")
def api_proc_suppliers_get(supplier_id: int):
    s = get_supplier(supplier_id)
    if not s:
        return JSONResponse({"success": False, "error": "供应商不存在"}, status_code=404)
    return {"success": True, "data": s}


@app.post("/api/procurement/suppliers")
def api_proc_suppliers_create(body: SupplierBody):
    try:
        s = create_supplier(name=body.name, email=body.email,
                            capability=body.capability or '')
        return {"success": True, "data": s}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.put("/api/procurement/suppliers/{supplier_id}")
def api_proc_suppliers_update(supplier_id: int, body: SupplierUpdateBody):
    try:
        s = update_supplier(supplier_id=supplier_id, name=body.name,
                            email=body.email, capability=body.capability)
        if s is None:
            return JSONResponse({"success": False, "error": "供应商不存在"}, status_code=404)
        return {"success": True, "data": s}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.delete("/api/procurement/suppliers/{supplier_id}")
def api_proc_suppliers_delete(supplier_id: int):
    try:
        r = delete_supplier(supplier_id)
        return {"success": True, **r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


# ============================================================
# 合同主数据表：5 个 CRUD（合同名 / 合同编号 / 项目经理 / 项目经理邮箱）
# ============================================================

class ProcContractBody(BaseModel):
    contract_no: str
    contract_name: Optional[str] = ''
    pm_name: Optional[str] = ''
    pm_email: Optional[str] = ''
    receiver_name: Optional[str] = ''
    receiver_phone: Optional[str] = ''
    receiver_address: Optional[str] = ''


class ProcContractUpdateBody(BaseModel):
    contract_no: Optional[str] = None
    contract_name: Optional[str] = None
    pm_name: Optional[str] = None
    pm_email: Optional[str] = None
    receiver_name: Optional[str] = None
    receiver_phone: Optional[str] = None
    receiver_address: Optional[str] = None


@app.get("/api/procurement/contracts")
def api_proc_contracts_list(keyword: Optional[str] = None, limit: int = 500):
    """合同主数据列表：按 合同编号 / 合同名 / 项目经理名 / 邮箱 搜索"""
    return {"success": True, "data": list_contracts(keyword=keyword, limit=limit)}


@app.get("/api/procurement/contracts/{contract_id}")
def api_proc_contracts_get(contract_id: int):
    s = get_contract(contract_id=contract_id)
    if not s:
        return JSONResponse({"success": False, "error": "合同不存在"}, status_code=404)
    return {"success": True, "data": s}


@app.post("/api/procurement/contracts")
def api_proc_contracts_create(body: ProcContractBody):
    try:
        c = proc_create_contract(
            contract_no=body.contract_no,
            contract_name=body.contract_name or '',
            pm_name=body.pm_name or '',
            pm_email=body.pm_email or '',
            receiver_name=body.receiver_name or '',
            receiver_phone=body.receiver_phone or '',
            receiver_address=body.receiver_address or '',
        )
        return {"success": True, "data": c}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.put("/api/procurement/contracts/{contract_id}")
def api_proc_contracts_update(contract_id: int, body: ProcContractUpdateBody):
    try:
        c = update_contract(contract_id=contract_id, contract_no=body.contract_no,
                            contract_name=body.contract_name, pm_name=body.pm_name,
                            pm_email=body.pm_email,
                            receiver_name=body.receiver_name, receiver_phone=body.receiver_phone,
                            receiver_address=body.receiver_address)
        if c is None:
            return JSONResponse({"success": False, "error": "合同不存在"}, status_code=404)
        return {"success": True, "data": c}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.delete("/api/procurement/contracts/{contract_id}")
def api_proc_contracts_delete(contract_id: int):
    try:
        r = proc_delete_contract(contract_id)
        return {"success": True, **r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


# ============================================================
# 【新增】全局邮件抄送配置：列表 / 新增 / 删除 + 给 neuops 用的 "只返回 [(name,email)]" 极简接口
# ============================================================

class ProcMailCCBody(BaseModel):
    name: str
    email: str


@app.get("/api/procurement/mail-cc")
def api_proc_mailcc_list(keyword: Optional[str] = None):
    return {"success": True, "data": list_mail_cc(keyword=keyword)}


@app.get("/api/procurement/mail-cc/emails")
def api_proc_mailcc_emails_plain():
    """给 neuops 调用的极简接口：只返回 CC 列表，不包裹 success/data。"""
    return {"cc": get_all_cc_emails()}


@app.post("/api/procurement/mail-cc")
def api_proc_mailcc_create(body: ProcMailCCBody):
    try:
        r = create_mail_cc(name=body.name, email=body.email)
        return {"success": True, "data": r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.delete("/api/procurement/mail-cc/{cc_id}")
def api_proc_mailcc_delete(cc_id: int):
    try:
        r = delete_mail_cc(cc_id)
        return {"success": True, **r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


# ===================== 备品备件 =====================

@app.get("/api/procurement/spare-parts")
def api_proc_spare_parts(keyword: str = Query(None), category: str = Query(None)):
    rows = list_spare_parts(keyword=keyword, category=category)
    return {"success": True, "data": rows, "total": len(rows)}


@app.get("/api/procurement/spare-parts/categories")
def api_proc_spare_part_categories():
    cats = list_spare_part_categories()
    return {"success": True, "data": cats}


@app.get("/api/procurement/spare-parts/{part_id}")
def api_proc_spare_part_get(part_id: int):
    r = get_spare_part(part_id)
    if not r:
        return JSONResponse({"success": False, "error": "备件不存在"}, status_code=404)
    return {"success": True, "data": r}


class SparePartBody(BaseModel):
    part_code: str
    part_name: str
    spec_model: str = ''
    brand: str = ''
    unit: str = '个'
    category: str = '通用'
    condition: str = ''
    remark: str = ''


@app.post("/api/procurement/spare-parts")
def api_proc_spare_part_create(body: SparePartBody):
    try:
        r = create_spare_part(
            part_code=body.part_code, part_name=body.part_name,
            spec_model=body.spec_model, brand=body.brand,
            unit=body.unit, category=body.category,
            condition=body.condition, remark=body.remark)
        return {"success": True, "data": r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.put("/api/procurement/spare-parts/{part_id}")
def api_proc_spare_part_update(part_id: int, body: SparePartBody):
    try:
        r = update_spare_part(part_id,
                              part_code=body.part_code, part_name=body.part_name,
                              spec_model=body.spec_model, brand=body.brand,
                              unit=body.unit, category=body.category,
                              condition=body.condition, remark=body.remark)
        return {"success": True, "data": r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.delete("/api/procurement/spare-parts/{part_id}")
def api_proc_spare_part_delete(part_id: int):
    try:
        r = delete_spare_part(part_id)
        return {"success": True, **r}
    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


# ===================== 合同管理 =====================

@app.get("/api/contracts")
def list_contracts_legacy(keyword: str = Query(None), status: str = Query(None)):
    """全部合同列表（首页）——【旧合同管理】Python 函数加 _legacy 避免与 procurement_models 导入的同名函数冲突"""
    conn = get_db()
    where = ["1=1"]
    params = []
    if keyword:
        where.append("(contract_name LIKE ? OR contract_no LIKE ?)")
        params.extend([f"%{keyword}%", f"%{keyword}%"])
    if status and status != '全部':
        where.append("status = ?")
        params.append(status)

    contracts = [dict(r) for r in conn.execute(
        f"""SELECT c.*, COALESCE(v.progress, 0) as progress, COALESCE(v.supplier_name, '') as latest_supplier,
                   (SELECT COUNT(DISTINCT supplier_name) FROM versions WHERE contract_id = c.id AND supplier_name != '') as supplier_count
            FROM contracts c
            LEFT JOIN (
                SELECT contract_id, MAX(id) as max_id FROM versions GROUP BY contract_id
            ) latest ON c.id = latest.contract_id
            LEFT JOIN versions v ON v.id = latest.max_id
            WHERE {' AND '.join(where)}
            ORDER BY c.id DESC""", params
    ).fetchall()]

    # 统计
    stats = {
        'total': conn.execute("SELECT COUNT(*) FROM contracts").fetchone()[0],
        'closed': conn.execute("SELECT COUNT(*) FROM contracts WHERE status = '已闭环(100%)'").fetchone()[0],
        'active': conn.execute("SELECT COUNT(*) FROM contracts WHERE status != '已闭环(100%)' AND status != '未上传基准'").fetchone()[0],
        'total_amount': conn.execute("SELECT COALESCE(SUM(total_amount), 0) FROM contracts").fetchone()[0],
    }
    conn.close()
    return JSONResponse({'contracts': contracts, 'stats': stats})


@app.post("/api/contracts")
async def create_new_contract(name: str = Query(...), no: str = Query(''), sign_date: str = Query('')):
    """CC-001 FR-1 新建合同（合同比对域）。create_contract 现为 models.py 的实现。"""
    if not (name or '').strip():
        return JSONResponse({'success': False, 'error': '合同名称不能为空'}, status_code=400)
    cid = create_contract(name.strip(), no or '', sign_date or '')
    return JSONResponse({'success': True, 'contract_id': cid})


@app.put("/api/contracts/{contract_id}")
async def update_contract_legacy(contract_id: int):
    """更新合同元信息（通过form data）- 函数加 _legacy 避免与采购模块同名冲突"""
    # Simple update via query params for now
    return JSONResponse({'success': True})


@app.delete("/api/contracts/{contract_id}")
def remove_contract(contract_id: int):
    """CC-001 FR-3 删除合同并级联清理（合同比对域）。"""
    delete_contract(contract_id)
    return JSONResponse({'success': True})


# ===================== 合同基准管理 =====================

@app.post("/api/contract/{contract_id}/upload")
async def upload_contract(contract_id: int, file: UploadFile = File(...)):
    filepath = os.path.join(UPLOAD_DIR, f'contract_{contract_id}_基准.xlsx')
    with open(filepath, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    try:
        result = import_contract_excel(contract_id, filepath)
        # 更新合同总金额
        conn = get_db()
        total = conn.execute(
            "SELECT COALESCE(SUM(contract_amount), 0) FROM contract_items WHERE contract_id = ?",
            (contract_id,)
        ).fetchone()[0]
        conn.execute("UPDATE contracts SET total_amount = ?, status = '比对进行中' WHERE id = ?",
                     (total, contract_id))
        conn.commit(); conn.close()
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)


@app.get("/api/contract/{contract_id}/items")
def list_contract_items(contract_id: int):
    conn = get_db()
    items = [dict(r) for r in conn.execute(
        "SELECT * FROM contract_items WHERE contract_id = ? ORDER BY id", (contract_id,)
    ).fetchall()]
    all_headers = []
    for item in items:
        try:
            raw = json.loads(item.get('raw_columns', '{}'))
            for k in raw:
                if k not in all_headers: all_headers.append(k)
        except: pass
    conn.close()
    return JSONResponse({'items': items, 'total': len(items), 'headers': all_headers})


# ===================== 供应商版本管理 =====================

@app.post("/api/contract/{contract_id}/supplier/upload")
async def upload_supplier(contract_id: int, file: UploadFile = File(...),
                          supplier_name: str = Query('')):
    if not supplier_name.strip():
        return JSONResponse({'success': False, 'error': '请填写供应商名称'}, status_code=400)
    filepath = os.path.join(UPLOAD_DIR, f'supplier_{contract_id}_{supplier_name}_{file.filename}')
    with open(filepath, 'wb') as f:
        shutil.copyfileobj(file.file, f)
    try:
        result = import_supplier_excel(contract_id, filepath, supplier_name.strip())
        update_contract_status(contract_id)
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)


@app.get("/api/contract/{contract_id}/supplier/versions")
def list_versions(contract_id: int, supplier_name: str = Query(None)):
    """版本列表，可按供应商筛选"""
    conn = get_db()
    where = "contract_id = ?"
    params = [contract_id]
    if supplier_name:
        where += " AND supplier_name = ?"
        params.append(supplier_name)
    versions = [dict(r) for r in conn.execute(
        f"SELECT * FROM versions WHERE {where} ORDER BY id DESC", params
    ).fetchall()]

    # 该合同下有哪些供应商
    suppliers = [dict(r) for r in conn.execute(
        "SELECT supplier_name, COUNT(*) as version_count, MAX(id) as latest_id FROM versions WHERE contract_id = ? AND supplier_name != '' GROUP BY supplier_name ORDER BY latest_id DESC",
        (contract_id,)
    ).fetchall()]
    conn.close()
    return JSONResponse({'versions': versions, 'suppliers': suppliers})


@app.get("/api/contract/{contract_id}/supplier/items")
def list_supplier_items(contract_id: int, version_id: int = Query(...)):
    conn = get_db()
    items = [dict(r) for r in conn.execute(
        "SELECT * FROM supplier_items WHERE contract_id = ? AND version_id = ? ORDER BY id",
        (contract_id, version_id)
    ).fetchall()]
    # 提取供应商全部原始列名（保持顺序，不删减）
    all_headers = []
    for item in items:
        try:
            raw = json.loads(item.get('raw_columns', '{}'))
            for k in raw:
                if k not in all_headers:
                    all_headers.append(k)
        except Exception:
            pass
    conn.close()
    return JSONResponse({'items': items, 'total': len(items), 'headers': all_headers})


@app.delete("/api/contract/{contract_id}/supplier/versions/{version_id}")
def delete_version(contract_id: int, version_id: int):
    """删除供应商版本，级联删除 comparison_results 和 supplier_items"""
    conn = get_db()
    c = conn.cursor()

    # 获取版本信息用于后续处理
    v = c.execute(
        "SELECT supplier_name FROM versions WHERE id = ? AND contract_id = ?",
        (version_id, contract_id)
    ).fetchone()
    if not v:
        conn.close()
        return JSONResponse({'success': False, 'error': '版本不存在'}, status_code=404)

    supplier_name = v['supplier_name']

    # 级联删除
    c.execute("DELETE FROM comparison_results WHERE version_id = ?", (version_id,))
    c.execute("DELETE FROM supplier_items WHERE version_id = ?", (version_id,))
    c.execute("DELETE FROM versions WHERE id = ?", (version_id,))

    # 如果该供应商还有其他版本，让最新的一个变活跃
    if supplier_name:
        latest = c.execute("""
            SELECT id FROM versions
            WHERE contract_id = ? AND supplier_name = ?
            ORDER BY id DESC LIMIT 1
        """, (contract_id, supplier_name)).fetchone()
        if latest:
            c.execute("UPDATE versions SET is_active = 1 WHERE id = ?", (latest[0],))

    conn.commit()
    conn.close()

    update_contract_status(contract_id)
    return JSONResponse({'success': True})


# ===================== 比对引擎 =====================

@app.post("/api/contract/{contract_id}/compare/run")
def run_compare(contract_id: int, version_id: int = Query(...)):
    try:
        result = run_comparison(contract_id, version_id)
        update_contract_status(contract_id)
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)


@app.get("/api/contract/{contract_id}/compare/results")
def get_results(contract_id: int, version_id: int = Query(None),
                status: str = Query(None), keyword: str = Query(None)):
    conn = get_db()
    where = ["r.contract_id = ?"]
    params = [contract_id]

    if version_id:
        where.append("r.version_id = ?"); params.append(version_id)
    else:
        latest = conn.execute(
            "SELECT MAX(id) FROM versions WHERE contract_id = ?", (contract_id,)
        ).fetchone()[0]
        if latest:
            where.append("r.version_id = ?"); params.append(latest)

    if status and status != '全部':
        where.append("r.match_status = ?"); params.append(status)

    query = f"""
        SELECT r.*, ct.device_name as ct_name, ct.device_model as ct_model,
            ct.specs_full as ct_specs, ct.contract_qty, ct.contract_unit,
            ct.raw_columns as ct_raw,
            sp.device_name as sp_name, sp.device_model as sp_model,
            sp.specs_full as sp_specs, sp.quote_qty, sp.quote_unit,
            sp.raw_columns as sp_raw
        FROM comparison_results r
        LEFT JOIN contract_items ct ON r.contract_item_id = ct.id
        LEFT JOIN supplier_items sp ON r.supplier_item_id = sp.id
        WHERE {' AND '.join(where)} ORDER BY r.match_status, ct.device_name
    """
    results = [dict(r) for r in conn.execute(query, params).fetchall()]

    # 收集合同和供应商的原始列名（取第一个有数据的）
    ct_headers = []
    sp_headers_raw = []
    for r in results:
        if not ct_headers and r.get('ct_raw'):
            try: ct_headers = list(json.loads(r['ct_raw']).keys())
            except: pass
        if not sp_headers_raw and r.get('sp_raw'):
            try: sp_headers_raw = list(json.loads(r['sp_raw']).keys())
            except: pass
        if ct_headers and sp_headers_raw: break

    # 供应商列保留全部原始列，不删减
    sp_headers = sp_headers_raw
    conn.close()

    if keyword:
        kw = keyword.lower()
        results = [r for r in results if kw in str(r.get('ct_name','')).lower()
                   or kw in str(r.get('ct_model','')).lower()
                   or kw in str(r.get('sp_name','')).lower()
                   or kw in str(r.get('anomaly_detail','')).lower()]

    for r in results:
        try: r['anomaly_types_list'] = json.loads(r.get('anomaly_types', '[]'))
        except: r['anomaly_types_list'] = []

    return JSONResponse({'results': results, 'total': len(results), 'ct_headers': ct_headers, 'sp_headers': sp_headers})


@app.get("/api/contract/{contract_id}/column-mapping")
def get_column_mapping(contract_id: int, version_id: int = Query(None)):
    """返回主合同列 ↔ 供应商列 的对齐关系"""
    conn = get_db()
    if not version_id:
        latest = conn.execute("SELECT MAX(id) FROM versions WHERE contract_id=?", (contract_id,)).fetchone()[0]
        version_id = latest
    row = conn.execute("SELECT column_mapping FROM versions WHERE id=?", (version_id,)).fetchone()
    conn.close()
    if row and row['column_mapping']:
        try:
            cm = json.loads(row['column_mapping'])
            cm['version_id'] = version_id
            return JSONResponse(cm)
        except Exception:
            pass
    return JSONResponse({'version_id': version_id})


@app.post("/api/contract/{contract_id}/column-mapping")
async def save_column_mapping(contract_id: int, request: Request):
    """保存手动调整的列对齐，重提供应商数据并重新比对"""
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({'success': False, 'error': '请求体不是合法 JSON'}, status_code=400)
    version_id = body.get('version_id')
    mapping = body.get('mapping', {})
    if not version_id:
        return JSONResponse({'success': False, 'error': '缺少 version_id'}, status_code=400)

    conn = get_db()
    c = conn.cursor()
    row = c.execute("SELECT column_mapping FROM versions WHERE id=?", (version_id,)).fetchone()
    if not row:
        conn.close()
        return JSONResponse({'success': False, 'error': '版本不存在'}, status_code=404)
    try:
        cm = json.loads(row['column_mapping'] or '{}')
    except Exception:
        cm = {}
    cm['mapping'] = mapping
    c.execute("UPDATE versions SET column_mapping=? WHERE id=?",
              (json.dumps(cm, ensure_ascii=False), version_id))
    conn.commit()
    conn.close()

    try:
        result = reapply_column_mapping(contract_id, version_id, mapping, cm.get('contract_semantics', {}))
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)
    update_contract_status(contract_id)
    result['column_mapping'] = cm
    return JSONResponse(result)


@app.post("/api/compare/{result_id}/confirm")
def confirm_result(result_id: int, confirmed: int = Query(1)):
    """手动确认：判断符合/匹配异常 → 确认后变为匹配成功；取消确认恢复原状态"""
    conn = get_db()
    c = conn.cursor()
    # 获取当前状态
    cur = dict(c.execute("SELECT * FROM comparison_results WHERE id = ?", (result_id,)).fetchone())
    original_status = cur['match_status']
    
    if confirmed:
        # 确认：判断符合/匹配异常 → 匹配成功
        if original_status in ('判断符合', '匹配异常'):
            c.execute("UPDATE comparison_results SET confirmed = 1, match_status = '匹配成功' WHERE id = ?",
                      (result_id,))
    else:
        # 取消确认：匹配成功 → 恢复原状态
        # 通过 match_note 推断原状态：有推理过程说明曾是判断符合，否则是匹配异常
        if cur.get('match_note', ''):
            c.execute("UPDATE comparison_results SET confirmed = 0, match_status = '判断符合' WHERE id = ?",
                      (result_id,))
        else:
            c.execute("UPDATE comparison_results SET confirmed = 0, match_status = '匹配异常' WHERE id = ?",
                      (result_id,))
    
    # 同步更新版本统计
    r = dict(c.execute("SELECT * FROM comparison_results WHERE id = ?", (result_id,)).fetchone())
    conn.commit()
    
    vid = r['version_id']
    stats = c.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN match_status='匹配成功' THEN 1 ELSE 0 END) as matched,
            SUM(CASE WHEN match_status='判断符合' THEN 1 ELSE 0 END) as judged,
            SUM(CASE WHEN match_status='匹配异常' THEN 1 ELSE 0 END) as anomaly,
            SUM(CASE WHEN match_status='待采购' THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN match_status='供应商增项' THEN 1 ELSE 0 END) as extra
        FROM comparison_results WHERE version_id = ?
    """, (vid,)).fetchone()
    contract_total = c.execute(
        "SELECT COUNT(*) FROM contract_items WHERE contract_id = ?", (r['contract_id'],)
    ).fetchone()[0]
    # 进度 = 匹配成功（含已确认的判断符合/异常）/ 合同总条目
    progress = round((stats['matched'] / max(contract_total, 1) * 100), 2)
    c.execute("UPDATE versions SET matched_count=?, judged_count=?, anomaly_count=?, pending_count=?, extra_count=?, progress=? WHERE id=?",
              (stats['matched'] or 0, stats['judged'] or 0, stats['anomaly'] or 0, stats['pending'] or 0, stats['extra'] or 0, progress, vid))
    conn.commit()
    conn.close()
    update_contract_status(r['contract_id'])
    return JSONResponse({'success': True, 'confirmed': bool(confirmed)})


# ===================== 统计（多合同版） =====================

@app.get("/api/contract/{contract_id}/stats")
def get_contract_stats(contract_id: int):
    conn = get_db()
    ct_count = conn.execute(
        "SELECT COUNT(*) FROM contract_items WHERE contract_id = ?", (contract_id,)
    ).fetchone()[0]
    ct_amount = conn.execute(
        "SELECT COALESCE(SUM(contract_amount), 0) FROM contract_items WHERE contract_id = ?",
        (contract_id,)
    ).fetchone()[0]

    stats = {'contract_total': ct_count, 'contract_amount': ct_amount}
    latest = conn.execute(
        "SELECT MAX(id) FROM versions WHERE contract_id = ?", (contract_id,)
    ).fetchone()[0]
    if latest:
        v = dict(conn.execute("SELECT * FROM versions WHERE id = ?", (latest,)).fetchone())
        stats.update({
            'version_id': v['id'], 'matched_count': v['matched_count'],
            'judged_count': v.get('judged_count', 0),
            'anomaly_count': v['anomaly_count'], 'pending_count': v['pending_count'],
            'extra_count': v['extra_count'], 'progress': v['progress'],
        })

    versions = [dict(r) for r in conn.execute(
        "SELECT id, progress FROM versions WHERE contract_id = ? ORDER BY id", (contract_id,)
    ).fetchall()]
    conn.close()
    return JSONResponse({'stats': stats, 'versions': versions})


# ===================== 全局统计（首页） =====================

@app.get("/api/stats")
def get_global_stats():
    """全局聚合统计（已废弃，用 /api/contracts 代替）"""
    conn = get_db()
    stats = {
        'total': conn.execute("SELECT COUNT(*) FROM contracts").fetchone()[0],
        'total_amount': conn.execute("SELECT COALESCE(SUM(total_amount), 0) FROM contracts").fetchone()[0],
    }
    conn.close()
    return JSONResponse({'stats': stats, 'versions': []})


# ===================== 报告导出 =====================

@app.get("/api/contract/{contract_id}/export/report")
def download_report(contract_id: int, version_id: int = Query(...)):
    try:
        filepath = export_report(version_id)
        return FileResponse(filepath,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=os.path.basename(filepath))
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)


# ===================== 资金占用分析 API =====================

FUND_DATA_DIR = os.path.join(BASE_DIR, '..', 'fund_data')
os.makedirs(FUND_DATA_DIR, exist_ok=True)


@app.get("/api/fund/status")
def fund_status():
    """查询付款/收款明细的上传状态：优先数据源管理，回退到 fund_data 目录"""
    from datetime import datetime
    result = {}
    meta = _load_ds_meta()
    name_map = {'payment': ('付款明细表', 'payment_details.xlsx'),
                'collection': ('收款明细表', 'collection_details.xlsx')}
    for key, (tname, fname) in name_map.items():
        vers = meta.get(tname, {}).get('versions', [])
        if vers:
            result[key] = f'{tname} v{vers[0]["id"]}'
            continue
        fpath = os.path.join(FUND_DATA_DIR, fname)
        if os.path.exists(fpath):
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%Y-%m-%d %H:%M')
                size = os.path.getsize(fpath) // 1024
                result[key] = f'{tname}（{size}KB · {mtime}）'
            except Exception:
                result[key] = f'{tname}（已上传）'
    return result


@app.post("/api/fund/upload")
async def fund_upload(file: UploadFile = File(...), type: str = Form(...)):
    """上传付款明细/收款明细/计算规则"""
    if type not in ('payment', 'collection', 'rule'):
        return {'success': False, 'error': '无效的文件类型'}
    name_map = {'payment': 'payment_details.xlsx',
                'collection': 'collection_details.xlsx',
                'rule': 'fund_rule.xlsx'}
    fname = name_map[type]
    fpath = os.path.join(FUND_DATA_DIR, fname)
    content = await file.read()
    with open(fpath, 'wb') as f:
        f.write(content)
    # 客户名/客户简称/项目名等敏感列一律删除，不进入系统
    sanitize_excel_file(fpath)
    return {'success': True, 'filename': fname}


# FIFO 垫资片段全局缓存
fund_segments_cache = {}
fund_flows_cache = {}


def _save_snapshot(job_key, result_dict):
    """保存分析结果快照到 SQLite（持久化，替代内存缓存）"""
    import json as _json
    from models import init_db
    init_db()
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO analysis_snapshots (job_key, result_json)
        VALUES (?,?)
        ON CONFLICT(job_key) DO UPDATE SET result_json=excluded.result_json, updated_at=datetime('now','localtime')
    """, (job_key, _json.dumps(result_dict, ensure_ascii=False, default=str)))
    conn.commit()
    conn.close()


def _load_snapshot(job_key):
    """读取分析结果快照"""
    import json as _json
    conn = get_db()
    c = conn.cursor()
    row = c.execute("SELECT result_json FROM analysis_snapshots WHERE job_key=?", (job_key,)).fetchone()
    conn.close()
    if row:
        return _json.loads(row['result_json'])
    return None

# ═══════════════════════════════════════════
# 资金多维度分析与预警（CC-006 FR-8 ~ FR-12）
# ═══════════════════════════════════════════

def _encode_customer_key(name):
    """对客户名称做确定性脱敏编码：已是编码（2-20 位字母数字）原样返回，否则 md5 前 8 位大写。

    只持久化编码，不持久化真实名称。
    """
    if not name:
        return ''
    s = str(name).strip()
    if re.fullmatch(r'[A-Za-z0-9]{2,20}', s):
        return s
    import hashlib
    return hashlib.md5(s.encode('utf-8')).hexdigest()[:8].upper()


def _seed_risk_config():
    """写入风险预警阈值默认值（幂等，仅缺省时插入）"""
    defaults = [
        ('days_green', 30, '占用≤N天为健康'),
        ('days_yellow', 90, '占用N天以内为关注'),
        ('days_orange', 180, '占用>N天进入预警/高危判定'),
        ('recv_rate', 0.5, '回款率阈值（<N 且占用90-180天→预警）'),
        ('intensity', 0.5, '占用强度阈值（>N 且占用>180天→高危）'),
        ('amount_high', 1000000, '回款率=0 且占用金额≥N→高危'),
        ('trend_months', 2, '占用金额环比连续上升N个月→趋势预警'),
    ]
    try:
        conn = get_db()
        c = conn.cursor()
        for k, v, desc in defaults:
            c.execute("INSERT OR IGNORE INTO risk_config (key, value, description) VALUES (?,?,?)", (k, v, desc))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[risk_config] seed 失败: {e}", flush=True)


def _get_risk_config():
    """读取风险阈值配置（dict），缺表/缺行时返回默认值"""
    defaults = {'days_green': 30, 'days_yellow': 90, 'days_orange': 180,
                'recv_rate': 0.5, 'intensity': 0.5, 'amount_high': 1000000,
                'trend_months': 2}
    try:
        conn = get_db()
        c = conn.cursor()
        rows = [dict(r) for r in c.execute("SELECT key, value FROM risk_config").fetchall()]
        conn.close()
        for r in rows:
            defaults[r['key']] = float(r['value'])
    except Exception:
        pass
    return defaults


def _calc_risk_level(occupy_days, recv_rate, occupy_intensity, occupy_amount, cfg):
    """风险分级纯函数：返回 (level, suggestion)。

    level ∈ healthy/yellow/orange/red；阈值来自 cfg（risk_config 表）。
    """
    days_green = cfg.get('days_green', 30)
    days_yellow = cfg.get('days_yellow', 90)
    days_orange = cfg.get('days_orange', 180)
    rr = cfg.get('recv_rate', 0.5)
    inten = cfg.get('intensity', 0.5)
    amount_high = cfg.get('amount_high', 1000000)

    # 回款率为 0 且占用金额超阈值 → 高危（强制）
    if recv_rate == 0 and occupy_amount >= amount_high:
        return 'red', '回款为0且占用金额超阈值，立即催收并上报'

    if occupy_days <= days_green:
        return 'healthy', '正常回款周期内'
    if occupy_days <= days_yellow:
        return 'yellow', '占用超过回款周期，提醒跟进回款'
    if occupy_days <= days_orange:
        if recv_rate >= rr:
            return 'yellow', '占用偏高但回款率达标，持续关注'
        return 'orange', '占用偏高且回款率不足，通知区域/部门负责人'
    # > days_orange
    if occupy_intensity <= inten:
        return 'orange', '长期占用但占用强度可控，安排对账催收'
    return 'red', '长期占用且占用强度过高，强制干预（催收/对账/上报）'


def _calc_trend_warning(monthly_occupy, trend_months=2):
    """趋势预警纯函数：按月占用序列判断是否环比连续上升 N 个月。

    monthly_occupy: [{'month': 'YYYY-MM', 'occupy': 100}, ...] 升序。
    返回 (bool, desc)。
    """
    if len(monthly_occupy) < trend_months + 1:
        return False, ''
    vals = [float(m.get('occupy', 0) or 0) for m in monthly_occupy]
    n = trend_months
    # 检查最后 n+1 个月是否连续上升
    tail = vals[-(n + 1):]
    ok = all(tail[i + 1] > tail[i] for i in range(n))
    if ok:
        desc = '、'.join(f"{m['month']}:{float(m.get('occupy',0) or 0):,.0f}" for m in monthly_occupy[-n - 1:])
        return True, f'占用金额连续{n}个月上升（{desc}）'
    return False, ''


def _load_contract_dims(contract_nos):
    """按合同编号 join 总合同表 + 项目里程碑表，返回 {contract_no: dims_dict}。

    dims 字段：region/province/dept/biz_line/industry/customer_key/
               contract_status/sign_year/project_status/payment_term/plan_recv_date
    join 不到的合同返回空 dims（归入「未知」桶，不影响主分析）。
    """
    import openpyxl
    from datetime import datetime, date

    def parse_date(v):
        if v is None or v == '' or v == '-': return None
        if isinstance(v, (datetime, date)): return v
        s = str(v).strip()
        if s.startswith('='): return None
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try: return datetime.strptime(s, fmt)
            except: pass
        return None

    def find_col(headers, keywords):
        for h in headers:
            hl = str(h).lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    dims_map = {cno: {} for cno in contract_nos}

    # ── 总合同表：区域/部门/业务线/行业/客户标识/合同状态/合同额/签约时间 ──
    h_fpath = _ds_latest_path('总合同表')
    if h_fpath and os.path.exists(h_fpath):
        try:
            wb = openpyxl.load_workbook(h_fpath, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            headers = [str(h) if h else '' for h in rows[0]]
            h_idx = {h: i for i, h in enumerate(headers)}
            col_no = find_col(headers, ['合同编号'])
            col_region = find_col(headers, ['区域'])
            col_prov = find_col(headers, ['省分', '省份', '省'])
            col_dept = find_col(headers, ['签定部门', '签订部门', '部门'])
            col_biz = find_col(headers, ['业务线'])
            col_ind = find_col(headers, ['签订行业', '行业'])
            col_cust = find_col(headers, ['客户标识', '客户编码', '客户编号', '客户名称', '责任人', '签定人'])
            col_status = find_col(headers, ['合同状态'])
            col_sign = find_col(headers, ['合同签定时间', '签订时间', '签约日期', '统计日期'])
            col_amt = find_col(headers, ['合同总金额', '合同金额', '合同额'])

            def gv(r, col):
                if not col or h_idx.get(col) is None: return None
                idx = h_idx[col]
                return r[idx] if idx < len(r) else None

            for r in rows[1:]:
                if r is None: continue
                cno = str(gv(r, col_no) or '').strip()
                if not cno or cno not in dims_map: continue
                region = str(gv(r, col_region) or '').strip()
                prov = str(gv(r, col_prov) or '').strip()
                dept = str(gv(r, col_dept) or '').strip()
                biz = str(gv(r, col_biz) or '').strip()
                ind = str(gv(r, col_ind) or '').strip()
                cust = str(gv(r, col_cust) or '').strip()
                status = str(gv(r, col_status) or '').strip()
                sign_dt = parse_date(gv(r, col_sign))
                amt = gv(r, col_amt)
                try: amt = float(amt) if amt not in (None, '', '-') else 0.0
                except: amt = 0.0
                dims_map[cno] = {
                    'region': region,
                    'province': prov,
                    'dept': dept,
                    'biz_line': biz,
                    'industry': ind,
                    'customer_key': _encode_customer_key(cust),
                    'contract_status': status,
                    'sign_year': str(sign_dt.year) if sign_dt else '',
                    'contract_amount': amt,
                }
        except Exception as e:
            print(f"[fund维度] 总合同表读取失败: {e}", flush=True)

    # ── 项目里程碑表：项目状态/账期/计划回款时间 ──
    m_fpath = _ds_latest_path('项目里程碑表')
    if m_fpath and os.path.exists(m_fpath):
        try:
            wb = openpyxl.load_workbook(m_fpath, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            headers = [str(h) if h else '' for h in rows[0]]
            h_idx = {h: i for i, h in enumerate(headers)}
            col_no = find_col(headers, ['合同编号'])
            col_pstatus = find_col(headers, ['项目状态'])
            col_term = find_col(headers, ['账期'])
            col_plan = find_col(headers, ['计划回款时间', '应收款应流入时间'])

            def gv(r, col):
                if not col or h_idx.get(col) is None: return None
                idx = h_idx[col]
                return r[idx] if idx < len(r) else None

            for r in rows[1:]:
                if r is None: continue
                cno = str(gv(r, col_no) or '').strip()
                if not cno or cno not in dims_map: continue
                pstatus = str(gv(r, col_pstatus) or '').strip()
                term = str(gv(r, col_term) or '').strip()
                plan_dt = parse_date(gv(r, col_plan))
                d = dims_map[cno]
                if pstatus: d['project_status'] = pstatus
                if term: d['payment_term'] = term
                if plan_dt: d['plan_recv_date'] = plan_dt.strftime('%Y-%m-%d')
        except Exception as e:
            print(f"[fund维度] 里程碑表读取失败: {e}", flush=True)

    return dims_map


def _fifo_occupy_upto(payments, collections, cutoff):
    """FIFO 口径：截至 cutoff 日期的资金占用（与 fund_analyze 主口径一致，CC-006 FR-13 同比用）"""
    if not payments:
        return 0
    first_pay = payments[0]['occur_date']
    if first_pay > cutoff:
        return 0  # cutoff 时点尚未发生付款，无占用
    # 预收款（日期 < 首付日）按付款顺序冲抵
    pre = sum(c['amount'] for c in collections if c['occur_date'] < first_pay)
    pool = []  # 每笔付款在 cutoff 前发生且未被预收款冲抵的剩余
    for p in payments:
        if p['occur_date'] > cutoff:
            continue
        remaining = p['amount']
        if pre > 0:
            off = min(pre, remaining)
            remaining -= off
            pre -= off
        if remaining > 0:
            pool.append(remaining)
    # 回款 FIFO（仅 cutoff 前且 >= 首付日的回款，同日先付后收由排序保证）
    for c in sorted(
        [c for c in collections if first_pay <= c['occur_date'] <= cutoff],
        key=lambda x: x['occur_date']):
        left = c['amount']
        while left > 0 and pool:
            if pool[0] <= left:
                left -= pool[0]
                pool.pop(0)
            else:
                pool[0] -= left
                left = 0
    return sum(pool)


@app.post("/api/fund/analyze")
def fund_analyze():
    """FIFO 先进先出资金占用计算"""
    import openpyxl
    from datetime import datetime, date
    from collections import defaultdict

    global fund_segments_cache, fund_flows_cache

    # 从数据源读「付款明细表」「收款明细表」最新版本（统一数据源），回退到 fund_data 旧目录
    def ds_latest_file(table_name):
        try:
            meta = _load_ds_meta()
            vers = meta.get(table_name, {}).get('versions', [])
            if vers:
                p = os.path.join(DATASOURCE_DIR, vers[0]['file'])
                if os.path.exists(p):
                    return p
        except Exception:
            pass
        return None

    pay_path = ds_latest_file('付款明细表') or os.path.join(FUND_DATA_DIR, 'payment_details.xlsx')
    coll_path = ds_latest_file('收款明细表') or os.path.join(FUND_DATA_DIR, 'collection_details.xlsx')

    if not os.path.exists(pay_path):
        return {'success': False, 'error': '请先上传付款明细表'}
    if not os.path.exists(coll_path):
        return {'success': False, 'error': '请先上传收款明细表'}

    # ── 工具函数 ──
    def safe_float(v):
        if v is None or v == '' or v == '-': return 0.0
        try: return float(v)
        except: return 0.0

    def parse_date(v):
        if v is None or v == '' or v == '-': return None
        if isinstance(v, (datetime, date)): return v
        s = str(v).strip()
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try: return datetime.strptime(s, fmt)
            except: pass
        return None

    def safe_str(v):
        if v is None: return ''
        return str(v).strip()

    def find_col(headers, keywords):
        for h in headers:
            hl = str(h).lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    def read_excel(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = [str(h) if h else '' for h in rows[0]]
        data = []
        for r in rows[1:]:
            row = {}
            for i, h in enumerate(headers):
                row[h] = r[i] if i < len(r) else None
            data.append(row)
        return headers, data

    # ── 系统参数 ──
    REPORT_CUTOFF = datetime(2026, 8, 12)
    ANNUAL_COST_RATE = 0.03

    # ── 读取数据并透视 ──
    pay_headers, pay_data = read_excel(pay_path)
    col_pay_no = find_col(pay_headers, ['合同编号', '编号'])
    col_pay_date = find_col(pay_headers, ['实际支付时间', '支付时间', '付款日期'])
    col_pay_amount = find_col(pay_headers, ['实际支付金额', '支付金额'])
    col_contract_amt = find_col(pay_headers, ['合同额', '合同金额'])

    coll_headers, coll_data = read_excel(coll_path)
    col_coll_no = find_col(coll_headers, ['合同号', '合同编号', '编号'])
    col_coll_date = find_col(coll_headers, ['回款日期', '收款日期', '回款时间'])
    col_coll_amount = find_col(coll_headers, ['到款金额', '回款金额', '收款金额'])

    # 合同基础信息（客户名/客户简称/项目名一律不采集，只保留合同编号等分析字段）
    contract_info = {}
    for row in pay_data:
        cno = safe_str(row.get(col_pay_no, ''))
        if not cno or cno in contract_info: continue
        contract_info[cno] = {
            '合同额': safe_float(row.get(col_contract_amt)) if col_contract_amt else 0,
        }

    # ── 透视聚合 ──
    pay_pivot = defaultdict(lambda: defaultdict(float))
    flow_idx = 0
    for row in pay_data:
        cno = safe_str(row.get(col_pay_no, ''))
        if not cno: continue
        dt = parse_date(row.get(col_pay_date))
        amt = safe_float(row.get(col_pay_amount))
        if dt and amt > 0:
            date_key = dt.strftime('%Y-%m-%d')
            pay_pivot[cno][date_key] += amt
            flow_idx += 1
        if cno not in contract_info:
            contract_info[cno] = {
                '合同额': safe_float(row.get(col_contract_amt)) if col_contract_amt else 0,
            }

    coll_pivot = defaultdict(lambda: defaultdict(float))
    for row in coll_data:
        cno = safe_str(row.get(col_coll_no, ''))
        if not cno: continue
        dt = parse_date(row.get(col_coll_date))
        amt = safe_float(row.get(col_coll_amount))
        if dt and amt > 0:
            date_key = dt.strftime('%Y-%m-%d')
            coll_pivot[cno][date_key] += amt
            flow_idx += 1
        if cno not in contract_info:
            contract_info[cno] = {
                '合同额': 0,
            }

    all_contracts = set(list(pay_pivot.keys()) + list(coll_pivot.keys()))

    # 维度关联：join 总合同表 + 项目里程碑表（CC-006 FR-8）
    contract_dims = _load_contract_dims(all_contracts)
    # 风险阈值配置（CC-006 FR-9）
    risk_cfg = _get_risk_config()

    # ── 清空全局缓存 ──
    fund_segments_cache = {}
    fund_flows_cache = {}

    # ── FIFO 计算 ──
    summary_rows = []
    # 全局逐笔现金流序列（跨合同合并，供同比分析 CC-006 FR-13）
    _global_cashflow = []
    # 去年同期（上年同日）FIFO 占用合计（CC-006 FR-13 同比）
    _grand_occupy_prev = 0

    for cno in all_contracts:
        # 构建付款列表
        payments = []
        p_idx = 0
        for date_key in sorted(pay_pivot.get(cno, {}).keys()):
            dt = parse_date(date_key)
            amt = pay_pivot[cno][date_key]
            if dt and amt > 0:
                p_idx += 1
                payments.append({
                    'flow_id': f'{cno}-PAY-{p_idx:03d}',
                    'occur_date': dt,
                    'amount': amt,
                })

        # 构建回款列表
        collections = []
        c_idx = 0
        for date_key in sorted(coll_pivot.get(cno, {}).keys()):
            dt = parse_date(date_key)
            amt = coll_pivot[cno][date_key]
            if dt and amt > 0:
                c_idx += 1
                collections.append({
                    'flow_id': f'{cno}-REC-{c_idx:03d}',
                    'occur_date': dt,
                    'amount': amt,
                })

        payments.sort(key=lambda x: x['occur_date'])
        collections.sort(key=lambda x: x['occur_date'])

        # 保存原始流水 + 现金流序列
        _pay_flows = [{'flow_id': p['flow_id'], 'flow_type': 'PAY', 'occur_date': p['occur_date'].strftime('%Y-%m-%d'), 'amount': p['amount']} for p in payments]
        _coll_flows = [{'flow_id': c['flow_id'], 'flow_type': 'RECEIVE', 'occur_date': c['occur_date'].strftime('%Y-%m-%d'), 'amount': c['amount']} for c in collections]

        # 现金流序列（逐笔：付款为负/回款为正，按日期排序，累计净现金余额）
        _events = []
        for p in payments:
            _events.append((p['occur_date'], 'PAY', -p['amount'], p['flow_id']))
        for c in collections:
            _events.append((c['occur_date'], 'RECEIVE', c['amount'], c['flow_id']))
        _events.sort(key=lambda x: (x[0], 1 if x[1] == 'RECEIVE' else 0))  # 同日先付后收
        _cashflow = []
        _balance = 0
        for _d, _t, _amt, _fid in _events:
            _balance += _amt
            _cashflow.append({
                'flow_id': _fid, 'date': _d.strftime('%Y-%m-%d'), 'type': _t,
                'amount': round(_amt), 'balance': round(_balance),
            })
        _global_cashflow.extend(_cashflow)  # 跨合同合并（CC-006 FR-13）
        # 按月汇总（每月付款/回款/净现金流/月末累计余额）
        _monthly = {}
        for _d, _t, _amt, _fid in _events:
            _mk = _d.strftime('%Y-%m')
            _m = _monthly.setdefault(_mk, {'month': _mk, 'pay_amount': 0, 'recv_amount': 0})
            if _t == 'PAY':
                _m['pay_amount'] += _amt
            else:
                _m['recv_amount'] += _amt
        _cashflow_monthly = []
        _mbalance = 0
        for _mk in sorted(_monthly.keys()):
            _m = _monthly[_mk]
            _mbalance += _m['recv_amount'] + _m['pay_amount']
            _cashflow_monthly.append({
                'month': _mk, 'pay_amount': round(_m['pay_amount']),
                'recv_amount': round(_m['recv_amount']),
                'net': round(_m['recv_amount'] + _m['pay_amount']), 'balance': round(_mbalance),
            })

        fund_flows_cache[cno] = {
            'payments': _pay_flows,
            'collections': _coll_flows,
            'cashflow': _cashflow,
            'cashflow_monthly': _cashflow_monthly,
        }

        # 无付款 → 纯收款合同，占用为0
        if not payments:
            total_receive = round(sum(c['amount'] for c in collections))
            info = contract_info.get(cno, {})
            dims = contract_dims.get(cno, {})
            fund_segments_cache[cno] = []
            summary_rows.append({
                '合同编号': cno,
                '合同额': round(info.get('合同额', 0)),
                '累计付款': 0,
                '累计收款': total_receive,
                '当前资金占用': 0,
                '元天合计': 0,
                '周期起始日': '-',
                '周期总天数': 0,
                '平均资金占用': 0,
                '预估资金成本': 0,
                '年化成本率': f'{ANNUAL_COST_RATE*100:.0f}%',
                '片段数': 0,
                '已结清片段': 0,
                '占用中片段': 0,
                '区域': dims.get('region', ''),
                '省份': dims.get('province', ''),
                '部门': dims.get('dept', ''),
                '业务线': dims.get('biz_line', ''),
                '行业': dims.get('industry', ''),
                '客户键': dims.get('customer_key', ''),
                '项目状态': dims.get('project_status', ''),
                '合同状态': dims.get('contract_status', ''),
                '签约年份': dims.get('sign_year', ''),
                '回款率': round(total_receive / info.get('合同额', 0), 4) if info.get('合同额', 0) else 0,
                '占用强度': 0,
                '风险等级': 'healthy',
            })
            continue

        # 第一步：预收款
        first_pay_date = payments[0]['occur_date']
        pre_receipts = [c for c in collections if c['occur_date'] < first_pay_date]
        regular_collections = [c for c in collections if c['occur_date'] >= first_pay_date]
        pre_receipt_balance = sum(c['amount'] for c in pre_receipts)

        # 第二步：垫资池（预收款冲抵时生成片段）
        advance_pool = []
        segments = []
        for p in payments:
            remaining = p['amount']
            if pre_receipt_balance > 0:
                offset = min(pre_receipt_balance, remaining)
                remaining -= offset
                pre_receipt_balance -= offset
                if offset > 0:
                    segments.append({
                        'contract_id': cno,
                        'origin_flow_id': p['flow_id'],
                        'pay_occur_date': p['occur_date'].strftime('%Y-%m-%d'),
                        'segment_status': 'PRESETTLED',
                        'segment_amount': round(offset),
                        'end_date': p['occur_date'].strftime('%Y-%m-%d'),
                        'occupy_days': 0,
                        'amount_day': 0,
                    })
            if remaining > 0:
                advance_pool.append({
                    'flow_id': p['flow_id'],
                    'pay_date': p['occur_date'],
                    'origin_amount': p['amount'],
                    'remain_amount': remaining,
                })

        # 第三步：回款 FIFO
        for c in regular_collections:
            receive_left = c['amount']
            receive_date = c['occur_date']
            while receive_left > 0 and advance_pool:
                item = advance_pool[0]
                days = (receive_date - item['pay_date']).days
                if days < 0: days = 0
                if item['remain_amount'] <= receive_left:
                    segments.append({
                        'contract_id': cno,
                        'origin_flow_id': item['flow_id'],
                        'pay_occur_date': item['pay_date'].strftime('%Y-%m-%d'),
                        'segment_status': 'SETTLED',
                        'segment_amount': round(item['remain_amount']),
                        'end_date': receive_date.strftime('%Y-%m-%d'),
                        'occupy_days': days,
                        'amount_day': round(item['remain_amount'] * days),
                    })
                    receive_left -= item['remain_amount']
                    advance_pool.pop(0)
                else:
                    segments.append({
                        'contract_id': cno,
                        'origin_flow_id': item['flow_id'],
                        'pay_occur_date': item['pay_date'].strftime('%Y-%m-%d'),
                        'segment_status': 'SETTLED',
                        'segment_amount': round(receive_left),
                        'end_date': receive_date.strftime('%Y-%m-%d'),
                        'occupy_days': days,
                        'amount_day': round(receive_left * days),
                    })
                    item['remain_amount'] -= receive_left
                    receive_left = 0

        # 第四步：剩余 → OCCUPYING
        for item in advance_pool:
            days = (REPORT_CUTOFF - item['pay_date']).days
            if days < 0: days = 0
            segments.append({
                'contract_id': cno,
                'origin_flow_id': item['flow_id'],
                'pay_occur_date': item['pay_date'].strftime('%Y-%m-%d'),
                'segment_status': 'OCCUPYING',
                'segment_amount': round(item['remain_amount']),
                'end_date': REPORT_CUTOFF.strftime('%Y-%m-%d'),
                'occupy_days': days,
                'amount_day': round(item['remain_amount'] * days),
            })

        fund_segments_cache[cno] = segments

        # 第五步：聚合
        total_pay = round(sum(p['amount'] for p in payments))
        total_receive = round(sum(c['amount'] for c in collections))
        current_occupy = round(sum(s['segment_amount'] for s in segments if s['segment_status'] == 'OCCUPYING'))
        # 去年同期（上年同日）FIFO 占用（CC-006 FR-13 同比；表格同比复用同一口径）
        prev_occupy = _fifo_occupy_upto(
            payments, collections,
            datetime(REPORT_CUTOFF.year - 1, REPORT_CUTOFF.month, REPORT_CUTOFF.day))
        _grand_occupy_prev += prev_occupy
        sum_amount_day = round(sum(s['amount_day'] for s in segments))
        cycle_start = payments[0]['occur_date']
        cycle_days = (REPORT_CUTOFF - cycle_start).days
        if cycle_days <= 0: cycle_days = 0
        avg_occupy = round(sum_amount_day / cycle_days) if cycle_days > 0 else 0
        estimate_cost = round(sum_amount_day * (ANNUAL_COST_RATE / 365))

        info = contract_info.get(cno, {})
        dims = contract_dims.get(cno, {})

        # 新指标：回款率 / 占用强度 / 占用天数 / 风险等级（CC-006 FR-8/FR-9）
        contract_amount = round(info.get('合同额', 0))
        if not contract_amount:
            contract_amount = dims.get('contract_amount', 0) or 0
        recv_rate = round(total_receive / contract_amount, 4) if contract_amount > 0 else 0
        occupy_intensity = round(current_occupy / contract_amount, 4) if contract_amount > 0 else (
            round(current_occupy / total_pay, 4) if total_pay > 0 else 0)
        # 占用天数：占用中片段的加权平均天数（元天/金额），无占用则为 0
        occupy_amount_total = sum(s['segment_amount'] for s in segments if s['segment_status'] == 'OCCUPYING')
        occupy_days = round(sum(s['amount_day'] for s in segments if s['segment_status'] == 'OCCUPYING') / occupy_amount_total) if occupy_amount_total > 0 else 0
        risk_level, suggestion = _calc_risk_level(occupy_days, recv_rate, occupy_intensity, current_occupy, risk_cfg)

        summary_rows.append({
            '合同编号': cno,
            '合同额': round(contract_amount),
            '累计付款': total_pay,
            '累计收款': total_receive,
            '净现金流': total_receive - total_pay,
            '当前资金占用': current_occupy,
            '上年同期占用': prev_occupy,
            '元天合计': sum_amount_day,
            '周期起始日': cycle_start.strftime('%Y-%m-%d'),
            '周期总天数': cycle_days,
            '平均资金占用': avg_occupy,
            '预估资金成本': estimate_cost,
            '年化成本率': f'{ANNUAL_COST_RATE*100:.0f}%',
            '片段数': len(segments),
            '已结清片段': sum(1 for s in segments if s['segment_status'] == 'SETTLED'),
            '占用中片段': sum(1 for s in segments if s['segment_status'] == 'OCCUPYING'),
            '区域': dims.get('region', ''),
            '省份': dims.get('province', ''),
            '部门': dims.get('dept', ''),
            '业务线': dims.get('biz_line', ''),
            '行业': dims.get('industry', ''),
            '客户键': dims.get('customer_key', ''),
            '项目状态': dims.get('project_status', ''),
            '合同状态': dims.get('contract_status', ''),
            '签约年份': dims.get('sign_year', ''),
            '回款率': recv_rate,
            '占用强度': occupy_intensity,
            '风险等级': risk_level,
            '风险建议': suggestion,
        })

    # 排序：纯付款 → 有付有收 → 纯收款，组内按资金占用降序
    def sort_key(row):
        pay = row['累计付款']
        recv = row['累计收款']
        if pay > 0 and recv == 0:
            cat = 0  # 纯付款
        elif pay > 0 and recv > 0:
            cat = 1  # 有付有收
        else:
            cat = 2  # 纯收款
        return (cat, -row['当前资金占用'])
    # 过滤纯收款合同（无付款、无占用）
    summary_rows = [r for r in summary_rows if r['累计付款'] > 0]
    summary_rows.sort(key=sort_key)

    n = len(summary_rows)
    grand_pay = sum(r['累计付款'] for r in summary_rows)
    grand_recv = sum(r['累计收款'] for r in summary_rows)
    grand_occupy = sum(r['当前资金占用'] for r in summary_rows)
    grand_amount_day = sum(r['元天合计'] for r in summary_rows)
    grand_cost = sum(r['预估资金成本'] for r in summary_rows)

    summary = {
        '合同总数': f'{n}个',
        '累计付款总额': f'¥{grand_pay:,}',
        '累计收款总额': f'¥{grand_recv:,}',
        '净现金流总额': f'¥{grand_recv - grand_pay:,}',
        '当前资金占用总额': f'¥{grand_occupy:,}',
        '总加权资金占用': f'{grand_amount_day:,}',
        '预估资金成本': f'¥{grand_cost:,}',
        '年化成本率': f'{ANNUAL_COST_RATE*100:.0f}%',
        '报表截止日': REPORT_CUTOFF.strftime('%Y-%m-%d'),
    }

    columns = ['合同编号', '累计付款', '累计收款', '净现金流',
               '当前资金占用', '平均资金占用', '预估资金成本', '周期总天数', '片段数']

    result = {
        'success': True,
        'message': f'分析完成：{n}个合同，当前资金占用 ¥{grand_occupy:,}',
        'data': {
            'summary': summary,
            'columns': columns,
            'rows': summary_rows,
            'flows': _global_cashflow,  # 逐笔现金流序列（CC-006 FR-13 同比分析）
            'yoy': {  # 同比辅助数据（CC-006 FR-13）
                'occupy_prev': round(_grand_occupy_prev),  # 上年同期日 FIFO 占用
            },
        }
    }
    # 写资金占用宽表（fund_metrics，每合同一行）
    try:
        conn_w = get_db()
        cw = conn_w.cursor()
        cw.execute("DELETE FROM fund_metrics")
        def _num(row, key):
            v = row.get(key, 0)
            if v is None or v == '-' or v == '': return 0
            if isinstance(v, (int, float)): return v
            try: return float(str(v).replace('¥', '').replace(',', '').replace('%', '').strip())
            except: return 0
        for row in summary_rows:
            cw.execute("""INSERT INTO fund_metrics
                (contract_no, customer_name, project_name, contract_amount, total_pay, total_recv,
                 current_occupy, prev_occupy, amount_day, cycle_start, cycle_days, avg_occupy, est_cost,
                 annual_rate, segment_count, settled_segments, occupying_segments,
                 region, province, dept, biz_line, industry, customer_key,
                 project_status, contract_status, sign_year, recv_rate, occupy_intensity, risk_level)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (row.get('合同编号', ''), '', '',
                 _num(row, '合同额'), _num(row, '累计付款'), _num(row, '累计收款'),
                 _num(row, '当前资金占用'), _num(row, '上年同期占用'), _num(row, '元天合计'), str(row.get('周期起始日', '')),
                 int(_num(row, '周期总天数')), _num(row, '平均资金占用'), _num(row, '预估资金成本'),
                 str(row.get('年化成本率', '')), int(_num(row, '片段数')),
                 int(_num(row, '已结清片段')), int(_num(row, '占用中片段')),
                 str(row.get('区域', '')), str(row.get('省份', '')), str(row.get('部门', '')),
                 str(row.get('业务线', '')), str(row.get('行业', '')), str(row.get('客户键', '')),
                 str(row.get('项目状态', '')), str(row.get('合同状态', '')), str(row.get('签约年份', '')),
                 _num(row, '回款率'), _num(row, '占用强度'), str(row.get('风险等级', 'healthy'))))
        conn_w.commit()
        conn_w.close()
    except Exception as e:
        print(f"[fund宽表] 写入失败: {e}")

    _save_snapshot('fund-occupancy', result)
    return result


@app.get("/api/fund/metrics")
def fund_metrics():
    """资金占用宽表查询：读 fund_metrics 宽表（ETL 结果），页面直接渲染"""
    conn = get_db()
    c = conn.cursor()
    rows = [dict(r) for r in c.execute("SELECT * FROM fund_metrics ORDER BY current_occupy DESC").fetchall()]
    conn.close()
    if not rows:
        return {'success': False, 'error': '资金占用宽表为空，请先执行「资金占用指标计算」定时任务'}

    n = len(rows)
    grand_pay = sum(r['total_pay'] for r in rows)
    grand_recv = sum(r['total_recv'] for r in rows)
    grand_occupy = sum(r['current_occupy'] for r in rows)
    grand_amount_day = sum(r['amount_day'] for r in rows)
    grand_cost = sum(r['est_cost'] for r in rows)
    annual_rate = rows[0]['annual_rate'] if rows else ''

    summary = {
        '合同总数': f'{n}个',
        '累计付款总额': f'¥{grand_pay:,.0f}',
        '累计收款总额': f'¥{grand_recv:,.0f}',
        '净现金流总额': f'¥{grand_recv - grand_pay:,.0f}',
        '当前资金占用总额': f'¥{grand_occupy:,.0f}',
        '总加权资金占用': f'{grand_amount_day:,.0f}',
        '预估资金成本': f'¥{grand_cost:,.0f}',
        '年化成本率': annual_rate,
    }

    # 每合同明细（列名对齐前端 renderFundResult；客户名/项目名一律不展示）
    detail_rows = [{
        '合同编号': r['contract_no'],
        '累计付款': r['total_pay'],
        '累计收款': r['total_recv'],
        '净现金流': r['total_recv'] - r['total_pay'],
        '当前资金占用': r['current_occupy'],
        '上年同期占用': r.get('prev_occupy') or 0,
        '平均资金占用': r['avg_occupy'],
        '预估资金成本': r['est_cost'],
        '周期总天数': r['cycle_days'],
        '周期起始日': r.get('cycle_start') or '',
        '片段数': r['segment_count'],
        '区域': r.get('region') or '',
        '省份': r.get('province') or '',
        '部门': r.get('dept') or '',
        '业务线': r.get('biz_line') or '',
        '客户键': r.get('customer_key') or '',
        '项目状态': r.get('project_status') or '',
        '回款率': r.get('recv_rate') or 0,
        '占用强度': r.get('occupy_intensity') or 0,
        '风险等级': r.get('risk_level') or 'healthy',
    } for r in rows]

    columns = ['合同编号', '累计付款', '累计收款', '净现金流',
               '当前资金占用', '平均资金占用', '预估资金成本', '周期总天数', '片段数']

    return {'success': True, 'data': {'summary': summary, 'columns': columns, 'rows': detail_rows}}


@app.get("/api/fund/segments/{contract_id}")
def fund_segments(contract_id: str):
    segments = fund_segments_cache.get(contract_id, [])
    flows = fund_flows_cache.get(contract_id, {'payments': [], 'collections': []})
    return {
        'success': True,
        'contract_id': contract_id,
        'segments': segments,
        'flows': flows,
        'cashflow': flows.get('cashflow', []),
        'cashflow_monthly': flows.get('cashflow_monthly', []),
        'local_summary': {
            'sum_amount_day': sum(s['amount_day'] for s in segments),
            'current_occupy': sum(s['segment_amount'] for s in segments if s['segment_status'] == 'OCCUPYING'),
            'total_segments': len(segments),
        }
    }


# ═══════════════════════════════════════════
# 资金多维度聚合 / 预警 API（CC-006 FR-8 ~ FR-12）
# ═══════════════════════════════════════════

# 维度列 → fund_metrics 列映射
DIM_COLUMN_MAP = {
    'region': 'region',
    'province': 'province',
    'dept': 'dept',
    'biz_line': 'biz_line',
    'industry': 'industry',
    'customer_key': 'customer_key',
    'project_status': 'project_status',
    'contract_status': 'contract_status',
    'sign_year': 'sign_year',
}
DIM_NAME_MAP = {
    'region': '区域', 'province': '省份', 'dept': '部门', 'biz_line': '业务线',
    'industry': '行业', 'customer_key': '客户集合', 'project_status': '项目状态',
    'contract_status': '合同状态', 'sign_year': '签约年份',
}
RISK_NAME_MAP = {
    'healthy': '健康', 'yellow': '关注', 'orange': '预警', 'red': '高危',
}


def _fund_rows_with_dims(where='', params=()):
    """读取 fund_metrics 宽表（含维度列），未分析时返回 None"""
    conn = get_db()
    c = conn.cursor()
    sql = "SELECT * FROM fund_metrics"
    if where:
        sql += f" WHERE {where}"
    sql += " ORDER BY current_occupy DESC"
    try:
        rows = [dict(r) for r in c.execute(sql, params).fetchall()]
    except Exception:
        rows = []
    conn.close()
    return rows or None


def _fund_dim_aggregate_inner(dim):
    """维度聚合计算（供 API 与测试复用）：返回 {rows, trend} 或抛错信息"""
    rows = _fund_rows_with_dims()
    if not rows:
        return {'error': '资金占用宽表为空，请先执行资金占用分析'}

    col = DIM_COLUMN_MAP[dim]
    agg = {}
    for r in rows:
        name = (r.get(col) or '').strip() or '未知'
        bucket = agg.setdefault(name, {
            'name': name,
            'contract_count': 0,
            'total_pay': 0, 'total_recv': 0,
            'current_occupy': 0, 'prev_occupy': 0,
            'avg_occupy_sum': 0, 'amount_day': 0, 'est_cost': 0,
            'contract_amount': 0,
            'risk_count': {'healthy': 0, 'yellow': 0, 'orange': 0, 'red': 0},
        })
        bucket['contract_count'] += 1
        bucket['total_pay'] += r.get('total_pay') or 0
        bucket['total_recv'] += r.get('total_recv') or 0
        bucket['current_occupy'] += r.get('current_occupy') or 0
        bucket['prev_occupy'] += r.get('prev_occupy') or 0
        bucket['avg_occupy_sum'] += r.get('avg_occupy') or 0
        bucket['amount_day'] += r.get('amount_day') or 0
        bucket['est_cost'] += r.get('est_cost') or 0
        bucket['contract_amount'] += r.get('contract_amount') or 0
        rl = r.get('risk_level') or 'healthy'
        bucket['risk_count'][rl if rl in bucket['risk_count'] else 'healthy'] += 1

    out = []
    for b in agg.values():
        n = b['contract_count']
        ca = b['contract_amount'] or b['total_pay'] or 0
        out.append({
            'name': b['name'],
            'contract_count': n,
            'total_pay': round(b['total_pay'], 2),
            'total_recv': round(b['total_recv'], 2),
            'net_cashflow': round(b['total_recv'] - b['total_pay'], 2),
            'current_occupy': round(b['current_occupy'], 2),
            'prev_occupy': round(b['prev_occupy'], 2),
            'avg_occupy': round(b['avg_occupy_sum'] / n, 2) if n else 0,
            'amount_day': round(b['amount_day'], 2),
            'est_cost': round(b['est_cost'], 2),
            'recv_rate': round(b['total_recv'] / ca, 4) if ca > 0 else 0,
            'occupy_intensity': round(b['current_occupy'] / ca, 4) if ca > 0 else 0,
            'risk_count': b['risk_count'],
            # 主风险等级：按 red > orange > yellow > healthy 取最大
            'risk_level': next((lv for lv in ['red', 'orange', 'yellow', 'healthy']
                                if b['risk_count'].get(lv, 0) > 0), 'healthy'),
        })
    out.sort(key=lambda x: x['current_occupy'], reverse=True)

    # 趋势数据：按客户键/区域分月的占用（供趋势预警折线）
    trend = None
    if dim in ('region', 'customer_key'):
        trend = _fund_monthly_trend(dim)
    return {'rows': out, 'trend': trend}


@app.get("/api/fund/dim/aggregate")
def fund_dim_aggregate(dim: str = Query('region'), month: str = Query('')):
    """维度聚合查询：GET /api/fund/dim/aggregate?dim=region&month=2026-07&level=1

    dim 取值见 DIM_COLUMN_MAP；month 可选（YYYY-MM，过滤签约/占用所属月暂按全量）；
    level 预留下钻层级（level=1 默认聚合；level=2 附加合同清单）。
    """
    month = (month or '').strip()
    if dim not in DIM_COLUMN_MAP:
        return {'success': False, 'error': f'不支持的维度: {dim}，可选: {",".join(DIM_COLUMN_MAP)}'}

    result = _fund_dim_aggregate_inner(dim)
    if 'error' in result:
        return {'success': False, 'error': result['error']}

    return {'success': True, 'dim': dim, 'dim_name': DIM_NAME_MAP.get(dim, dim),
            'rows': result['rows'], 'trend': result['trend']}


@app.get("/api/fund/dim/drill")
def fund_dim_drill(dim: str = Query('region'), value: str = Query('')):
    """穿透下钻：GET /api/fund/dim/drill?dim=region&value=华东 → 该维度下合同清单"""
    value = (value or '').strip()
    if dim not in DIM_COLUMN_MAP:
        return {'success': False, 'error': f'不支持的维度: {dim}'}
    if not value:
        return {'success': False, 'error': '缺少 value 参数'}

    rows = _fund_rows_with_dims()
    if not rows:
        return {'success': False, 'error': '资金占用宽表为空，请先执行资金占用分析'}

    col = DIM_COLUMN_MAP[dim]
    matched = [r for r in rows if (r.get(col) or '').strip() == value]
    if not matched:
        return {'success': True, 'dim': dim, 'value': value, 'rows': [], 'total_occupy': 0}

    out = [{
        '合同编号': r['contract_no'],
        '客户键': r.get('customer_key') or '',
        '区域': r.get('region') or '',
        '部门': r.get('dept') or '',
        '业务线': r.get('biz_line') or '',
        '项目状态': r.get('project_status') or '',
        '合同额': r.get('contract_amount') or 0,
        '累计付款': r.get('total_pay') or 0,
        '累计收款': r.get('total_recv') or 0,
        '当前资金占用': r.get('current_occupy') or 0,
        '上年同期占用': r.get('prev_occupy') or 0,
        '回款率': r.get('recv_rate') or 0,
        '占用强度': r.get('occupy_intensity') or 0,
        '风险等级': r.get('risk_level') or 'healthy',
    } for r in matched]
    out.sort(key=lambda x: x['当前资金占用'], reverse=True)
    return {'success': True, 'dim': dim, 'value': value,
            'rows': out, 'total_occupy': round(sum(x['当前资金占用'] for x in out), 2)}


@app.get("/api/fund/risk/config")
def fund_risk_config_get():
    """读取风险预警阈值配置"""
    _seed_risk_config()
    conn = get_db()
    c = conn.cursor()
    rows = [dict(r) for r in c.execute("SELECT * FROM risk_config ORDER BY key").fetchall()]
    conn.close()
    return {'success': True, 'config': rows}


@app.post("/api/fund/risk/config")
async def fund_risk_config_set(request: Request):
    """更新风险预警阈值配置，更新后重算风险等级"""
    try:
        payload = await request.json() or {}
    except Exception:
        payload = {}
    if not payload:
        return {'success': False, 'error': '缺少配置项'}

    allowed = {'days_green', 'days_yellow', 'days_orange', 'recv_rate',
               'intensity', 'amount_high', 'trend_months'}
    conn = get_db()
    c = conn.cursor()
    for k, v in payload.items():
        if k not in allowed:
            continue
        try:
            fv = float(v)
        except Exception:
            continue
        c.execute("INSERT INTO risk_config (key, value) VALUES (?,?) "
                  "ON CONFLICT(key) DO UPDATE SET value=excluded.value, "
                  "updated_at=datetime('now','localtime')", (k, fv))
    conn.commit()
    conn.close()

    # 重算风险等级：基于 fund_metrics 的占用天数近似（周期天数），重跑分析更准确
    try:
        fund_analyze()
        msg = '配置已更新，并已重算风险等级'
    except Exception as e:
        msg = f'配置已更新，但重算失败: {e}'
    return {'success': True, 'message': msg}


@app.get("/api/fund/risk/list")
def fund_risk_list(level: str = Query(''), dim: str = Query(''), dim_value: str = Query('')):
    """预警清单：GET /api/fund/risk/list?level=red|orange&dim=region&dim_value=华东"""
    level = (level or '').strip()
    dim = (dim or '').strip()
    dim_value = (dim_value or '').strip()

    rows = _fund_rows_with_dims()
    if not rows:
        return {'success': False, 'error': '资金占用宽表为空，请先执行资金占用分析'}

    def _ok(r):
        rl = r.get('risk_level') or 'healthy'
        if level and rl != level:
            return False
        if dim and dim in DIM_COLUMN_MAP:
            v = (r.get(DIM_COLUMN_MAP[dim]) or '').strip()
            if dim_value and v != dim_value:
                return False
            if not dim_value and v == '':
                return False
        return True

    matched = [r for r in rows if _ok(r)]
    matched.sort(key=lambda x: x.get('current_occupy') or 0, reverse=True)

    out = [{
        'contract_no': r['contract_no'],
        'customer_key': r.get('customer_key') or '',
        'region': r.get('region') or '',
        'dept': r.get('dept') or '',
        'biz_line': r.get('biz_line') or '',
        'project_status': r.get('project_status') or '',
        'contract_amount': r.get('contract_amount') or 0,
        'current_occupy': r.get('current_occupy') or 0,
        'recv_rate': r.get('recv_rate') or 0,
        'occupy_intensity': r.get('occupy_intensity') or 0,
        'risk_level': r.get('risk_level') or 'healthy',
        'risk_name': RISK_NAME_MAP.get(r.get('risk_level'), r.get('risk_level')),
        'suggestion': _risk_suggestion(r.get('risk_level') or 'healthy'),
    } for r in matched]

    # 分维度统计预警数量
    stat = {'red': 0, 'orange': 0, 'yellow': 0, 'healthy': 0}
    for r in rows:
        rl = r.get('risk_level') or 'healthy'
        stat[rl if rl in stat else 'healthy'] += 1

    return {'success': True, 'rows': out, 'count': len(out), 'stat': stat}


def _risk_suggestion(level):
    return {
        'healthy': '正常回款周期内',
        'yellow': '提醒跟进回款',
        'orange': '通知区域/部门负责人，安排对账',
        'red': '强制干预：催收/对账/风险上报',
    }.get(level, '')


def _fund_monthly_trend(dim):
    """按月聚合占用金额（用于趋势折线）：返回 [{dim_value, month, occupy}]"""
    rows = _fund_rows_with_dims()
    if not rows:
        return []
    from collections import defaultdict
    buckets = defaultdict(float)
    # 月度近似：用 cycle_start 月份（无则用当前月）
    from datetime import datetime
    for r in rows:
        key = (r.get(DIM_COLUMN_MAP[dim]) or '').strip() or '未知'
        cs = r.get('cycle_start') or ''
        month = cs[:7] if cs and len(cs) >= 7 else datetime.now().strftime('%Y-%m')
        buckets[(key, month)] += r.get('current_occupy') or 0
    return [{'dim_value': k, 'month': m, 'occupy': round(v, 2)}
            for (k, m), v in sorted(buckets.items())]


@app.get("/api/fund/risk/trend")
def fund_risk_trend(dim: str = Query('region')):
    """趋势预警：区域/客户集合维度占用金额环比连续上升 N 个月 → 预警"""
    cfg = _get_risk_config()
    trend_months = int(cfg.get('trend_months', 2))
    dim = (dim or 'region').strip()
    if dim not in ('region', 'customer_key'):
        return {'success': False, 'error': '趋势预警仅支持 dim=region 或 customer_key'}

    trend = _fund_monthly_trend(dim)
    from collections import defaultdict
    by_key = defaultdict(list)
    for t in trend:
        by_key[t['dim_value']].append({'month': t['month'], 'occupy': t['occupy']})
    for k in by_key:
        by_key[k].sort(key=lambda x: x['month'])

    warnings = []
    for key, series in by_key.items():
        ok, desc = _calc_trend_warning(series, trend_months)
        if ok:
            warnings.append({'dim': dim, 'dim_value': key, 'message': desc})
    warnings.sort(key=lambda x: x['dim_value'])
    return {'success': True, 'trend_months': trend_months, 'warnings': warnings,
            'trend': trend}


@app.get("/api/fund/dim/export")
def fund_dim_export(dim: str = Query('region')):
    """导出维度聚合数据为 Excel：GET /api/fund/dim/export?dim=region"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    result = _fund_dim_aggregate_inner(dim)
    if 'error' in result:
        return {'success': False, 'error': result['error']}
    rows = result['rows']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'资金{DIM_NAME_MAP.get(dim, dim)}维度聚合'
    ws.merge_cells('A1:K1')
    ws['A1'] = f'{DIM_NAME_MAP.get(dim, dim)}维度资金占用聚合（{datetime.now().strftime("%Y-%m-%d %H:%M")}）'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['维度', '合同数', '累计付款', '累计收款', '净现金流', '当前资金占用',
               '平均资金占用', '预估资金成本', '回款率', '占用强度', '主风险等级']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDEBF7')

    for ri, r in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=r['name'])
        ws.cell(row=ri, column=2, value=r['contract_count'])
        ws.cell(row=ri, column=3, value=r['total_pay'])
        ws.cell(row=ri, column=4, value=r['total_recv'])
        ws.cell(row=ri, column=5, value=r['net_cashflow'])
        ws.cell(row=ri, column=6, value=r['current_occupy'])
        ws.cell(row=ri, column=7, value=r['avg_occupy'])
        ws.cell(row=ri, column=8, value=r['est_cost'])
        ws.cell(row=ri, column=9, value=r['recv_rate'])
        ws.cell(row=ri, column=10, value=r['occupy_intensity'])
        ws.cell(row=ri, column=11, value=RISK_NAME_MAP.get(r['risk_level'], r['risk_level']))

    for col in 'ABCDEFGHIJK':
        ws.column_dimensions[col].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=fund_{dim}.xlsx'})


@app.get("/api/fund/risk/export")
def fund_risk_export(level: str = Query('')):
    """导出预警清单为 Excel：GET /api/fund/risk/export?level=red"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    rows = _fund_rows_with_dims()
    if not rows:
        return {'success': False, 'error': '资金占用宽表为空，请先执行资金占用分析'}
    if level:
        rows = [r for r in rows if (r.get('risk_level') or '') == level]
    out = [{
        'contract_no': r['contract_no'],
        'customer_key': r.get('customer_key') or '',
        'region': r.get('region') or '',
        'dept': r.get('dept') or '',
        'biz_line': r.get('biz_line') or '',
        'project_status': r.get('project_status') or '',
        'contract_amount': r.get('contract_amount') or 0,
        'current_occupy': r.get('current_occupy') or 0,
        'recv_rate': r.get('recv_rate') or 0,
        'occupy_intensity': r.get('occupy_intensity') or 0,
        'risk_level': r.get('risk_level') or 'healthy',
        'risk_name': RISK_NAME_MAP.get(r.get('risk_level'), r.get('risk_level')),
        'suggestion': _risk_suggestion(r.get('risk_level') or 'healthy'),
    } for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '资金风险预警清单'
    ws.merge_cells('A1:J1')
    ws['A1'] = f'资金风险预警清单（{datetime.now().strftime("%Y-%m-%d %H:%M")}）'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['合同编号', '客户键', '区域', '部门', '业务线', '项目状态',
               '合同额', '当前资金占用', '回款率', '占用强度', '风险等级', '干预建议']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FCE4D6')

    for ri, r in enumerate(out, start=3):
        ws.cell(row=ri, column=1, value=r['contract_no'])
        ws.cell(row=ri, column=2, value=r['customer_key'])
        ws.cell(row=ri, column=3, value=r['region'])
        ws.cell(row=ri, column=4, value=r['dept'])
        ws.cell(row=ri, column=5, value=r['biz_line'])
        ws.cell(row=ri, column=6, value=r['project_status'])
        ws.cell(row=ri, column=7, value=r['contract_amount'])
        ws.cell(row=ri, column=8, value=r['current_occupy'])
        ws.cell(row=ri, column=9, value=r['recv_rate'])
        ws.cell(row=ri, column=10, value=r['occupy_intensity'])
        ws.cell(row=ri, column=11, value=r['risk_name'])
        ws.cell(row=ri, column=12, value=r['suggestion'])

    for col in 'ABCDEFGHIJKL':
        ws.column_dimensions[col].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=fund_risk.xlsx'})


@app.get("/api/fund/analyze/export")
def fund_analyze_export():
    """导出资金占用分析结果"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    # 先执行分析
    result = fund_analyze()
    if not result['success']:
        return result

    data = result['data']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '资金占用分析'

    # 标题行
    ws.merge_cells('A1:E1')
    ws['A1'] = f'资金占用分析报告（{datetime.now().strftime("%Y-%m-%d %H:%M")}）'
    ws['A1'].font = Font(bold=True, size=14)

    # 汇总指标
    row_idx = 3
    for k, v in data['summary'].items():
        ws.cell(row=row_idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=v)
        row_idx += 1

    # 明细表
    row_idx += 1
    headers = data['columns']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='165DFF', end_color='165DFF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

    for row in data['rows']:
        for ci, h in enumerate(headers, 1):
            v = row.get(h, '')
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.alignment = Alignment(horizontal='center')
            if isinstance(v, (int, float)) and v < 0:
                cell.font = Font(color='FF0000')
        row_idx += 1

    # 列宽
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'资金占用分析_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    encoded = urllib.parse.quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'{encoded}'}
    )


# ═══════════════════════════════════════════
# 定时 ETL 框架（轨道A：固定指标计算链路）
# 原始明细表 → 定时任务聚合 → 指标汇总宽表
# ═══════════════════════════════════════════





def _ds_latest_path(table_name):
    """获取数据源某表最新版本的文件路径"""
    try:
        meta = _load_ds_meta()
        vers = meta.get(table_name, {}).get('versions', [])
        if vers:
            p = os.path.join(DATASOURCE_DIR, vers[0]['file'])
            if os.path.exists(p):
                return p
    except Exception:
        pass
    return None


def run_etl_gross_margin():
    """签单毛利 ETL：读总合同表 → 按年份/区域聚合 → 写指标宽表"""
    import openpyxl
    from datetime import datetime
    from collections import defaultdict

    h_fpath = _ds_latest_path('总合同表')
    if not h_fpath:
        return {'success': False, 'error': '总合同表未上传'}

    def safe_float(v):
        if v is None or v == '' or v == '-': return 0.0
        try: return float(v)
        except: return 0.0

    def find_col(headers, keywords):
        for h in headers:
            hl = str(h).lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    wb = openpyxl.load_workbook(h_fpath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) if h is not None else '' for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    col_amt = find_col(headers, ['合同总金额'])
    col_gross = find_col(headers, ['签单毛利'])
    col_region = find_col(headers, ['区域'])
    col_date = find_col(headers, ['统计日期'])
    col_dept = find_col(headers, ['业务类型'])

    def dept_cn(biz):
        """从业务类型映射部门（智能计算与集成事业部 ICID 下属）：
        系统集成业务→系统集成部、运维服务业务→运维服务部、资产运营业务→资产运营部、一体化运维→运维平台；
        其余业务类型不纳入部门维度"""
        biz = str(biz).strip()
        if not biz:
            return ''
        if '一体化运维' in biz:
            return '运维平台'
        if '运维服务' in biz:
            return '运维服务部'
        if '系统集成' in biz:
            return '系统集成部'
        if '资产运营' in biz:
            return '资产运营部'
        return ''

    def parse_year(v):
        if isinstance(v, datetime): return v.year
        if hasattr(v, 'year'): return v.year
        try: return int(str(v)[:4])
        except: return None

    year_agg = defaultdict(lambda: {'amt': 0.0, 'gross': 0.0})
    region_year_agg = defaultdict(lambda: defaultdict(lambda: {'amt': 0.0, 'gross': 0.0}))
    dept_year_agg = defaultdict(lambda: defaultdict(lambda: {'amt': 0.0, 'gross': 0.0}))
    dept_region_year_agg = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'amt': 0.0, 'gross': 0.0})))
    for r in rows[1:]:
        if r is None: continue
        amt = safe_float(r[col_idx[col_amt]]) if col_amt else 0.0
        gross = safe_float(r[col_idx[col_gross]]) if col_gross else 0.0
        region = str(r[col_idx[col_region]]).strip() if col_region and r[col_idx[col_region]] else ''
        dept = str(r[col_idx[col_dept]]).strip() if col_dept and r[col_idx[col_dept]] else ''
        dept = dept_cn(dept) if dept else ''
        y = parse_year(r[col_idx[col_date]]) if col_date else None
        if y is not None:
            year_agg[y]['amt'] += amt
            year_agg[y]['gross'] += gross
            if region:
                region_year_agg[region][y]['amt'] += amt
                region_year_agg[region][y]['gross'] += gross
            if dept:
                dept_year_agg[dept][y]['amt'] += amt
                dept_year_agg[dept][y]['gross'] += gross
            if region and dept:
                dept_region_year_agg[dept][region][y]['amt'] += amt
                dept_region_year_agg[dept][region][y]['gross'] += gross

    def rate(g, a): return round(g / a, 6) if a else 0.0

    conn = get_db()
    c = conn.cursor()
    # 兼容旧数据库：若 indicator_metrics 无 extra_json 列则添加
    cols = [row[1] for row in c.execute("PRAGMA table_info(indicator_metrics)").fetchall()]
    if 'extra_json' not in cols:
        c.execute("ALTER TABLE indicator_metrics ADD COLUMN extra_json TEXT DEFAULT '{}' ")
    c.execute("DELETE FROM indicator_metrics WHERE job_key='gross-margin'")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0
    for y, v in year_agg.items():
        c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, calc_time) VALUES (?,?,?,?,?,?,?,?,?)",
                  ('gross-margin', '签单毛利率', 'year', str(y), str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), now))
        n += 1
    for region, yd in region_year_agg.items():
        for y, v in yd.items():
            c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, calc_time) VALUES (?,?,?,?,?,?,?,?,?)",
                      ('gross-margin', '签单毛利率', 'region', region, str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), now))
            n += 1
    for dept, yd in dept_year_agg.items():
        for y, v in yd.items():
            c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, calc_time) VALUES (?,?,?,?,?,?,?,?,?)",
                      ('gross-margin', '签单毛利率', 'dept', dept, str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), now))
            n += 1
    for dept, rd in dept_region_year_agg.items():
        for region, yd in rd.items():
            for y, v in yd.items():
                extra = json.dumps({'dept': dept, 'region': region})
                c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, extra_json, calc_time) VALUES (?,?,?,?,?,?,?,?,?,?)",
                          ('gross-margin', '签单毛利率', 'dept_region', f"{dept}|{region}", str(y), v['amt'], v['gross'], rate(v['gross'], v['amt']), extra, now))
                n += 1
    conn.commit()
    conn.close()
    return {'success': True, 'rows': n}




def run_etl_fund_multidim():
    """资金占用多维度 ETL：读 fund_metrics 宽表 → 按维度×月份聚合 → 写 indicator_metrics

    产出 dim_type='fund_dim'，metric_name 形如 'fund_dim:region'。
    """
    from datetime import datetime
    from collections import defaultdict
    import json

    rows = _fund_rows_with_dims()
    if not rows:
        return {'success': False, 'error': 'fund_metrics 为空，请先执行「资金占用指标计算」'}

    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM indicator_metrics WHERE job_key='fund-multidim'")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0

    for dim in DIM_COLUMN_MAP:
        col = DIM_COLUMN_MAP[dim]
        agg = defaultdict(lambda: {
            'count': 0, 'pay': 0.0, 'recv': 0.0, 'occupy': 0.0, 'amt_day': 0.0,
            'cost': 0.0, 'ca': 0.0, 'risk': {'healthy': 0, 'yellow': 0, 'orange': 0, 'red': 0},
        })
        for r in rows:
            name = (r.get(col) or '').strip() or '未知'
            month = (r.get('cycle_start') or '')[:7] or '未知'
            b = agg[(name, month)]
            b['count'] += 1
            b['pay'] += r.get('total_pay') or 0
            b['recv'] += r.get('total_recv') or 0
            b['occupy'] += r.get('current_occupy') or 0
            b['amt_day'] += r.get('amount_day') or 0
            b['cost'] += r.get('est_cost') or 0
            b['ca'] += r.get('contract_amount') or 0
            rl = r.get('risk_level') or 'healthy'
            b['risk'][rl if rl in b['risk'] else 'healthy'] += 1

        for (name, month), b in agg.items():
            ca = b['ca'] or b['pay'] or 0
            extra = {
                'dim': dim, 'dim_name': DIM_NAME_MAP.get(dim, dim), 'month': month,
                'contract_count': b['count'],
                'total_pay': round(b['pay'], 2), 'total_recv': round(b['recv'], 2),
                'current_occupy': round(b['occupy'], 2), 'amount_day': round(b['amt_day'], 2),
                'est_cost': round(b['cost'], 2),
                'recv_rate': round(b['recv'] / ca, 4) if ca > 0 else 0,
                'occupy_intensity': round(b['occupy'] / ca, 4) if ca > 0 else 0,
                'risk_count': b['risk'],
                'risk_level': next((lv for lv in ['red', 'orange', 'yellow', 'healthy']
                                    if b['risk'].get(lv, 0) > 0), 'healthy'),
            }
            c.execute("INSERT INTO indicator_metrics (job_key, metric_name, dim_type, dim_value, year, contract_amt, gross_profit, gross_rate, extra_json, calc_time) VALUES (?,?,?,?,?,?,?,?,?,?)",
                      ('fund-multidim', f'fund_dim:{dim}', 'fund_dim', name, month,
                       b['count'], b['occupy'], extra['occupy_intensity'],
                       json.dumps(extra, ensure_ascii=False), now))
            n += 1

    conn.commit()
    conn.close()
    return {'success': True, 'rows': n}


def run_etl_payment_cycle():
    """回款周期 ETL：读总合同表 + 项目里程碑表 → 算每合同回款周期 → 写宽表"""
    import openpyxl
    from datetime import datetime, date

    h_fpath = _ds_latest_path('总合同表')
    r_fpath = _ds_latest_path('项目里程碑表')
    if not h_fpath:
        return {'success': False, 'error': '总合同表未上传'}

    def parse_date(v):
        if v is None or v == '' or v == '-': return None
        if isinstance(v, (datetime, date)): return v
        s = str(v).strip()
        if s.startswith('='): return None
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try: return datetime.strptime(s, fmt)
            except: pass
        return None

    def safe_float(v):
        if v is None or v == '' or v == '-': return 0.0
        try: return float(v)
        except: return 0.0

    def find_col(headers, keywords):
        for h in headers:
            hl = str(h).lower().replace(' ', '').replace('_', '').replace('-', '')
            for kw in keywords:
                if kw.lower().replace(' ', '').replace('_', '').replace('-', '') in hl:
                    return h
        return None

    # 读总合同表
    wb = openpyxl.load_workbook(h_fpath, data_only=True)
    ws = wb.active
    hrows = list(ws.iter_rows(values_only=True))
    wb.close()
    h_headers = [str(h) if h else '' for h in hrows[0]]
    h_col_no = find_col(h_headers, ['合同编号'])
    h_col_date = find_col(h_headers, ['统计日期', '签约日期'])
    h_col_amt = find_col(h_headers, ['合同总金额', '合同金额', '合同额'])
    h_idx = {h: i for i, h in enumerate(h_headers)}
    contracts = {}
    for r in hrows[1:]:
        if r is None: continue
        cno = str(r[h_idx[h_col_no]]).strip() if h_col_no and h_idx.get(h_col_no) is not None else ''
        if not cno: continue
        cdate = parse_date(r[h_idx[h_col_date]]) if h_col_date and h_idx.get(h_col_date) is not None else None
        amt = safe_float(r[h_idx[h_col_amt]]) if h_col_amt and h_idx.get(h_col_amt) is not None else 0.0
        contracts[cno] = {'date': cdate, 'amount': amt}

    # 读项目里程碑表，找每合同的最后一笔回款日期
    last_pay = {}
    if r_fpath:
        wb = openpyxl.load_workbook(r_fpath, data_only=True)
        ws = wb.active
        rrows = list(ws.iter_rows(values_only=True))
        wb.close()
        r_headers = [str(h) if h else '' for h in rrows[0]]
        r_col_no = find_col(r_headers, ['合同编号'])
        r_col_pay_date = find_col(r_headers, ['回款时间', '回款日期', '到款时间'])
        r_idx = {h: i for i, h in enumerate(r_headers)}
        for r in rrows[1:]:
            if r is None: continue
            cno = str(r[r_idx[r_col_no]]).strip() if r_col_no and r_idx.get(r_col_no) is not None else ''
            if not cno: continue
            pdate = parse_date(r[r_idx[r_col_pay_date]]) if r_col_pay_date and r_idx.get(r_col_pay_date) is not None else None
            if pdate and (cno not in last_pay or pdate > last_pay[cno]):
                last_pay[cno] = pdate

    # 算回款周期，写宽表
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM payment_cycle_metrics")
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n = 0
    for cno, info in contracts.items():
        cdate = info['date']
        pdate = last_pay.get(cno)
        cycle_days = 0
        if cdate and pdate:
            cycle_days = (pdate - cdate).days
            if cycle_days < 0: cycle_days = 0
        c.execute("INSERT INTO payment_cycle_metrics (contract_no, contract_date, last_payment_date, cycle_days, amount, calc_time) VALUES (?,?,?,?,?,?)",
                  (cno, cdate.strftime('%Y-%m-%d') if cdate else '', pdate.strftime('%Y-%m-%d') if pdate else '', cycle_days, info['amount'], now))
        n += 1
    conn.commit()
    conn.close()
    return {'success': True, 'rows': n}




@app.get("/api/etl/jobs")
def etl_jobs():
    """ETL 任务列表（供 9007 长期任务关联）"""
    conn = get_db()
    c = conn.cursor()
    jobs = [dict(r) for r in c.execute("SELECT * FROM etl_jobs ORDER BY id").fetchall()]
    conn.close()
    return {'jobs': jobs}


def _record_execution(job_key, result):
    """记录一次执行到 etl_executions 表"""
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ok = bool(result.get('success'))
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO etl_executions (job_key, run_time, status, detail, rows_written) VALUES (?,?,?,?,?)",
              (job_key, now, 'success' if ok else 'failed', result.get('error', ''), result.get('rows', 0)))
    c.execute("UPDATE etl_jobs SET last_run=?, last_result=? WHERE job_key=?",
              (now, '成功' if ok else result.get('error', '失败'), job_key))
    conn.commit()
    conn.close()


@app.post("/api/etl/run/{job_key}")
def etl_run(job_key: str):
    """手动触发 ETL 任务"""
    if job_key == 'gross-margin':
        result = run_etl_gross_margin()
    elif job_key == 'fund-occupancy':
        result = fund_analyze()
        if result.get('success'):
            result['rows'] = len(result.get('data', {}).get('rows', []))
    elif job_key == 'fund-multidim':
        result = run_etl_fund_multidim()
    elif job_key == 'payment-cycle':
        result = run_etl_payment_cycle()
    else:
        result = {'success': False, 'error': '该任务的计算逻辑尚未实现（骨架阶段）'}
    _record_execution(job_key, result)
    return result


@app.post("/api/etl/jobs/{job_key}/start")
def etl_start(job_key: str):
    """启动任务（进入自动调度）"""
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE etl_jobs SET status='running' WHERE job_key=?", (job_key,))
    conn.commit()
    conn.close()
    return {'success': True, 'job_key': job_key, 'status': 'running'}


@app.post("/api/etl/jobs/{job_key}/stop")
def etl_stop(job_key: str):
    """停止任务（退出自动调度）"""
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE etl_jobs SET status='stopped' WHERE job_key=?", (job_key,))
    conn.commit()
    conn.close()
    return {'success': True, 'job_key': job_key, 'status': 'stopped'}


@app.get("/api/etl/jobs/{job_key}")
def etl_job_detail(job_key: str):
    """任务详情（含计算逻辑 + 执行记录）"""
    conn = get_db()
    c = conn.cursor()
    job = c.execute("SELECT * FROM etl_jobs WHERE job_key=?", (job_key,)).fetchone()
    if not job:
        conn.close()
        return JSONResponse({'error': '任务不存在'}, status_code=404)
    job_dict = dict(job)
    exes = [dict(r) for r in c.execute("SELECT * FROM etl_executions WHERE job_key=? ORDER BY id DESC LIMIT 20", (job_key,)).fetchall()]
    conn.close()
    job_dict['executions'] = exes
    return job_dict


@app.get("/api/etl/metrics")
def etl_metrics(job_key: str = '', metric_name: str = '', dim_type: str = ''):
    """查询指标汇总宽表（指标数据集MCP 基础，只读）"""
    conn = get_db()
    c = conn.cursor()
    sql = "SELECT * FROM indicator_metrics WHERE 1=1"
    params = []
    if job_key:
        sql += " AND job_key=?"
        params.append(job_key)
    if metric_name:
        sql += " AND metric_name=?"
        params.append(metric_name)
    if dim_type:
        sql += " AND dim_type=?"
        params.append(dim_type)
    sql += " ORDER BY id"
    rows = [dict(r) for r in c.execute(sql, params).fetchall()]
    conn.close()
    return {'metrics': rows}


# ═══════════════════════════════════════════
# 原子本体 MCP（本体明细网关，只读）
# 对外暴露原始明细访问通道，不做大规模聚合
# ═══════════════════════════════════════════





# ═══════════════════════════════════════════
# 后台调度（轨道A 定时任务自动执行）
# ═══════════════════════════════════════════

import threading as _threading
import time as _time


def _cron_should_run(schedule, last_run, now):
    """简单 cron 匹配（仅支持「分 时」字段，每天执行一次）"""
    try:
        parts = schedule.split()
        if len(parts) < 2:
            return False
        minute = int(parts[0]); hour = int(parts[1])
    except Exception:
        return False
    today = now.strftime('%Y-%m-%d')
    if last_run and last_run.startswith(today):
        return False
    if (now.hour > hour) or (now.hour == hour and now.minute >= minute):
        return True
    return False


def _scheduler_loop():
    """后台调度：每 60 秒检查 running 任务，按 schedule 每天执行"""
    while True:
        try:
            from datetime import datetime
            now = datetime.now()
            conn = get_db()
            jobs = [dict(r) for r in conn.execute("SELECT * FROM etl_jobs WHERE status='running'").fetchall()]
            conn.close()
            for job in jobs:
                if _cron_should_run(job.get('schedule', ''), job.get('last_run', ''), now):
                    print(f"[ETL调度] 触发 {job['job_key']} @ {now.strftime('%Y-%m-%d %H:%M:%S')}", flush=True)
                    try:
                        if job['job_key'] == 'gross-margin':
                            result = run_etl_gross_margin()
                        elif job['job_key'] == 'fund-occupancy':
                            result = fund_analyze()
                            if result.get('success'):
                                result['rows'] = len(result.get('data', {}).get('rows', []))
                        elif job['job_key'] == 'payment-cycle':
                            result = run_etl_payment_cycle()
                        else:
                            result = {'success': False, 'error': '计算逻辑尚未实现（骨架阶段）'}
                    except Exception as e:
                        result = {'success': False, 'error': str(e)}
                    _record_execution(job['job_key'], result)
        except Exception as e:
            print(f"[ETL调度] 异常: {e}", flush=True)
        _time.sleep(60)


def _start_scheduler():
    t = _threading.Thread(target=_scheduler_loop, daemon=True, name='etl-scheduler')
    t.start()
    print("[ETL调度] 已启动，每 60 秒检查一次 running 任务", flush=True)


# 启动时注册 ETL 任务（幂等）
_register_etl_jobs()

# 启动时写入风险预警阈值默认值（幂等）
_seed_risk_config()

# 启动后台调度线程
_start_scheduler()


# ===================== 项目全生命周期管理（CC-010） =====================
# 注意：必须以命名空间方式引用。历史上 `from procurement_models import create_contract`
# 覆盖了 models.py 的同名函数，迫使 /api/contracts 用 `import contract_models` 兜底。
import plm_models as plm
from fastapi.responses import Response as _RawResponse


def _plm_op(payload=None):
    """取操作人（本期无鉴权，仅留痕）。"""
    if isinstance(payload, dict) and payload.get('operator'):
        return str(payload['operator'])[:64]
    return 'admin'


def _plm_ret(res):
    """统一响应包装：读接口 → {success, data}；写接口透出业务结果。"""
    if isinstance(res, dict) and 'success' in res:
        return res
    return {'success': True, 'data': res}


# ---- 页面 ----
@app.get("/plm")
def plm_page():
    return FileResponse(os.path.join(FRONTEND_DIR, 'plm.html'))


@app.get("/plm.app.js")
def plm_app_js():
    return FileResponse(os.path.join(FRONTEND_DIR, 'plm.app.js'))


# ---- 总览 / 配置 / 字典 / 日志 ----
@app.get("/api/plm/overview")
def api_plm_overview():
    return _plm_ret(plm.overview())


@app.get("/api/plm/dict")
def api_plm_dict(category: Optional[str] = None):
    return _plm_ret(plm.list_dict(category))


@app.post("/api/plm/dict")
def api_plm_dict_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_dict(payload.get('category', ''), payload.get('key', ''),
                                    payload.get('label', ''), payload.get('sort', 0),
                                    payload.get('remark', ''), _plm_op(payload)))


@app.delete("/api/plm/dict/{dict_id}")
def api_plm_dict_delete(dict_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_dict(dict_id, operator))


@app.get("/api/plm/config")
def api_plm_config():
    return _plm_ret(plm.list_config())


@app.put("/api/plm/config")
def api_plm_config_update(payload: Dict[str, Any]):
    key = payload.get('key')
    if not key:
        return {'success': False, 'error': 'key 必填'}
    return _plm_ret({'key': key,
                     'value': plm.set_config(key, payload.get('value', ''),
                                             payload.get('description', ''), _plm_op(payload))})


@app.get("/api/plm/logs")
def api_plm_logs(target_type: Optional[str] = None, target_id: Optional[str] = None,
                 limit: int = Query(200, le=1000)):
    return _plm_ret(plm.list_logs(target_type, target_id, limit))


# ---- 模块一：商机与投标概算 ----
@app.get("/api/plm/opportunities")
def api_plm_opp_list(keyword: Optional[str] = None, status: Optional[str] = None):
    return _plm_ret(plm.list_opportunities(keyword, status))


@app.post("/api/plm/opportunities")
def api_plm_opp_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_opportunity(payload, _plm_op(payload)))


@app.get("/api/plm/opportunities/{opp_id}")
def api_plm_opp_get(opp_id: int):
    r = plm.get_opportunity(opp_id)
    if not r:
        return JSONResponse({'success': False, 'error': '商机不存在'}, status_code=404)
    return _plm_ret(r)


@app.put("/api/plm/opportunities/{opp_id}")
def api_plm_opp_update(opp_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_opportunity(opp_id, payload, _plm_op(payload)))


@app.delete("/api/plm/opportunities/{opp_id}")
def api_plm_opp_delete(opp_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_opportunity(opp_id, operator))


@app.post("/api/plm/opportunities/{opp_id}/follow")
def api_plm_opp_follow(opp_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.add_follow_record(opp_id, payload.get('content', ''),
                                          _plm_op(payload), payload.get('time')))


@app.get("/api/plm/opportunities/{opp_id}/estimate")
def api_plm_opp_estimate(opp_id: int):
    return _plm_ret(plm.get_opportunity_estimate(opp_id))


@app.post("/api/plm/opportunities/{opp_id}/estimate")
def api_plm_opp_estimate_save(opp_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.save_opportunity_estimate(opp_id, payload, _plm_op(payload)))


@app.get("/api/plm/opportunities/{opp_id}/docs")
def api_plm_opp_docs(opp_id: int):
    return _plm_ret(plm.list_presale_docs(opp_id))


@app.post("/api/plm/opportunities/{opp_id}/docs")
def api_plm_opp_doc_create(opp_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p['opportunity_id'] = opp_id
    return _plm_ret(plm.create_presale_doc(p, _plm_op(payload)))


@app.delete("/api/plm/presale-docs/{doc_id}")
def api_plm_opp_doc_delete(doc_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_presale_doc(doc_id, operator))


# ---- 模块二：中标商机联动立项 ----
@app.post("/api/plm/opportunities/convert")
def api_plm_convert(payload: Dict[str, Any]):
    return _plm_ret(plm.convert_opportunity(payload, _plm_op(payload)))


# ---- 模块二：合同 ----
@app.get("/api/plm/contracts")
def api_plm_ct_list(keyword: Optional[str] = None):
    return _plm_ret(plm.list_contracts(keyword))


@app.post("/api/plm/contracts")
def api_plm_ct_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_contract(payload, _plm_op(payload)))


@app.get("/api/plm/contracts/{contract_id}")
def api_plm_ct_get(contract_id: int):
    r = plm.get_contract(contract_id)
    if not r:
        return JSONResponse({'success': False, 'error': '合同不存在'}, status_code=404)
    return _plm_ret(r)


@app.put("/api/plm/contracts/{contract_id}")
def api_plm_ct_update(contract_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_contract(contract_id, payload, _plm_op(payload)))


@app.delete("/api/plm/contracts/{contract_id}")
def api_plm_ct_delete(contract_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_contract(contract_id, operator))


# ---- 项目：列表 / 详情 / 全景 / 进度 / 财务 ----
@app.get("/api/plm/projects")
def api_plm_proj_list(keyword: Optional[str] = None, status: Optional[str] = None):
    return _plm_ret(plm.list_projects(keyword, status))


@app.post("/api/plm/projects")
def api_plm_proj_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_project(payload, _plm_op(payload)))


@app.get("/api/plm/projects/{project_id}")
def api_plm_proj_get(project_id: int):
    r = plm.get_project(project_id)
    if not r:
        return JSONResponse({'success': False, 'error': '项目不存在'}, status_code=404)
    return _plm_ret(r)


@app.put("/api/plm/projects/{project_id}")
def api_plm_proj_update(project_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_project(project_id, payload, _plm_op(payload)))


@app.delete("/api/plm/projects/{project_id}")
def api_plm_proj_delete(project_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_project(project_id, operator))


@app.get("/api/plm/projects/{project_id}/panorama")
def api_plm_proj_panorama(project_id: int):
    r = plm.project_panorama(project_id)
    if not r:
        return JSONResponse({'success': False, 'error': '项目不存在'}, status_code=404)
    return _plm_ret(r)


@app.get("/api/plm/projects/{project_id}/progress")
def api_plm_proj_progress(project_id: int):
    return _plm_ret(plm.project_progress(project_id))


@app.get("/api/plm/projects/{project_id}/finance")
def api_plm_proj_finance(project_id: int):
    return _plm_ret(plm.project_finance(project_id))


# ---- 模块二/三：四算基线 ----
@app.get("/api/plm/projects/{project_id}/baselines")
def api_plm_baseline_list(project_id: int):
    return _plm_ret(plm.list_baselines(project_id=project_id))


@app.get("/api/plm/projects/{project_id}/baseline-compare")
def api_plm_baseline_compare(project_id: int):
    return _plm_ret(plm.compare_baselines(project_id))


@app.post("/api/plm/projects/{project_id}/baselines")
def api_plm_baseline_save(project_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p.setdefault('project_id', project_id)
    p.setdefault('scope_type', 'project')
    p.setdefault('scope_id', project_id)
    return _plm_ret(plm.save_baseline(p, _plm_op(payload)))


@app.get("/api/plm/baselines/{baseline_id}")
def api_plm_baseline_get(baseline_id: int):
    r = plm.get_baseline(baseline_id)
    if not r:
        return JSONResponse({'success': False, 'error': '基线不存在'}, status_code=404)
    return _plm_ret(r)


@app.put("/api/plm/baselines/{baseline_id}")
def api_plm_baseline_update(baseline_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p['id'] = baseline_id
    return _plm_ret(plm.save_baseline(p, _plm_op(payload)))


@app.post("/api/plm/baselines/{baseline_id}/confirm")
def api_plm_baseline_confirm(baseline_id: int, payload: Dict[str, Any] = None):
    return _plm_ret(plm.confirm_baseline(baseline_id, _plm_op(payload)))


@app.post("/api/plm/baselines/{baseline_id}/lock")
def api_plm_baseline_lock(baseline_id: int, payload: Dict[str, Any] = None):
    return _plm_ret(plm.lock_baseline(baseline_id, _plm_op(payload)))


@app.delete("/api/plm/baselines/{baseline_id}")
def api_plm_baseline_delete(baseline_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_baseline(baseline_id, operator))


# ---- 模块三：里程碑与任务 ----
@app.get("/api/plm/projects/{project_id}/milestones")
def api_plm_ms_list(project_id: int):
    return _plm_ret(plm.list_milestones(project_id))


@app.post("/api/plm/projects/{project_id}/milestones")
def api_plm_ms_create(project_id: int, payload: Dict[str, Any]):
    p = dict(payload)
    p.setdefault('project_id', project_id)
    return _plm_ret(plm.create_milestone(p, _plm_op(payload)))


@app.put("/api/plm/milestones/{milestone_id}")
def api_plm_ms_update(milestone_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_milestone(milestone_id, payload, _plm_op(payload)))


@app.delete("/api/plm/milestones/{milestone_id}")
def api_plm_ms_delete(milestone_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_milestone(milestone_id, operator))


@app.get("/api/plm/projects/{project_id}/tasks")
def api_plm_task_list(project_id: int, milestone_id: Optional[int] = None):
    return _plm_ret(plm.list_tasks(project_id, milestone_id))


@app.post("/api/plm/tasks")
def api_plm_task_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_task(payload, _plm_op(payload)))


@app.get("/api/plm/tasks/{task_id}")
def api_plm_task_get(task_id: int):
    r = plm.get_task(task_id)
    if not r:
        return JSONResponse({'success': False, 'error': '任务不存在'}, status_code=404)
    return _plm_ret(r)


@app.put("/api/plm/tasks/{task_id}")
def api_plm_task_update(task_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_task(task_id, payload, _plm_op(payload)))


@app.delete("/api/plm/tasks/{task_id}")
def api_plm_task_delete(task_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_task(task_id, operator))


# ---- 模块四：人力池 / 分配 / 工时 ----
@app.get("/api/plm/staff")
def api_plm_staff_list(keyword: Optional[str] = None, status: Optional[str] = None):
    return _plm_ret(plm.list_staff(keyword, status))


# 必须先于 /staff/{staff_id} 注册，否则 'load' 会被当作整型 id 解析
@app.get("/api/plm/staff/load")
def api_plm_staff_load():
    return _plm_ret(plm.staff_load())


@app.post("/api/plm/staff")
def api_plm_staff_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_staff(payload, _plm_op(payload)))


@app.get("/api/plm/staff/{staff_id}")
def api_plm_staff_get(staff_id: int):
    r = plm.get_staff(staff_id)
    if not r:
        return JSONResponse({'success': False, 'error': '人员不存在'}, status_code=404)
    return _plm_ret(r)


@app.put("/api/plm/staff/{staff_id}")
def api_plm_staff_update(staff_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_staff(staff_id, payload, _plm_op(payload)))


@app.delete("/api/plm/staff/{staff_id}")
def api_plm_staff_delete(staff_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_staff(staff_id, operator))


@app.get("/api/plm/assignments")
def api_plm_asg_list(project_id: Optional[int] = None, staff_id: Optional[int] = None,
                     status: Optional[str] = None):
    return _plm_ret(plm.list_assignments(project_id, staff_id, status))


@app.post("/api/plm/assignments")
def api_plm_asg_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_assignment(payload, _plm_op(payload)))


@app.put("/api/plm/assignments/{assign_id}")
def api_plm_asg_update(assign_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_assignment(assign_id, payload, _plm_op(payload)))


@app.delete("/api/plm/assignments/{assign_id}")
def api_plm_asg_delete(assign_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_assignment(assign_id, operator))


@app.get("/api/plm/timesheets")
def api_plm_ts_list(project_id: Optional[int] = None, staff_id: Optional[int] = None):
    return _plm_ret(plm.list_timesheets(project_id, staff_id))


@app.post("/api/plm/timesheets")
def api_plm_ts_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_timesheet(payload, _plm_op(payload)))


@app.put("/api/plm/timesheets/{ts_id}")
def api_plm_ts_update(ts_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_timesheet(ts_id, payload, _plm_op(payload)))


@app.delete("/api/plm/timesheets/{ts_id}")
def api_plm_ts_delete(ts_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_timesheet(ts_id, operator))


@app.post("/api/plm/timesheets/sync")
def api_plm_ts_sync(payload: Dict[str, Any] = None):
    p = payload or {}
    return _plm_ret(plm.sync_labor_cost(p.get('project_id'), p.get('staff_id')))


# ---- 模块五：收支台账 ----
@app.get("/api/plm/ledger")
def api_plm_ledger_list(project_id: Optional[int] = None, kind: Optional[str] = None,
                        category: Optional[str] = None, source: Optional[str] = None):
    return _plm_ret(plm.list_ledger(project_id, kind, category, source))


@app.post("/api/plm/ledger")
def api_plm_ledger_create(payload: Dict[str, Any]):
    return _plm_ret(plm.create_ledger(payload, _plm_op(payload)))


@app.put("/api/plm/ledger/{ledger_id}")
def api_plm_ledger_update(ledger_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.update_ledger(ledger_id, payload, _plm_op(payload)))


@app.delete("/api/plm/ledger/{ledger_id}")
def api_plm_ledger_delete(ledger_id: int, operator: str = Query('admin')):
    return _plm_ret(plm.delete_ledger(ledger_id, operator))


# ---- 模块七：预警 ----
@app.get("/api/plm/alert-rules")
def api_plm_rule_list():
    return _plm_ret(plm.list_alert_rules())


@app.put("/api/plm/alert-rules/{rule_key}")
def api_plm_rule_update(rule_key: str, payload: Dict[str, Any]):
    return _plm_ret(plm.update_alert_rule(rule_key, payload, _plm_op(payload)))


@app.get("/api/plm/alerts")
def api_plm_alert_list(project_id: Optional[int] = None, dim: Optional[str] = None,
                       status: Optional[str] = None, level: Optional[str] = None):
    return _plm_ret(plm.list_alerts(project_id, dim, status, level))


@app.post("/api/plm/alerts/scan")
def api_plm_alert_scan(payload: Dict[str, Any] = None):
    return _plm_ret(plm.scan_alerts(_plm_op(payload)))


@app.put("/api/plm/alerts/{alert_id}/handle")
def api_plm_alert_handle(alert_id: int, payload: Dict[str, Any]):
    return _plm_ret(plm.handle_alert(alert_id, payload, _plm_op(payload)))


# ---- 模块八：报表导出 ----
@app.get("/api/plm/export/{report}")
def api_plm_export(report: str, project_id: Optional[int] = None):
    try:
        filename, data = plm.export_report(report, project_id)
    except ValueError as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=400)
    return _RawResponse(
        content=data,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': "attachment; filename*=UTF-8''%s"
                                        % urllib.parse.quote(filename)})


if __name__ == '__main__':
    import os
    import uvicorn
    port = int(os.environ.get('CC_PORT', '9006'))
    uvicorn.run(app, host='0.0.0.0', port=port)
