# -*- coding: utf-8 -*-
"""资金占用多维度分析与预警测试
# 规格编号: CC-006 资金占用分析（FR-8 维度关联/FR-9 风险预警/FR-12 趋势预警/FR-11 导出）
"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'backend'))

import openpyxl  # noqa: E402
import pytest  # noqa: E402

import main  # noqa: E402
from models import init_db  # noqa: E402


# ==================== 测试数据构造 ====================

def _make_pay_xlsx(path):
    """付款明细表：列名命中 find_col 匹配规则"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['合同编号', '实际支付时间', '实际支付金额', '合同额'])
    rows = [
        ('C1', '2025-01-01', 1000000, 1000000),
        ('C2', '2026-07-20', 500000, 1000000),
        ('C3', '2026-08-01', 0, 0),
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _make_coll_xlsx(path):
    """收款明细表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['合同号', '回款日期', '到款金额'])
    rows = [
        ('C1', '2025-03-01', 200000),
        ('C2', '2026-07-25', 100000),
        ('C3', '2026-08-01', 300000),
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _make_h_xlsx(path):
    """总合同表：维度字段（区域/省/部门/业务线/行业/客户标识/合同状态/签约时间/合同总额）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['合同编号', '区域', '省', '部门', '业务线', '行业', '客户标识',
               '合同状态', '合同签定时间', '合同总金额'])
    rows = [
        ('C1', '华东', '江苏', '政企一部', '政务', '政府', 'QDHEKJ', '执行中', '2024-12-01', 1000000),
        ('C2', '华东', '江苏', '政企二部', '教育', '教育', 'QDHEKJ', '执行中', '2026-05-01', 1000000),
        ('C3', '华北', '北京', '政企一部', '政务', '政府', 'ZGYZCXYH', '已完工', '2026-03-01', 0),
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _make_r_xlsx(path):
    """项目里程碑表：项目状态/账期/计划回款时间"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['合同编号', '项目状态', '账期', '计划回款时间'])
    rows = [
        ('C1', '实施中', '90天', '2025-03-01'),
        ('C2', '实施中', '30天', '2026-07-31'),
        ('C3', '已完工', '0天', '2026-07-01'),
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _write_datasource(tmp_path):
    """构造临时 datasource 目录（versions.json + 四张表）"""
    ds = tmp_path / 'datasource'
    ds.mkdir(exist_ok=True)
    _make_pay_xlsx(str(ds / 'pay_v1.xlsx'))
    _make_coll_xlsx(str(ds / 'coll_v1.xlsx'))
    _make_h_xlsx(str(ds / 'H_v1.xlsx'))
    _make_r_xlsx(str(ds / 'R_v1.xlsx'))
    meta = {
        '付款明细表': {'versions': [{'id': 1, 'file': 'pay_v1.xlsx'}], 'next_id': 2},
        '收款明细表': {'versions': [{'id': 1, 'file': 'coll_v1.xlsx'}], 'next_id': 2},
        '总合同表': {'versions': [{'id': 1, 'file': 'H_v1.xlsx'}], 'next_id': 2},
        '项目里程碑表': {'versions': [{'id': 1, 'file': 'R_v1.xlsx'}], 'next_id': 2},
    }
    (ds / 'versions.json').write_text(json.dumps(meta, ensure_ascii=False), encoding='utf-8')
    return str(ds)


def _patch_ds(monkeypatch, d):
    monkeypatch.setattr(main, 'DATASOURCE_DIR', d)
    monkeypatch.setattr(main, 'DS_META_FILE', os.path.join(d, 'versions.json'))


@pytest.fixture
def ds_dir(tmp_path, monkeypatch):
    d = _write_datasource(tmp_path)
    _patch_ds(monkeypatch, d)
    init_db()
    main._seed_risk_config()
    return d


# ==================== 用例 ====================

def test_fund_analyze_has_dims(ds_dir):
    """FR-8 维度关联：分析结果携带区域/部门/业务线/客户键/项目状态"""
    res = main.fund_analyze()
    assert res['success'] is True
    rows = res['data']['rows']
    by_no = {r['合同编号']: r for r in rows}

    c1 = by_no['C1']
    assert c1['区域'] == '华东'
    assert c1['省份'] == '江苏'
    assert c1['部门'] == '政企一部'
    assert c1['业务线'] == '政务'
    assert c1['行业'] == '政府'
    assert c1['客户键'] == 'QDHEKJ'
    assert c1['项目状态'] == '实施中'
    assert c1['签约年份'] == '2024'
    assert c1['合同状态'] == '执行中'

    # 纯收款合同（无付款）被过滤，不进入明细
    assert 'C3' not in by_no


def test_fund_analyze_risk_level(ds_dir):
    """FR-9 风险分级：占用>180天且强度>50%→高危；占用≤30天→健康"""
    res = main.fund_analyze()
    rows = res['data']['rows']
    by_no = {r['合同编号']: r for r in rows}

    # C1：付款100万(2025-01-01)，回款20万(2025-03-01) → 占用80万约529天，强度80%，回款率20% → 高危
    c1 = by_no['C1']
    assert c1['当前资金占用'] == 800000
    assert c1['回款率'] == pytest.approx(0.2, abs=0.01)
    assert c1['占用强度'] == pytest.approx(0.8, abs=0.01)
    assert c1['风险等级'] == 'red'

    # C2：付款50万(2026-07-20)，回款10万(2026-07-25) → 占用40万约23天 → 健康
    c2 = by_no['C2']
    assert c2['当前资金占用'] == 400000
    assert c2['风险等级'] == 'healthy'


def test_calc_risk_level_edges():
    """FR-9 风险分级纯函数边界：30/90/180 天、回款率、强度组合"""
    cfg = main._get_risk_config()
    assert main._calc_risk_level(30, 1.0, 0, 0, cfg)[0] == 'healthy'
    assert main._calc_risk_level(31, 1.0, 0, 0, cfg)[0] == 'yellow'
    assert main._calc_risk_level(90, 0.3, 0, 0, cfg)[0] == 'yellow'   # 边界：≤90 关注
    assert main._calc_risk_level(91, 0.3, 0, 0, cfg)[0] == 'orange'   # 90-180 且回款率<50%
    assert main._calc_risk_level(95, 0.8, 0, 0, cfg)[0] == 'yellow'   # 90-180 且回款率≥50%
    assert main._calc_risk_level(200, 0.5, 0.4, 0, cfg)[0] == 'orange'  # >180 且强度≤50%
    assert main._calc_risk_level(200, 0.5, 0.6, 0, cfg)[0] == 'red'    # >180 且强度>50%
    # 回款率为 0 且占用金额超阈值 → 高危
    assert main._calc_risk_level(10, 0, 0, 2000000, cfg)[0] == 'red'


def test_trend_warning(ds_dir):
    """FR-12 趋势预警：连续 2 个月上升 → 预警"""
    cfg = main._get_risk_config()
    months = int(cfg.get('trend_months', 2))
    ok, desc = main._calc_trend_warning(
        [{'month': '2026-05', 'occupy': 1000},
         {'month': '2026-06', 'occupy': 1200},
         {'month': '2026-07', 'occupy': 1400}], months)
    assert ok is True
    assert '连续' in desc

    ok2, _ = main._calc_trend_warning(
        [{'month': '2026-05', 'occupy': 1000},
         {'month': '2026-06', 'occupy': 900},
         {'month': '2026-07', 'occupy': 1400}], months)
    assert ok2 is False


def test_customer_key_encode():
    """FR-2 敏感列清洗：客户键确定性编码，不存真实名称"""
    assert main._encode_customer_key('QDHEKJ') == 'QDHEKJ'   # 已是编码原样返回
    enc = main._encode_customer_key('某集团有限公司')
    assert enc and enc != '某集团有限公司'
    assert len(enc) == 8
    assert main._encode_customer_key('某集团有限公司') == enc  # 确定性


def test_fund_dim_aggregate(ds_dir):
    """FR-8 维度聚合：按区域/客户键聚合，回款率与占用强度正确"""
    main.fund_analyze()
    # 直接查宽表聚合（不走 HTTP，验证计算逻辑）
    agg = main._fund_dim_aggregate_inner('region')
    rows = agg['rows']
    by_name = {r['name']: r for r in rows}
    east = by_name['华东']
    assert east['contract_count'] == 2
    assert east['current_occupy'] == 1200000  # C1 80万 + C2 40万
    assert east['total_recv'] == 300000
    assert east['recv_rate'] == pytest.approx(0.15, abs=0.01)  # 30万/200万
    assert east['risk_count']['red'] == 1
    assert east['risk_count']['healthy'] == 1

    cust = main._fund_dim_aggregate_inner('customer_key')
    cust_rows = {r['name']: r for r in cust['rows']}
    qd = cust_rows['QDHEKJ']
    assert qd['contract_count'] == 2
    assert qd['current_occupy'] == 1200000
    assert qd['risk_count']['red'] == 1


def test_fund_risk_list(ds_dir):
    """FR-9 预警清单：高危合同进入清单"""
    main.fund_analyze()
    rows = main._fund_rows_with_dims()
    assert rows is not None
    red = [r for r in rows if r.get('risk_level') == 'red']
    assert len(red) == 1
    assert red[0]['contract_no'] == 'C1'


def test_fund_metrics_columns(ds_dir):
    """FR-8 宽表持久化：维度列与新指标列入库"""
    main.fund_analyze()
    conn = main.get_db()
    c = conn.cursor()
    col_names = [r[1] for r in c.execute("PRAGMA table_info(fund_metrics)").fetchall()]
    conn.close()
    for col in ['region', 'province', 'dept', 'biz_line', 'industry', 'customer_key',
                'project_status', 'contract_status', 'sign_year',
                'recv_rate', 'occupy_intensity', 'risk_level']:
        assert col in col_names, f'missing column: {col}'
