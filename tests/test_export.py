# -*- coding: utf-8 -*-
"""整改报告导出测试（对应规格 CC-004：双 Sheet 结构 / 总览数据 / 异常明细着色）"""
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__)), 'backend'))

import excel_handler  # noqa: E402
import models  # noqa: E402
import openpyxl  # noqa: E402


@pytest.fixture()
def export_env(tmp_path, monkeypatch):
    """隔离测试环境：临时 DB + 临时上传目录"""
    db_file = tmp_path / 'test_export.db'
    upload_dir = tmp_path / 'uploads'
    upload_dir.mkdir()
    monkeypatch.setattr(models, 'DB_PATH', str(db_file))
    monkeypatch.setattr(models, 'DB_DIR', str(tmp_path))
    monkeypatch.setattr(excel_handler, 'UPLOAD_DIR', str(upload_dir))
    models.init_db()
    # 插入基础合同（versions/comparison_results 的外键依赖）
    conn = models.get_db()
    conn.execute("INSERT INTO contracts (contract_name) VALUES (?)", ('测试合同',))
    conn.commit()
    conn.close()
    return tmp_path


def _make_version(env, **overrides):
    conn = models.get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO versions (contract_id, supplier_name, total_items, matched_count,"
        " anomaly_count, pending_count, extra_count, progress)"
        " VALUES (?,?,?,?,?,?,?,?)",
        (1, '示例供应商',
         overrides.get('total_items', 4), overrides.get('matched_count', 2),
         overrides.get('anomaly_count', 1), overrides.get('pending_count', 1),
         overrides.get('extra_count', 0), overrides.get('progress', 50.0)),
    )
    version_id = c.lastrowid
    conn.commit()
    conn.close()
    return version_id


def _make_result(env, version_id, match_status, anomaly_types='[]', device_name='设备A'):
    conn = models.get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO comparison_results (contract_id, contract_item_id, supplier_item_id,"
        " match_status, anomaly_types, anomaly_detail, version_id)"
        " VALUES (?,?,?,?,?,?,?)",
        (1, None, None, match_status, anomaly_types, '', version_id),
    )
    conn.commit()
    conn.close()


def test_export_two_sheets(export_env):  # CC-004 FR-2
    """双 Sheet 结构：整改报告总览 + 异常明细，顺序固定"""
    version_id = _make_version(export_env)
    filepath = excel_handler.export_report(version_id)
    assert os.path.exists(filepath)
    wb = openpyxl.load_workbook(filepath)
    assert wb.sheetnames == ['整改报告总览', '异常明细']


def test_export_overview_data(export_env):  # CC-004 FR-3
    """总览页数据：版本号/条目数/各状态计数/整体进度"""
    version_id = _make_version(
        export_env, total_items=4, matched_count=2,
        anomaly_count=1, pending_count=1, progress=50.0,
    )
    filepath = excel_handler.export_report(version_id)
    wb = openpyxl.load_workbook(filepath)
    ws = wb['整改报告总览']
    values = {ws[f'A{i}'].value: ws[f'B{i}'].value for i in range(3, 11)}
    assert values['总条目数'] == 4
    assert values['成功匹配'] == 2
    assert values['匹配异常'] == 1
    assert values['待采购漏报'] == 1
    assert values['供应商增项'] == 0
    assert values['整体采购进度'] == '50.0%'


def test_export_anomaly_detail(export_env):  # CC-004 FR-4 / FR-5
    """异常明细：全部条目按状态着色，异常类型列可读"""
    version_id = _make_version(export_env)
    _make_result(export_env, version_id, '匹配异常', '["数量少报异常"]')
    _make_result(export_env, version_id, '待采购', '[]', device_name='设备B')
    filepath = excel_handler.export_report(version_id)
    wb = openpyxl.load_workbook(filepath)
    ws = wb['异常明细']
    types_col = [ws.cell(row=r, column=7).value for r in range(2, 4)]
    assert '数量少报异常' in types_col
    # 状态着色：匹配异常行字体为橙色系、待采购行字体为红色系
    color_map = {}
    for r in range(2, 4):
        cell = ws.cell(row=r, column=1)
        color_map[cell.font.color.rgb if cell.font.color else None] = cell.value
    assert len(color_map) >= 2  # 至少两种颜色区分不同状态
