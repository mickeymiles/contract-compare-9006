"""对 datasource/ 下的原始 xlsx 做脱敏，生成脱敏版覆盖原文件（先备份）

背景：
- 服务器分析接口（回款周期 / 毛利 / 资金占用等）直接读取 datasource/*.xlsx
- 原始 xlsx 含真实客户名、人员姓名、项目名，必须脱敏后才能部署
- 脱敏规则与既有体系一致：
    * 机构名（甲方名称/最终用户/合同甲方/客户名称）→ 拼音缩写
      （映射表 full+core → 补充词典 → pypinyin 兜底，保证无中文残留）
    * 人名（合同签定人/项目经理/原件归档人/里程碑达成审批人）→ 姓 + 叉叉
    * 项目名/描述 → 映射表子串替换 + 补充词典 + pypinyin 兜底
    * 金额/日期/周期等数值原样保留

用法：
    python3 scripts/sanitize_datasource.py            # 处理 datasource/*.xlsx
    python3 scripts/sanitize_datasource.py --dry-run  # 只统计不写文件
"""
import argparse
import json
import os
import re
import shutil
from datetime import datetime

import openpyxl

try:
    from pypinyin import lazy_pinyin
except ImportError:
    print("缺少 pypinyin，请先安装：pip3 install pypinyin")
    raise SystemExit(1)

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DS_DIR = os.path.join(BASE, "datasource")
MAP_PATH = os.path.join(BASE, "name_abbr_mapping.json")

# 项目名补充词典（与 export_seed.py 保持一致）
EXTRA_PROJECT_REPLACE = [
    ("东软教育健康科技", "DRJYKJ"),
    ("东软教育", "DRJY"),
    ("口腔医院", "KQYY"),
    ("华为", "HW"),
    ("深信服", "SXF"),
    ("青新网络", "QXWL"),
    ("国创", "GC"),
    ("摆渡平台", "BD平台"),
    ("西藏自治区", "XZZZQ"),
    ("涉藏地区", "SZDQ"),
    ("青海省", "QHS"),
    ("松江区", "SJQ"),
    ("杭州湾", "HZW"),
    ("黄岛", "HD"),
    ("青岛", "QD"),
    ("大连", "DL"),
    ("成都", "CD"),
    ("东北基地", "DBJD"),
]

# 机构名补充词典（整串/子串兜底，优先于 pypinyin 缩写）
EXTRA_ORG_REPLACE = EXTRA_PROJECT_REPLACE + [
    ("人民医院", "RMYY"),
    ("银行", "YH"),
    ("大学", "DX"),
    ("学院", "XY"),
    ("研究院", "YJY"),
    ("科技", "KJ"),
    ("信息", "XX"),
    ("技术", "JS"),
    ("软件", "RJ"),
    ("网络", "WL"),
    ("通信", "TX"),
    ("电力", "DL"),
    ("能源", "NY"),
    ("工程", "GC"),
    ("建设", "JS"),
    ("集团", "JT"),
    ("公司", "GS"),
]

# 需要按机构名脱敏的列（关键词模糊匹配）
ORG_COL_KEYWORDS = ["甲方名称", "最终用户", "合同甲方", "客户名称", "客户简称"]
# 需要按人名脱敏的列
PERSON_COL_KEYWORDS = ["合同签定人", "项目经理", "原件归档人", "里程碑达成审批人"]
# 需要按项目名脱敏的列
PROJECT_COL_KEYWORDS = ["项目描述", "项目名称"]

CN_RE = re.compile(r"[\u4e00-\u9fff]{2,}")


def load_mapping():
    with open(MAP_PATH, encoding="utf-8") as f:
        data = json.load(f)
    full = data.get("full", {})
    core = data.get("core", {})
    items = sorted(
        list(full.items()) + list(core.items()),
        key=lambda kv: len(kv[0]),
        reverse=True,
    )
    return full, items


FULL, MAP_ITEMS = load_mapping()


def cn_abbr(s):
    """连续中文字符串 -> 拼音首字母大写（如 广东省水文局 -> GDSWSJ）"""
    parts = [p[0].upper() for p in lazy_pinyin(s) if p]
    return "".join(parts)


def scrub_cn(out):
    """把字符串中剩余的中文（含单字）转拼音缩写，其余保留"""
    out = CN_RE.sub(lambda m: cn_abbr(m.group()), out)
    out = re.sub(r"[\u4e00-\u9fff]", lambda m: cn_abbr(m.group()), out)
    return out


def is_number(v):
    return isinstance(v, (int, float))


def sanitize_org(v):
    """机构名脱敏：整串映射 -> 子串映射 -> 补充词典 -> pypinyin 兜底"""
    if v is None or is_number(v):
        return v
    s = str(v).strip()
    if not s or s in ("无", "-"):
        return v
    if s in FULL:
        return FULL[s]
    out = s
    for src, dst in MAP_ITEMS:
        if src in out:
            out = out.replace(src, dst)
    for src, dst in EXTRA_ORG_REPLACE:
        if src in out:
            out = out.replace(src, dst)
    out = scrub_cn(out)
    return out


def sanitize_person(v):
    """人名脱敏：姓 + 叉叉（已脱敏的 X叉叉 保持，非中文保留）"""
    if v is None or is_number(v):
        return v
    s = str(v).strip()
    if not s or s in ("无", "-"):
        return v
    if re.fullmatch(r"[\u4e00-\u9fff]叉叉", s):
        return s
    if re.fullmatch(r"[\u4e00-\u9fff]{2,4}", s):
        return s[0] + "叉叉"
    # 含中文的混合串（如 "张三/李四"）：逐个人名处理
    if CN_RE.search(s):
        return CN_RE.sub(lambda m: m.group()[0] + "叉叉", s)
    return v


def sanitize_project(v):
    """项目名/描述脱敏：映射表子串 + 补充词典 + pypinyin 兜底"""
    if v is None or is_number(v):
        return v
    s = str(v).strip()
    if not s or s in ("无", "-"):
        return v
    out = s
    for src, dst in MAP_ITEMS:
        if src in out:
            out = out.replace(src, dst)
    for src, dst in EXTRA_PROJECT_REPLACE:
        if src in out:
            out = out.replace(src, dst)
    out = scrub_cn(out)
    return out


def col_category(header):
    h = str(header) if header else ""
    for kw in ORG_COL_KEYWORDS:
        if kw in h:
            return "org"
    for kw in PERSON_COL_KEYWORDS:
        if kw in h:
            return "person"
    for kw in PROJECT_COL_KEYWORDS:
        if kw in h:
            return "project"
    return None


def sanitize_value(v, cat):
    if cat == "org":
        return sanitize_org(v)
    if cat == "person":
        return sanitize_person(v)
    if cat == "project":
        return sanitize_project(v)
    return v


def process_workbook(path, stats, dry_run):
    wb = openpyxl.load_workbook(path)  # 保留格式
    for ws in wb.worksheets:
        # 表头在首行
        headers = [c.value for c in ws[1]]
        cats = [col_category(h) for h in headers]
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=2, max_row=max_row):
            for cell, cat in zip(row, cats):
                if cat is None:
                    continue
                old = cell.value
                new = sanitize_value(old, cat)
                if new != old:
                    stats["changed"] += 1
                    stats["by_col"][cat] = stats["by_col"].get(cat, 0) + 1
                    if old is not None and len(stats["samples"][cat]) < 5:
                        stats["samples"][cat].append(f"{old}  =>  {new}")
                    if not dry_run:
                        cell.value = new
    if not dry_run:
        wb.save(path)
    wb.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="只统计不写文件")
    ap.add_argument("--file", default=None, help="只处理指定文件（可选）")
    args = ap.parse_args()

    targets = []
    if args.file:
        targets.append(os.path.join(DS_DIR, args.file))
    else:
        for fn in sorted(os.listdir(DS_DIR)):
            if fn.endswith(".xlsx") and not fn.startswith("_"):
                targets.append(os.path.join(DS_DIR, fn))

    if not args.dry_run:
        bak_dir = os.path.join(DS_DIR, f"_orig_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        os.makedirs(bak_dir, exist_ok=True)
        for t in targets:
            shutil.copy2(t, os.path.join(bak_dir, os.path.basename(t)))
        print(f"已备份原始文件到: {bak_dir}")

    total_stats = {"changed": 0, "by_col": {}, "samples": {"org": [], "person": [], "project": []}}
    for t in targets:
        stats = {"changed": 0, "by_col": {}, "samples": {"org": [], "person": [], "project": []}}
        process_workbook(t, stats, args.dry_run)
        name = os.path.basename(t)
        print(f"\n== {name}  修改 {stats['changed']} 个单元格")
        for cat, n in sorted(stats["by_col"].items()):
            print(f"   [{cat}] {n} 处")
        for cat, samples in stats["samples"].items():
            for s in samples:
                print(f"     {s}")
        total_stats["changed"] += stats["changed"]
        for k, v in stats["by_col"].items():
            total_stats["by_col"][k] = total_stats["by_col"].get(k, 0) + v

    print(f"\n总计修改 {total_stats['changed']} 个单元格")
    if args.dry_run:
        print("（--dry-run 模式，未写文件）")
    else:
        print("已生成脱敏版 xlsx（覆盖原文件），原始文件保留在 _orig_backup_* 目录")


if __name__ == "__main__":
    main()
